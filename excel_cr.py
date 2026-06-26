"""
excel_cr.py · W26+ · Excel CheckRates — estructura simplificada
5 hojas: Severity | Maestra | Críticos | Bajo Rendimiento | Sin Conversión

Severity: cada sección es una Tabla Excel (filtro propio independiente).
Maestra: una fila por Hotel × Canasta — Banda Ef + Banda CV
Bandas: Banda Ef Global + Banda CV Global coloreadas; exposición cruzada en texto plano.
"""
import pickle, os, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_eficacia, banda_convrate, get_dest_tier

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_CR', f'cr_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D = pickle.load(f)

CANASTA = D['CANASTA']
TAB_EF  = D['TAB_EF_BY_CANASTA']
TAB_CV  = D['TAB_CV_BY_CANASTA']
M       = D['M']
p80_all = D['p80_hotel'].copy()
hcm     = D.get('hotel_channel_map', {})

try:
    import pickle as _pkl
    _bk_path = os.getenv('PICKLE_BK', f'bk_w{VOL_NUM}_data.pkl')
    with open(_bk_path, 'rb') as _f:
        _BK = _pkl.load(_f)
    BK_HOTEL = _BK.get('BK_HOTEL', {})
except Exception:
    BK_HOTEL = {}

def clean(n):
    return re.sub(r'^\(\d+\)\s*-\s*', '', str(n)).strip() if n else str(n)

hcm_clean = {clean(k): v for k, v in hcm.items()}
p80_all['Hotel']   = p80_all['Hotel'].apply(clean)
p80_all['Channel'] = p80_all['Hotel'].map(hcm_clean).fillna('—')

CR = '5C469C'

def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP   = fill('EAF3DE'); WOW_UP_F   = fnt('2F6C34', bold=True)
WOW_DN   = fill('FCE8E6'); WOW_DN_F   = fnt('C0392B', bold=True)
WOW_NEU  = fill('F2EEE6'); WOW_NEU_F  = fnt('8A8377', bold=True)

RANGOS_EF = {
    'Exitosa':       '≥ 97%',
    'Aceptable':     '93% – 97%',
    'Revisar':       '85% – 93%',
    'Crítica':       '60% – 85%',
    'Súper Crítica': '< 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0',
    'Crítica':        '< 0,8%',
    'Revisar':        '0,8% – 1,5%',
    'Aceptable':      '1,5% – 2,5%',
    'Exitosa':        '≥ 2,5%',
}

_table_counter = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    return round(float(v), 4)

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(CR, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(CR, 11, True)

def mk_hdr_plain(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(CR)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_excel_table(ws, hdr_row, data_row_end, n_cols, name_prefix):
    _table_counter[0] += 1
    tname = f'T_{name_prefix}_{_table_counter[0]}'
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{data_row_end}'
    t = Table(displayName=tname, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center', fmt=None):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if fmt: cell.number_format = fmt
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

# ── Fuentes de datos ──────────────────────────────────────────────────────────
CANASTAS_DEF = [
    ('global', 'Global',      None,     None),
    ('b2c',    'B2C',         'B2C',    'B2C'),
    ('op',     'Opaco',       'B2B-OP', 'B2B-OP'),
    ('cug',    'Ultra Opaco', 'CUG',    'CUG'),
]

def hotel_src_cr(can_dist):
    if can_dist is None:
        df = p80_all.copy()
    elif 'DistributionCategory' in p80_all.columns:
        df = p80_all[p80_all['DistributionCategory'] == can_dist].copy()
    else:
        can = CANASTA.get(can_dist, {})
        src = can.get('p80_hotel')
        if src is None: src = can.get('p80')
        df = (src if src is not None else p80_all).copy()
    if 'Channel' not in df.columns and 'Hotel' in df.columns:
        df['Hotel_c'] = df['Hotel'].apply(clean)
        df['Channel'] = df['Hotel_c'].map(hcm_clean).fillna('—')
    return df

def get_bk_hotel(hotel_name, can_label):
    h = BK_HOTEL.get(hotel_name, {})
    return h.get(can_label, h.get('Global', None)) if h else None

def build_banda_lookup_cr():
    lookup = {}
    for _, can_label, _, can_dist in CANASTAS_DEF:
        if can_dist is None: continue
        df = hotel_src_cr(can_dist)
        if df is None or len(df) == 0: continue
        for _, row in df.iterrows():
            h  = clean(str(row.get('Hotel', '')))
            ef = fmt_pct(row.get('Eficacia'))
            cv = fmt_pct(row.get('ConvRate'))
            bk = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            b_ef = banda_eficacia(ef) if ef is not None else '—'
            b_cv = banda_convrate(cv, bk) if cv is not None else '—'
            if h not in lookup: lookup[h] = {}
            lookup[h][can_label] = {'ef': b_ef, 'cv': b_cv}
    return lookup

def band_split_ef(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['Eficacia'].apply(
        lambda v: banda_eficacia(round(float(v), 4)) if pd.notna(v) else '—'
    )
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Hoja Severity ─────────────────────────────────────────────────────────────
def write_severity(ws):
    title(ws, f'Severity CR W{VOL_NUM}', f'Resumen ejecutivo · {PERIODO}')
    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))

    ef_curr = m_curr.get('eficacia', 0); ef_prev = m_prev.get('eficacia', 0)
    cv_curr = m_curr.get('conv_rate', 0); cv_prev = m_prev.get('conv_rate', 0)
    ef_wow  = (ef_curr - ef_prev) * 100 if ef_prev else None
    cv_wow  = (cv_curr - cv_prev) * 100 if cv_prev else None

    r = 4
    section_title(ws, r, 'KPI Global'); r += 1
    r = mk_hdr_plain(ws, r, ['Métrica', f'W{int(VOL_NUM)-1}', f'W{VOL_NUM}', 'WoW'])
    for label, prev, curr, wow in [
        ('Eficacia',  ef_prev, ef_curr, ef_wow),
        ('Conv Rate', cv_prev, cv_curr, cv_wow),
    ]:
        ws.cell(r, 1, label).font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
        ws.cell(r, 2, round(prev, 4) if prev else None).border = BD
        ws.cell(r, 3, round(curr, 4) if curr else None).border = BD
        ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
        ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        apply_wow(ws, r, 4, wow, invert=False)
        r += 1
    r += 1

    sev_ef, sev_cv = {}, {}
    for _, row in p80_all.iterrows():
        ef = fmt_pct(row.get('Eficacia'))
        cv = fmt_pct(row.get('ConvRate'))
        bk = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        if ef is not None: b = banda_eficacia(ef); sev_ef[b] = sev_ef.get(b, 0) + 1
        if cv is not None: b = banda_convrate(cv, bk); sev_cv[b] = sev_cv.get(b, 0) + 1
    total_ef = sum(sev_ef.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    # Severity Eficacia — Tabla Excel
    section_title(ws, r, 'Severity Eficacia · Hoteles Global'); r += 1
    hdr_r = r; r = mk_hdr_plain(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_ef.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True); ws.cell(r, 1).border = BD
        mk_cell(ws, r, 2, RANGOS_EF.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_ef, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_excel_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevEf')
    r += 1

    # Severity Conv Rate — Tabla Excel
    section_title(ws, r, 'Severity Conv Rate · Hoteles Global'); r += 1
    hdr_r = r; r = mk_hdr_plain(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True); ws.cell(r, 1).border = BD
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_excel_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    tab_ef_g = TAB_EF.get('global', {})
    tab_cv_g = TAB_CV.get('global', {})
    DIM_COLS = ['Nombre', 'Banda Ef', 'Banda CV', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']

    def write_dim_table(df_ef, df_cv, key_col, label, n=30):
        nonlocal r
        if df_ef is None or len(df_ef) == 0: return
        section_title(ws, r, label); r += 1
        hdr_r = r; r = mk_hdr_plain(ws, r, DIM_COLS)
        df_m = df_ef.sort_values('Eficacia', ascending=True).head(n).copy()
        if df_cv is not None and 'ConvRate_WoW_pp' in df_cv.columns and key_col in df_cv.columns:
            df_m = df_m.merge(df_cv[[key_col, 'ConvRate', 'ConvRate_WoW_pp']], on=key_col, how='left', suffixes=('', '_cv'))
        for _, row in df_m.iterrows():
            ef  = fmt_pct(row.get('Eficacia'))
            cv  = fmt_pct(row.get('ConvRate'))
            bk  = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef = banda_eficacia(ef) if ef is not None else '—'
            bcv = banda_convrate(cv, bk) if cv is not None else '—'
            mk_cell(ws, r, 1, str(row.get(key_col, '—')), align='left')
            mk_cell(ws, r, 2, bef, bef, is_sev=True)
            mk_cell(ws, r, 3, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 4, cru)
            pct_cell(ws, r, 5, ef)
            apply_wow(ws, r, 6, sf(row.get('Eficacia_WoW_pp')), invert=False)
            pct_cell(ws, r, 7, cv)
            apply_wow(ws, r, 8, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 9, bk)
            r += 1
        add_excel_table(ws, hdr_r, r-1, len(DIM_COLS), label[:4].replace(' ','_'))
        r += 1

    write_dim_table(tab_ef_g.get('destino'), tab_cv_g.get('destino'), 'Destino',  'Top Destinos · Eficacia ASC')
    write_dim_table(tab_ef_g.get('corp'),    tab_cv_g.get('corp'),    'CorpName', 'Top Corporativos · Eficacia ASC')

    # Channel — Tabla Excel
    df_ch_ef = tab_ef_g.get('channel')
    df_ch_cv = tab_cv_g.get('channel')
    if df_ch_ef is not None and len(df_ch_ef) > 0:
        CH_COLS = ['Channel', 'Banda Ef', 'Banda CV', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']
        section_title(ws, r, 'Channel · Eficacia ASC'); r += 1
        hdr_r = r; r = mk_hdr_plain(ws, r, CH_COLS)
        df_ch_m = df_ch_ef.sort_values('Eficacia', ascending=True).copy()
        if df_ch_cv is not None and 'ConvRate_WoW_pp' in df_ch_cv.columns:
            df_ch_m = df_ch_m.merge(df_ch_cv[['ExternalProviderName', 'ConvRate_WoW_pp']], on='ExternalProviderName', how='left')
        for _, row in df_ch_m.iterrows():
            ef  = fmt_pct(row.get('Eficacia'))
            cv  = fmt_pct(row.get('ConvRate'))
            bk  = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef = banda_eficacia(ef) if ef is not None else '—'
            bcv = banda_convrate(cv, bk) if cv is not None else '—'
            mk_cell(ws, r, 1, str(row.get('ExternalProviderName', '—')), align='left')
            mk_cell(ws, r, 2, bef, bef, is_sev=True)
            mk_cell(ws, r, 3, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 4, cru)
            pct_cell(ws, r, 5, ef)
            apply_wow(ws, r, 6, sf(row.get('Eficacia_WoW_pp')), invert=False)
            pct_cell(ws, r, 7, cv)
            apply_wow(ws, r, 8, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 9, bk)
            r += 1
        add_excel_table(ws, hdr_r, r-1, len(CH_COLS), 'Chan')

    autofit(ws, [28, 16, 16, 10, 10, 10, 10, 10, 10])

# ── Hoja Maestra ──────────────────────────────────────────────────────────────
MAESTRA_COLS = [
    'Destino', 'Clasificación Destino', 'Corporativo', 'Hotel', 'Channel', 'Canasta',
    'CR Únicos', 'WoW CR', 'Eficacia', 'WoW Ef', 'Banda Ef',
    'Conv Rate', 'WoW CV', 'Banda CV', 'Bookability', 'Bookings'
]

def write_maestra(ws):
    title(ws, f'Maestra CR · CheckRates W{VOL_NUM}',
          f'Una fila por Hotel × Canasta · {PERIODO} · Filtrar por cualquier columna')
    hdr_r = 4
    r = mk_hdr_plain(ws, hdr_r, MAESTRA_COLS)

    for _, can_label, _, can_dist in CANASTAS_DEF:
        df = hotel_src_cr(can_dist)
        if df is None or len(df) == 0: continue
        for _, row in df.iterrows():
            ef   = fmt_pct(row.get('Eficacia'))
            cv   = fmt_pct(row.get('ConvRate'))
            bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef  = banda_eficacia(ef) if ef is not None else '—'
            bcv  = banda_convrate(cv, bk) if cv is not None else '—'
            hotel   = clean(str(row.get('Hotel', '—')))
            channel = str(row.get('Channel', '—'))
            destino = str(row.get('Destino', '—'))
            corp    = str(row.get('CorpName', row.get('Corp', '—')))
            tier    = get_dest_tier(destino).capitalize()
            bk_pct  = get_bk_hotel(hotel, can_label)
            for ci, val in enumerate([destino, tier, corp, hotel, channel, can_label], 1):
                mk_cell(ws, r, ci, val, align='left')
            mk_cell(ws, r, 7, cru)
            apply_wow(ws, r, 8, sf(row.get('CR_WoW_pct')), invert=False)
            pct_cell(ws, r, 9, ef)
            apply_wow(ws, r, 10, sf(row.get('Eficacia_WoW_pp')), invert=False)
            mk_cell(ws, r, 11, bef, bef, is_sev=True)
            pct_cell(ws, r, 12, cv)
            apply_wow(ws, r, 13, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 14, bcv, bcv, is_sev=True)
            pct_cell(ws, r, 15, bk_pct)
            mk_cell(ws, r, 16, bk)
            r += 1

    add_excel_table(ws, hdr_r, r-1, len(MAESTRA_COLS), 'Maestra')
    autofit(ws, [22, 14, 22, 40, 18, 14, 10, 10, 10, 10, 18, 10, 10, 18, 10, 10])

# ── Hojas de banda ────────────────────────────────────────────────────────────
BANDA_COLS = [
    'Destino', 'Clasificación Destino', 'Corporativo', 'Hotel', 'Channel',
    'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV',
    'Bookability', 'Bookings',
    'Banda Ef Global', 'Banda CV Global',
    'Ef B2C', 'Ef Opaco', 'Ef Ultra Opaco',
    'CV B2C', 'CV Opaco', 'CV Ultra Opaco',
]

def write_banda(ws, df_banda, sheet_title, banda_lookup):
    title(ws, sheet_title, f'Hoteles Global · {PERIODO} · Top 500')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return

    df_s = df_banda.sort_values('Eficacia', ascending=True).head(500)
    tab_cv_g = TAB_CV.get('global', {})
    if 'ConvRate_WoW_pp' not in df_s.columns and tab_cv_g.get('hotel') is not None:
        df_s = df_s.merge(tab_cv_g['hotel'][['Hotel', 'ConvRate_WoW_pp']], on='Hotel', how='left')

    hdr_r = 4
    r = mk_hdr_plain(ws, hdr_r, BANDA_COLS)

    for _, row in df_s.iterrows():
        ef   = fmt_pct(row.get('Eficacia'))
        cv   = fmt_pct(row.get('ConvRate'))
        bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
        bef  = banda_eficacia(ef) if ef is not None else '—'
        bcv  = banda_convrate(cv, bk) if cv is not None else '—'
        hotel   = clean(str(row.get('Hotel', '—')))
        channel = str(row.get('Channel', '—'))
        destino = str(row.get('Destino', '—'))
        corp    = str(row.get('CorpName', row.get('Corp', '—')))
        tier    = get_dest_tier(destino).capitalize()
        bk_pct  = get_bk_hotel(hotel, 'Global')
        exp     = banda_lookup.get(hotel, {})

        for ci, val in enumerate([destino, tier, corp, hotel, channel], 1):
            mk_cell(ws, r, ci, val, align='left')
        mk_cell(ws, r, 6, cru)
        pct_cell(ws, r, 7, ef)
        apply_wow(ws, r, 8, sf(row.get('Eficacia_WoW_pp')), invert=False)
        pct_cell(ws, r, 9, cv)
        apply_wow(ws, r, 10, sf(row.get('ConvRate_WoW_pp')), invert=False)
        pct_cell(ws, r, 11, bk_pct)
        mk_cell(ws, r, 12, bk)
        # Global: coloreadas
        mk_cell(ws, r, 13, bef, bef, is_sev=True)
        mk_cell(ws, r, 14, bcv, bcv, is_sev=True)
        # Exposición cruzada: texto plano
        for ci, can_label in enumerate(['B2C', 'Opaco', 'Ultra Opaco'], 15):
            mk_cell(ws, r, ci, exp.get(can_label, {}).get('ef', '—'))
        for ci, can_label in enumerate(['B2C', 'Opaco', 'Ultra Opaco'], 18):
            mk_cell(ws, r, ci, exp.get(can_label, {}).get('cv', '—'))
        r += 1

    add_excel_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [22, 14, 22, 40, 18, 10, 10, 10, 10, 10, 10, 10, 16, 16, 12, 12, 14, 12, 12, 14])

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
banda_lookup = build_banda_lookup_cr()
df_global = hotel_src_cr(None)
crit, bajo, sinc = band_split_ef(df_global)

ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = CR
write_severity(ws)
ws = wb.create_sheet('Maestra');          ws.sheet_properties.tabColor = CR
write_maestra(ws)
ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
write_banda(ws, crit, f'Críticos · CR W{VOL_NUM}', banda_lookup)
ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
write_banda(ws, bajo, f'Bajo Rendimiento · CR W{VOL_NUM}', banda_lookup)
ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
write_banda(ws, sinc, f'Sin Conversión · CR W{VOL_NUM}', banda_lookup)

out = f'{OUTPUTS}/Analisis_CheckRates_W{VOL_NUM}.xlsx'
wb.save(out)
print(f'✅ Excel CR: {out}')
print(f'   {len(wb.sheetnames)} hojas: {" | ".join(wb.sheetnames)}')
