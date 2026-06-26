"""
excel_rnd_regional.py · W26+ · Excels RND por región
Genera 4 archivos: Analisis_RND_{Region}_WNN.xlsx

Filtro sobre pickle global — P80 y métricas calculadas sobre universo completo.
MIN_TRAFICO_REGIONAL = 10K (vs 50K global) para capturar hoteles relevantes
a nivel regional que no pasan el corte del universo completo.

4 hojas por archivo: Severity | Críticos | Bajo Rendimiento | Sin Conversión
Sin hoja Maestra — el equipo comercial necesita saber qué hoteles contactar,
no cruzar canastas.
"""
import pickle, os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_nodispo, banda_convrate
from regional_config import REGIONES, MIN_TRAFICO_REGIONAL

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_RND', f'rnd_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D = pickle.load(f)

TAB_ND  = D['TAB_NoDispo']
M       = D['M']

RND = 'EA0074'

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP   = fill('EAF3DE'); WOW_UP_F   = fnt('2F6C34', bold=True)
WOW_DN   = fill('FCE8E6'); WOW_DN_F   = fnt('C0392B', bold=True)
WOW_NEU  = fill('F2EEE6'); WOW_NEU_F  = fnt('8A8377', bold=True)

RANGOS_ND = {
    'Exitosa':       '< 3%',
    'Aceptable':     '3% – 5%',
    'Revisar':       '5% – 20%',
    'Crítica':       '20% – 60%',
    'Súper Crítica': '> 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0',
    'Crítica':        '< 0,8%',
    'Revisar':        '0,8% – 1,5%',
    'Aceptable':      '1,5% – 2,5%',
    'Exitosa':        '≥ 2,5%',
}

_tc = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(RND, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(RND, 11, True)

def mk_hdr(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(RND)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_table(ws, hdr_row, end_row, n_cols, prefix):
    _tc[0] += 1
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{end_row}'
    t = Table(displayName=f'T_{prefix}_{_tc[0]}', ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center'):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Split bandas ──────────────────────────────────────────────────────────────
def band_split(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['%NoDispo'].apply(lambda v: banda_nodispo(sf(v)) if sf(v) is not None else '—')
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Hoja Severity regional ────────────────────────────────────────────────────
def write_severity(ws, df_reg, reg_label):
    title(ws, f'{reg_label} · Severity RND W{VOL_NUM}', f'Filtro regional · {PERIODO}')

    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))
    nd_curr = m_curr.get('pct_nodispo', 0); nd_prev = m_prev.get('pct_nodispo', 0)
    nd_wow  = (nd_curr - nd_prev) * 100 if nd_prev else None

    # KPI regional
    r = 4
    section_title(ws, r, f'KPI Regional {reg_label}'); r += 1
    r = mk_hdr(ws, r, ['Métrica', f'W{int(VOL_NUM)-1} Global', f'W{VOL_NUM} Regional', 'WoW Global'])

    # Regional %NoDispo
    nd_reg = None
    if len(df_reg) > 0 and '%NoDispo' in df_reg.columns:
        trf_sum = df_reg['Trafico'].sum() if 'Trafico' in df_reg.columns else 0
        if trf_sum > 0:
            nd_reg = (df_reg['%NoDispo'] * df_reg['Trafico']).sum() / trf_sum

    ws.cell(r, 1, '%NoDispo').font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
    ws.cell(r, 2, round(nd_prev, 4) if nd_prev else None).border = BD
    ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
    ws.cell(r, 3, round(nd_reg, 4) if nd_reg else None).border = BD
    ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
    apply_wow(ws, r, 4, nd_wow, invert=True); r += 1

    ws.cell(r, 1, 'Hoteles en región').font = Font(name='Arial', size=10); ws.cell(r, 1).border = BD
    ws.cell(r, 3, len(df_reg)).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
    r += 2

    # Severity %NoDispo regional
    sev_nd, sev_cv = {}, {}
    for _, row in df_reg.iterrows():
        nd  = sf(row.get('%NoDispo'))
        bk  = int(sf(row.get('Bookings')) or 0)
        trf = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        if nd is not None: b = banda_nodispo(nd); sev_nd[b] = sev_nd.get(b, 0) + 1
        b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else None)
        if b_cv: sev_cv[b_cv] = sev_cv.get(b_cv, 0) + 1
    total_nd = sum(sev_nd.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    section_title(ws, r, 'Severity %NoDispo'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_nd.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_ND.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_nd, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevND')
    r += 1

    section_title(ws, r, 'Severity %Conv Rate'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    # Top Destinos de la región
    DIM_COLS = ['País', 'Destino', 'Banda ND', 'Banda CV', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    df_dest_reg = df_reg.groupby(['PaisDestino', 'Destino'], as_index=False).agg(
        Trafico=('Trafico', 'sum'),
        Bookings=('Bookings', 'sum'),
        NoDispo_WoW_pp=('NoDispo_WoW_pp', 'mean'),
    )
    if len(df_dest_reg) > 0:
        df_dest_reg['%NoDispo'] = df_reg.groupby(['PaisDestino','Destino']).apply(
            lambda g: (g['%NoDispo'] * g['Trafico']).sum() / g['Trafico'].sum()
            if g['Trafico'].sum() > 0 else None
        ).reset_index(drop=True)

    # Usar TAB_ND global filtrado por países de la región como fuente de destinos
    df_dest_global = TAB_ND.get('destino', pd.DataFrame())
    paises_reg = set(df_reg['PaisDestino'].unique()) if 'PaisDestino' in df_reg.columns else set()

    # Reconstruir top destinos desde df_reg (tiene todos los campos necesarios)
    dest_agg = {}
    for _, row in df_reg.iterrows():
        dest = str(row.get('Destino', '—'))
        pais = str(row.get('PaisDestino', '—'))
        nd   = sf(row.get('%NoDispo'))
        bk   = int(sf(row.get('Bookings')) or 0)
        trf  = int(sf(row.get('Trafico')) or 0)
        wow  = sf(row.get('NoDispo_WoW_pp'))
        if dest not in dest_agg:
            dest_agg[dest] = {'pais': pais, 'trf': 0, 'bk': 0, 'nd_trf': 0.0, 'wow': wow}
        dest_agg[dest]['trf']    += trf
        dest_agg[dest]['bk']     += bk
        dest_agg[dest]['nd_trf'] += (nd * trf) if nd is not None else 0

    dest_rows = []
    for dest, v in dest_agg.items():
        nd_w = v['nd_trf'] / v['trf'] if v['trf'] > 0 else None
        conv = v['bk'] / v['trf'] if v['trf'] > 0 else None
        dest_rows.append({
            'Destino': dest, 'PaisDestino': v['pais'],
            'Trafico': v['trf'], 'Bookings': v['bk'],
            '%NoDispo': nd_w, 'ConvRate': conv, 'NoDispo_WoW_pp': v['wow'],
        })
    df_dest_r = pd.DataFrame(dest_rows).sort_values('%NoDispo', ascending=False).head(30)

    if len(df_dest_r) > 0:
        section_title(ws, r, 'Top Destinos · %NoDispo DESC'); r += 1
        hdr_r = r; r = mk_hdr(ws, r, DIM_COLS)
        for _, row in df_dest_r.iterrows():
            nd   = sf(row.get('%NoDispo'))
            conv = sf(row.get('ConvRate'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            b_nd = banda_nodispo(nd) if nd is not None else '—'
            b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            mk_cell(ws, r, 1, str(row.get('PaisDestino', '—')), align='left')
            mk_cell(ws, r, 2, str(row.get('Destino', '—')), align='left')
            mk_cell(ws, r, 3, b_nd, b_nd, is_sev=True)
            mk_cell(ws, r, 4, b_cv, b_cv, is_sev=True)
            mk_cell(ws, r, 5, trf)
            pct_cell(ws, r, 6, nd)
            apply_wow(ws, r, 7, sf(row.get('NoDispo_WoW_pp')), invert=True)
            pct_cell(ws, r, 8, conv)
            mk_cell(ws, r, 9, bk)
            r += 1
        add_table(ws, hdr_r, r-1, len(DIM_COLS), 'Dest')
        r += 1

    # Top Corp de la región
    CORP_COLS = ['Corporativo', 'Banda ND', 'Banda CV', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    corp_agg = {}
    for _, row in df_reg.iterrows():
        corp = str(row.get('CorpName', row.get('Corp', '—')))
        nd   = sf(row.get('%NoDispo'))
        bk   = int(sf(row.get('Bookings')) or 0)
        trf  = int(sf(row.get('Trafico')) or 0)
        wow  = sf(row.get('NoDispo_WoW_pp'))
        if corp not in corp_agg:
            corp_agg[corp] = {'trf': 0, 'bk': 0, 'nd_trf': 0.0, 'wow': wow}
        corp_agg[corp]['trf']    += trf
        corp_agg[corp]['bk']     += bk
        corp_agg[corp]['nd_trf'] += (nd * trf) if nd is not None else 0

    corp_rows = []
    for corp, v in corp_agg.items():
        nd_w = v['nd_trf'] / v['trf'] if v['trf'] > 0 else None
        conv = v['bk'] / v['trf'] if v['trf'] > 0 else None
        corp_rows.append({
            'Corp': corp, 'Trafico': v['trf'], 'Bookings': v['bk'],
            '%NoDispo': nd_w, 'ConvRate': conv, 'NoDispo_WoW_pp': v['wow'],
        })
    df_corp_r = pd.DataFrame(corp_rows).sort_values('%NoDispo', ascending=False).head(30)

    if len(df_corp_r) > 0:
        section_title(ws, r, 'Top Corporativos · %NoDispo DESC'); r += 1
        hdr_r = r; r = mk_hdr(ws, r, CORP_COLS)
        for _, row in df_corp_r.iterrows():
            nd   = sf(row.get('%NoDispo'))
            conv = sf(row.get('ConvRate'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            b_nd = banda_nodispo(nd) if nd is not None else '—'
            b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            mk_cell(ws, r, 1, str(row.get('Corp', '—')), align='left')
            mk_cell(ws, r, 2, b_nd, b_nd, is_sev=True)
            mk_cell(ws, r, 3, b_cv, b_cv, is_sev=True)
            mk_cell(ws, r, 4, trf)
            pct_cell(ws, r, 5, nd)
            apply_wow(ws, r, 6, sf(row.get('NoDispo_WoW_pp')), invert=True)
            pct_cell(ws, r, 7, conv)
            mk_cell(ws, r, 8, bk)
            r += 1
        add_table(ws, hdr_r, r-1, len(CORP_COLS), 'Corp')

    autofit(ws, [20, 28, 16, 16, 10, 10, 10, 10, 10])

# ── Hojas de banda regional ───────────────────────────────────────────────────
BANDA_COLS = [
    'País', 'Destino', 'Corporativo', 'Hotel',
    'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'WoW CV', 'Bookings',
    'Banda ND', 'Banda CV',
]

def write_banda(ws, df_banda, sheet_title):
    title(ws, sheet_title, f'Top 500 · banda por %NoDispo · {PERIODO}')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return

    df_s = df_banda.sort_values('%NoDispo', ascending=False).head(500)
    hdr_r = 4
    r = mk_hdr(ws, hdr_r, BANDA_COLS)

    for _, row in df_s.iterrows():
        nd   = sf(row.get('%NoDispo'))
        bk   = int(sf(row.get('Bookings')) or 0)
        trf  = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        wow_nd = sf(row.get('NoDispo_WoW_pp'))
        wow_cv = sf(row.get('ConvRate_WoW_pp') if 'ConvRate_WoW_pp' in row.index else None)
        bnd_nd = banda_nodispo(nd) if nd is not None else '—'
        bnd_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
        pais    = str(row.get('PaisDestino', row.get('Pais', '—')))
        destino = str(row.get('Destino', '—'))
        corp    = str(row.get('CorpName', row.get('Corp', '—')))
        hotel   = str(row.get('Hotel', '—'))

        mk_cell(ws, r, 1, pais, align='left')
        mk_cell(ws, r, 2, destino, align='left')
        mk_cell(ws, r, 3, corp, align='left')
        mk_cell(ws, r, 4, hotel, align='left')
        mk_cell(ws, r, 5, trf)
        pct_cell(ws, r, 6, nd)
        apply_wow(ws, r, 7, wow_nd, invert=True)
        pct_cell(ws, r, 8, conv)
        apply_wow(ws, r, 9, wow_cv, invert=False)
        mk_cell(ws, r, 10, bk)
        mk_cell(ws, r, 11, bnd_nd, bnd_nd, is_sev=True)
        mk_cell(ws, r, 12, bnd_cv, bnd_cv, is_sev=True)
        r += 1

    add_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [18, 24, 24, 40, 10, 10, 10, 10, 10, 10, 16, 16])

# ── Generar un Excel por región ───────────────────────────────────────────────
df_global = D.get('p80_hotel', pd.DataFrame())

# Crear carpeta regional
out_dir = Path(OUTPUTS) / f'regional'
out_dir.mkdir(parents=True, exist_ok=True)

generated = []
for reg_key, reg in REGIONES.items():
    # Filtrar por países de la región + umbral de tráfico regional
    mask_pais = df_global['PaisDestino'].isin(reg['paises'])
    df_reg = df_global[mask_pais].copy()

    if 'Trafico' in df_reg.columns:
        df_reg = df_reg[df_reg['Trafico'] >= MIN_TRAFICO_REGIONAL]

    df_reg = df_reg.reset_index(drop=True)

    # Split bandas
    crit, bajo, sinc = band_split(df_reg)

    wb = Workbook(); wb.remove(wb.active)
    _tc[0] = 0  # reset counter por archivo

    ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = RND
    write_severity(ws, df_reg, reg['label'])

    ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
    write_banda(ws, crit, f"{reg['label']} · Críticos W{VOL_NUM}")

    ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
    write_banda(ws, bajo, f"{reg['label']} · Bajo Rendimiento W{VOL_NUM}")

    ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
    write_banda(ws, sinc, f"{reg['label']} · Sin Conversión W{VOL_NUM}")

    fname = f'Analisis_RND_{reg["file"]}_W{VOL_NUM}.xlsx'
    fpath = out_dir / fname
    wb.save(fpath)
    generated.append((reg['label'], fname, len(df_reg), len(crit), len(bajo), len(sinc)))

print(f'✅ Excels RND regionales → {out_dir}')
print(f'   {"Región":<25} {"Archivo":<38} {"Total":>6} {"Crit":>5} {"BR":>5} {"SC":>5}')
for label, fname, total, crit_n, bajo_n, sinc_n in generated:
    print(f'   {label:<25} {fname:<38} {total:>6} {crit_n:>5} {bajo_n:>5} {sinc_n:>5}')
