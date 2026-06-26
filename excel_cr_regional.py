"""
excel_cr_regional.py · W26+ · Excels CR por región
Genera 4 archivos: Analisis_CR_{Region}_WNN.xlsx

El pickle CR no tiene PaisDestino — se construye el mapa Destino→País
cruzando con el pickle RND (que sí lo tiene).
Filtro sobre pickle global — P80 calculado sobre universo completo.

4 hojas por archivo: Severity | Críticos | Bajo Rendimiento | Sin Conversión
"""
import pickle, os, re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_eficacia, banda_convrate
from regional_config import REGIONES, MIN_TRAFICO_REGIONAL

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

# Cargar pickles
with open(os.getenv('PICKLE_CR', f'cr_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D_CR = pickle.load(f)
with open(os.getenv('PICKLE_RND', f'rnd_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D_RND = pickle.load(f)

M       = D_CR['M']
TAB_EF  = D_CR['TAB_EF_BY_CANASTA']
TAB_CV  = D_CR['TAB_CV_BY_CANASTA']
hcm     = D_CR.get('hotel_channel_map', {})

# Mapa Destino → País desde RND
df_rnd = D_RND.get('p80_hotel', pd.DataFrame())
DEST_PAIS = (
    df_rnd[['Destino', 'PaisDestino']]
    .drop_duplicates()
    .set_index('Destino')['PaisDestino']
    .to_dict()
) if 'PaisDestino' in df_rnd.columns else {}

def clean(n):
    return re.sub(r'^\(\d+\)\s*-\s*', '', str(n)).strip() if n else str(n)

hcm_clean = {clean(k): v for k, v in hcm.items()}

p80_all = D_CR['p80_hotel'].copy()
p80_all['Hotel']   = p80_all['Hotel'].apply(clean)
p80_all['Channel'] = p80_all['Hotel'].map(hcm_clean).fillna('—')
# Agregar País desde mapa RND
p80_all['PaisDestino'] = p80_all['Destino'].map(DEST_PAIS).fillna('—')

CR_COLOR = '5C469C'

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP   = fill('EAF3DE'); WOW_UP_F   = fnt('2F6C34', bold=True)
WOW_DN   = fill('FCE8E6'); WOW_DN_F   = fnt('C0392B', bold=True)
WOW_NEU  = fill('F2EEE6'); WOW_NEU_F  = fnt('8A8377', bold=True)

RANGOS_EF = {
    'Exitosa':       '≥ 97%',
    'Aceptable':     '93% – 97%',
    'Revisar':       '85% – 93%',
    'Crítica':       '60% – 85%',
    'Súper Crítica': '< 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0',
    'Crítica':        '< 0,8%',
    'Revisar':        '0,8% – 1,5%',
    'Aceptable':      '1,5% – 2,5%',
    'Exitosa':        '≥ 2,5%',
}

_tc = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    return round(float(v), 4)

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(CR_COLOR, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(CR_COLOR, 11, True)

def mk_hdr(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(CR_COLOR)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_table(ws, hdr_row, end_row, n_cols, prefix):
    _tc[0] += 1
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{end_row}'
    t = Table(displayName=f'T_{prefix}_{_tc[0]}', ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center'):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Split bandas ──────────────────────────────────────────────────────────────
def band_split(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['Eficacia'].apply(
        lambda v: banda_eficacia(round(float(v), 4)) if pd.notna(v) else '—'
    )
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Hoja Severity regional ────────────────────────────────────────────────────
def write_severity(ws, df_reg, reg_label):
    title(ws, f'{reg_label} · Severity CR W{VOL_NUM}', f'Filtro regional · {PERIODO}')

    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))
    ef_curr = m_curr.get('eficacia', 0); ef_prev = m_prev.get('eficacia', 0)
    cv_curr = m_curr.get('conv_rate', 0); cv_prev = m_prev.get('conv_rate', 0)
    ef_wow  = (ef_curr - ef_prev) * 100 if ef_prev else None
    cv_wow  = (cv_curr - cv_prev) * 100 if cv_prev else None

    r = 4
    section_title(ws, r, f'KPI Regional {reg_label}'); r += 1
    r = mk_hdr(ws, r, ['Métrica', f'W{int(VOL_NUM)-1} Global', f'W{VOL_NUM} Regional', 'WoW Global'])

    # KPI regional ponderado
    def regional_kpi(col, weight_col):
        if col not in df_reg.columns or weight_col not in df_reg.columns:
            return None
        w = df_reg[weight_col].fillna(0)
        return (df_reg[col].fillna(0) * w).sum() / w.sum() if w.sum() > 0 else None

    ef_reg = regional_kpi('Eficacia', 'CR_Unicos')
    cv_reg = regional_kpi('ConvRate', 'CR_Unicos')

    for label, prev, curr, wow in [
        ('Eficacia',  ef_prev, ef_reg, ef_wow),
        ('Conv Rate', cv_prev, cv_reg, cv_wow),
    ]:
        ws.cell(r, 1, label).font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
        ws.cell(r, 2, round(prev, 4) if prev else None).border = BD
        ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
        ws.cell(r, 3, round(curr, 4) if curr else None).border = BD
        ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        apply_wow(ws, r, 4, wow, invert=False)
        r += 1

    ws.cell(r, 1, 'Hoteles en región').font = Font(name='Arial', size=10)
    ws.cell(r, 1).border = BD
    ws.cell(r, 3, len(df_reg)).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
    r += 2

    # Severity Eficacia
    sev_ef, sev_cv = {}, {}
    for _, row in df_reg.iterrows():
        ef = fmt_pct(row.get('Eficacia'))
        cv = fmt_pct(row.get('ConvRate'))
        bk = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        if ef is not None: b = banda_eficacia(ef); sev_ef[b] = sev_ef.get(b, 0) + 1
        if cv is not None: b = banda_convrate(cv, bk); sev_cv[b] = sev_cv.get(b, 0) + 1
    total_ef = sum(sev_ef.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    section_title(ws, r, 'Severity Eficacia'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_ef.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_EF.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_ef, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevEf')
    r += 1

    section_title(ws, r, 'Severity Conv Rate'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    # Top Destinos de la región — reconstruido desde df_reg
    dest_agg = {}
    for _, row in df_reg.iterrows():
        dest = str(row.get('Destino', '—'))
        pais = str(row.get('PaisDestino', '—'))
        ef   = fmt_pct(row.get('Eficacia'))
        cv   = fmt_pct(row.get('ConvRate'))
        bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
        if dest not in dest_agg:
            dest_agg[dest] = {'pais': pais, 'cru': 0, 'bk': 0, 'ef_cru': 0.0, 'cv_cru': 0.0,
                              'wow_ef': sf(row.get('Eficacia_WoW_pp')),
                              'wow_cv': sf(row.get('ConvRate_WoW_pp'))}
        dest_agg[dest]['cru']    += cru
        dest_agg[dest]['bk']     += bk
        dest_agg[dest]['ef_cru'] += (ef * cru) if ef is not None else 0
        dest_agg[dest]['cv_cru'] += (cv * cru) if cv is not None else 0

    dest_rows = []
    for dest, v in dest_agg.items():
        ef_w = v['ef_cru'] / v['cru'] if v['cru'] > 0 else None
        cv_w = v['cv_cru'] / v['cru'] if v['cru'] > 0 else None
        dest_rows.append({
            'Destino': dest, 'PaisDestino': v['pais'],
            'CR_Unicos': v['cru'], 'Bookings': v['bk'],
            'Eficacia': ef_w, 'ConvRate': cv_w,
            'Eficacia_WoW_pp': v['wow_ef'], 'ConvRate_WoW_pp': v['wow_cv'],
        })
    _df_dest_tmp = pd.DataFrame(dest_rows)
    df_dest_r = _df_dest_tmp.sort_values('Eficacia', ascending=True).head(30) if len(_df_dest_tmp) > 0 and 'Eficacia' in _df_dest_tmp.columns else pd.DataFrame()

    DEST_COLS = ['País', 'Destino', 'Banda Ef', 'Banda CV', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']
    if len(df_dest_r) > 0:
        section_title(ws, r, 'Top Destinos · Eficacia ASC'); r += 1
        hdr_r = r; r = mk_hdr(ws, r, DEST_COLS)
        for _, row in df_dest_r.iterrows():
            ef  = fmt_pct(row.get('Eficacia'))
            cv  = fmt_pct(row.get('ConvRate'))
            bk  = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef = banda_eficacia(ef) if ef is not None else '—'
            bcv = banda_convrate(cv, bk) if cv is not None else '—'
            mk_cell(ws, r, 1, str(row.get('PaisDestino', '—')), align='left')
            mk_cell(ws, r, 2, str(row.get('Destino', '—')), align='left')
            mk_cell(ws, r, 3, bef, bef, is_sev=True)
            mk_cell(ws, r, 4, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 5, cru)
            pct_cell(ws, r, 6, ef)
            apply_wow(ws, r, 7, sf(row.get('Eficacia_WoW_pp')), invert=False)
            pct_cell(ws, r, 8, cv)
            apply_wow(ws, r, 9, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 10, bk)
            r += 1
        add_table(ws, hdr_r, r-1, len(DEST_COLS), 'Dest')
        r += 1

    # Top Corp de la región
    corp_agg = {}
    for _, row in df_reg.iterrows():
        corp = str(row.get('CorpName', row.get('Corp', '—')))
        ef   = fmt_pct(row.get('Eficacia'))
        cv   = fmt_pct(row.get('ConvRate'))
        bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
        if corp not in corp_agg:
            corp_agg[corp] = {'cru': 0, 'bk': 0, 'ef_cru': 0.0, 'cv_cru': 0.0,
                              'wow_ef': sf(row.get('Eficacia_WoW_pp')),
                              'wow_cv': sf(row.get('ConvRate_WoW_pp'))}
        corp_agg[corp]['cru']    += cru
        corp_agg[corp]['bk']     += bk
        corp_agg[corp]['ef_cru'] += (ef * cru) if ef is not None else 0
        corp_agg[corp]['cv_cru'] += (cv * cru) if cv is not None else 0

    corp_rows = []
    for corp, v in corp_agg.items():
        ef_w = v['ef_cru'] / v['cru'] if v['cru'] > 0 else None
        cv_w = v['cv_cru'] / v['cru'] if v['cru'] > 0 else None
        corp_rows.append({
            'Corp': corp, 'CR_Unicos': v['cru'], 'Bookings': v['bk'],
            'Eficacia': ef_w, 'ConvRate': cv_w,
            'Eficacia_WoW_pp': v['wow_ef'], 'ConvRate_WoW_pp': v['wow_cv'],
        })
    _df_corp_tmp = pd.DataFrame(corp_rows)
    df_corp_r = _df_corp_tmp.sort_values('Eficacia', ascending=True).head(30) if len(_df_corp_tmp) > 0 and 'Eficacia' in _df_corp_tmp.columns else pd.DataFrame()

    CORP_COLS = ['Corporativo', 'Banda Ef', 'Banda CV', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']
    if len(df_corp_r) > 0:
        section_title(ws, r, 'Top Corporativos · Eficacia ASC'); r += 1
        hdr_r = r; r = mk_hdr(ws, r, CORP_COLS)
        for _, row in df_corp_r.iterrows():
            ef  = fmt_pct(row.get('Eficacia'))
            cv  = fmt_pct(row.get('ConvRate'))
            bk  = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef = banda_eficacia(ef) if ef is not None else '—'
            bcv = banda_convrate(cv, bk) if cv is not None else '—'
            mk_cell(ws, r, 1, str(row.get('Corp', '—')), align='left')
            mk_cell(ws, r, 2, bef, bef, is_sev=True)
            mk_cell(ws, r, 3, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 4, cru)
            pct_cell(ws, r, 5, ef)
            apply_wow(ws, r, 6, sf(row.get('Eficacia_WoW_pp')), invert=False)
            pct_cell(ws, r, 7, cv)
            apply_wow(ws, r, 8, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 9, bk)
            r += 1
        add_table(ws, hdr_r, r-1, len(CORP_COLS), 'Corp')

    autofit(ws, [20, 26, 16, 16, 10, 10, 10, 10, 10, 10])


# ── Hoja Maestra regional CR ──────────────────────────────────────────────────
MAESTRA_COLS_CR = [
    'País', 'Destino', 'Corporativo', 'Hotel', 'Channel', 'Canasta',
    'CR Únicos', 'WoW CR', 'Eficacia', 'WoW Ef', 'Banda Ef',
    'Conv Rate', 'WoW CV', 'Banda CV', 'Bookings'
]

def write_maestra_cr(ws, df_reg, reg_label):
    """Una fila por Hotel × Canasta filtrada por región."""
    title(ws, f'{reg_label} · Maestra CR W{VOL_NUM}',
          f'Una fila por Hotel × Canasta · {PERIODO} · Filtrar por cualquier columna')
    hdr_r = 4
    r = mk_hdr(ws, hdr_r, MAESTRA_COLS_CR)

    paises_reg = set(df_reg['PaisDestino'].unique()) if 'PaisDestino' in df_reg.columns else set()
    CANASTA = D_CR.get('CANASTA', {})
    canastas = [
        ('Global',      None),
        ('B2C',         'B2C'),
        ('Opaco',       'B2B-OP'),
        ('Ultra Opaco', 'CUG'),
    ]
    for can_label, can_id in canastas:
        if can_id is None:
            df_can = df_reg.copy()
        else:
            if 'DistributionCategory' in p80_all.columns:
                df_can = p80_all[p80_all['DistributionCategory'] == can_id].copy()
            else:
                can = CANASTA.get(can_id, {})
                src = can.get('p80_hotel') or can.get('p80')
                if src is None: continue
                df_can = src.copy()
                if 'Channel' not in df_can.columns and 'Hotel' in df_can.columns:
                    df_can['Hotel_c'] = df_can['Hotel'].apply(clean)
                    df_can['Channel'] = df_can['Hotel_c'].map(hcm_clean).fillna('—')
            df_can['PaisDestino'] = df_can['Destino'].map(DEST_PAIS).fillna('—')
            df_can = df_can[df_can['PaisDestino'].isin(paises_reg)].copy()
        if len(df_can) == 0: continue

        for _, row in df_can.iterrows():
            ef   = fmt_pct(row.get('Eficacia'))
            cv   = fmt_pct(row.get('ConvRate'))
            bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef  = banda_eficacia(ef) if ef is not None else '—'
            bcv  = banda_convrate(cv, bk) if cv is not None else '—'
            hotel   = clean(str(row.get('Hotel', '—')))
            channel = str(row.get('Channel', '—'))
            destino = str(row.get('Destino', '—'))
            pais    = str(row.get('PaisDestino', '—'))
            corp    = str(row.get('CorpName', row.get('Corp', '—')))
            for ci, val in enumerate([pais, destino, corp, hotel, channel, can_label], 1):
                mk_cell(ws, r, ci, val, align='left')
            mk_cell(ws, r, 7, cru)
            apply_wow(ws, r, 8, sf(row.get('CR_WoW_pct')), invert=False)
            pct_cell(ws, r, 9, ef)
            apply_wow(ws, r, 10, sf(row.get('Eficacia_WoW_pp')), invert=False)
            mk_cell(ws, r, 11, bef, bef, is_sev=True)
            pct_cell(ws, r, 12, cv)
            apply_wow(ws, r, 13, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 14, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 15, bk)
            r += 1

    if r > hdr_r + 1:
        add_table(ws, hdr_r, r-1, len(MAESTRA_COLS_CR), 'Maestra')
    autofit(ws, [16, 22, 22, 40, 18, 14, 10, 10, 10, 10, 18, 10, 10, 18, 10])

# ── Hojas de banda ────────────────────────────────────────────────────────────
BANDA_COLS = [
    'País', 'Destino', 'Corporativo', 'Hotel', 'Channel',
    'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings',
    'Banda Ef', 'Banda CV',
]

def write_banda(ws, df_banda, sheet_title):
    title(ws, sheet_title, f'Top 500 · banda por Eficacia ASC · {PERIODO}')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return

    df_s = df_banda.sort_values('Eficacia', ascending=True).head(500)
    # Merge ConvRate_WoW_pp si no está
    tab_cv_g = TAB_CV.get('global', {})
    if 'ConvRate_WoW_pp' not in df_s.columns and tab_cv_g.get('hotel') is not None:
        df_s = df_s.merge(tab_cv_g['hotel'][['Hotel', 'ConvRate_WoW_pp']], on='Hotel', how='left')

    hdr_r = 4
    r = mk_hdr(ws, hdr_r, BANDA_COLS)

    for _, row in df_s.iterrows():
        ef   = fmt_pct(row.get('Eficacia'))
        cv   = fmt_pct(row.get('ConvRate'))
        bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
        bef  = banda_eficacia(ef) if ef is not None else '—'
        bcv  = banda_convrate(cv, bk) if cv is not None else '—'
        hotel   = clean(str(row.get('Hotel', '—')))
        channel = str(row.get('Channel', '—'))
        destino = str(row.get('Destino', '—'))
        pais    = str(row.get('PaisDestino', '—'))
        corp    = str(row.get('CorpName', row.get('Corp', '—')))

        mk_cell(ws, r, 1, pais, align='left')
        mk_cell(ws, r, 2, destino, align='left')
        mk_cell(ws, r, 3, corp, align='left')
        mk_cell(ws, r, 4, hotel, align='left')
        mk_cell(ws, r, 5, channel, align='left')
        mk_cell(ws, r, 6, cru)
        pct_cell(ws, r, 7, ef)
        apply_wow(ws, r, 8, sf(row.get('Eficacia_WoW_pp')), invert=False)
        pct_cell(ws, r, 9, cv)
        apply_wow(ws, r, 10, sf(row.get('ConvRate_WoW_pp')), invert=False)
        mk_cell(ws, r, 11, bk)
        mk_cell(ws, r, 12, bef, bef, is_sev=True)
        mk_cell(ws, r, 13, bcv, bcv, is_sev=True)
        r += 1

    add_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [18, 24, 24, 40, 18, 10, 10, 10, 10, 10, 10, 16, 16])

# ── Generar un Excel por región ───────────────────────────────────────────────
out_dir = Path(OUTPUTS) / 'regional'
out_dir.mkdir(parents=True, exist_ok=True)

generated = []
for reg_key, reg in REGIONES.items():
    # Filtrar p80_hotel CR por países de la región
    mask = p80_all['PaisDestino'].isin(reg['paises'])
    df_reg = p80_all[mask].copy()

    # p80_hotel CR ya tiene filtro MIN_CR=100 aplicado en el pipeline
    # No se aplica umbral adicional — todos los hoteles del p80 son válidos

    df_reg = df_reg.reset_index(drop=True)
    crit, bajo, sinc = band_split(df_reg)

    wb = Workbook(); wb.remove(wb.active)
    _tc[0] = 0

    ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = CR_COLOR
    write_severity(ws, df_reg, reg['label'])

    ws = wb.create_sheet('Maestra');          ws.sheet_properties.tabColor = CR_COLOR
    write_maestra_cr(ws, df_reg, reg['label'])

    ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
    write_banda(ws, crit, f"{reg['label']} · Críticos CR W{VOL_NUM}")

    ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
    write_banda(ws, bajo, f"{reg['label']} · Bajo Rendimiento CR W{VOL_NUM}")

    ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
    write_banda(ws, sinc, f"{reg['label']} · Sin Conversión CR W{VOL_NUM}")

    fname = f'Analisis_CR_{reg["file"]}_W{VOL_NUM}.xlsx'
    fpath = out_dir / fname
    wb.save(fpath)
    generated.append((reg['label'], fname, len(df_reg), len(crit), len(bajo), len(sinc)))

print(f'✅ Excels CR regionales → {out_dir}')
print(f'   {"Región":<25} {"Archivo":<38} {"Total":>6} {"Crit":>5} {"BR":>5} {"SC":>5}')
for label, fname, total, crit_n, bajo_n, sinc_n in generated:
    print(f'   {label:<25} {fname:<38} {total:>6} {crit_n:>5} {bajo_n:>5} {sinc_n:>5}')
