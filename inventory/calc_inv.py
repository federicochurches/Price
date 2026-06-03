"""
calc_inv.py — Hotel Inventory · Standalone
Genera INVENTORY_WNN.html completo desde el dataset de contratos.

Uso:
    python calc_inv.py

Configurar las variables de la sección CONFIG antes de correr.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, date
import json, sys

# ─────────────────────────────────────────────
# CONFIG — editar cada semana
# ─────────────────────────────────────────────
WEEK          = "W21"
WEEK_NUM      = 21
VOL_NUM       = "21"
YEAR_ACTUAL   = 2026
SNAPSHOT_DATE = "23 de Mayo de 2026"
INPUT_FILE    = "detalleHoteles_contratos_w21.xlsx"
OUTPUT_FILE   = f"INVENTORY_{WEEK}.html"
OUTPUT_DIR    = Path(f"inventory/week-{WEEK_NUM:02d}")  # outputs van acá
OUTPUT_DIR    = Path(f"inventory/week-{WEEK_NUM:02d}")  # carpeta de outputs
TARGET_PROPIO = 70_000
TOTAL_WEEKS_2026 = 52
SEMANAS_RESTANTES = max(1, TOTAL_WEEKS_2026 - WEEK_NUM)

# Logo PriceTravel (base64 — mismo que CR/RND)
LOGO_B64 = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAQMAAAA/CAYAAADkHq2pAAAAAXNSR0IArs4c6QAAAIRlWElmTU0AKgAAAAgABQESAAMAAAABAAEAAAEaAAUAAAABAAAASgEbAAUAAAABAAAAUgEoAAMAAAABAAIAAIdpAAQAAAABAAAAWgAAAAAAAACWAAAAAQAAAJYAAAABAAOgAQADAAAAAQABAACgAgAEAAAAAQAAAQOgAwAEAAAAAQAAAD8AAAAArgvL1QAAAAlwSFlzAAAXEgAAFxIBZ5/SUgAAIThJREFUeAHtnQl8VNW9x+8yM1kgCUtCWAIkgFq1dlHrUqviGnABqaKt74EsClSxVmtra+srj9rXPuvyqlgRjECtolKXCm2Ahwq2H1/fU9TWYktFAiQgYSchy2Tm3vu+/ztzJzOTmWQmiaL9nPNhcu4553/+539+55z/+Z/lXjRNOYWAQkAhoBBQCCgEFAIKAYWAQkAhoBBQCCgEFAIKAYWAQkAhoBBQCCgEFAIKAYWAQkAhoBBQCCgEFAIKAYWAQkAhoBBQCCgEFAIKAYWAQkAhoBBQCCgEFAKfAAQmb3IGTH72WbM3RBn/p/2Fk1+vzesNXoqHQkAhkBkCRmZk6anG/35/4aUbDv665XDze81Dxm2ofHnvyempu0hxHP3S9QfuNK2cd1ucfm9d8sq+iV3kUMkKAYVALyGg95TP+PUHvhHoW/TLcPMRzfAHNDsc2mdZ1qzVY/u/kA3v8b9/v1DvW7zA9OdOgYdm+PyaFWzZ4fP1/+xLX9Ebs+GlaBUCCoHsEeixZWDoRh/HcTTHtmTwigTFps/3zCXrD3wnU3EuXVc/yigY9DszkDdFeDhWWJSKZM8PHmnwZ8pH0SkEFALdR6DHlsGEtXuHhnN8q81A7klRZaDphhGxEkKti4381m+tPHVoczoRK9cfONtv+pbpPn+Flx8G5Pdrdlvwu78fO+Dn6fJmGa+Xl5ePDIfDXSoXwzAsv99/pF+/fgc3btzoaqUsy8qYvKysbJimBfI1rY08AS0UatpVX1/flDGDj5Fw2LBhZbqu98peDnzs0aNHb1+/fn34Y6zCp66o4447rqCpKTRYBPf7nYaampr6VJWoqKgYGQqFCpmYd+3cuXN/Kpqu4rqlDJzCWwfoDQ8c8Jhf8vLBkXrAeNHwBb4QG9Carpm5eTKgXw61HJm+tnJYrUfv+eM3HJxqmOYCXTcK7JAMBpwoAp9Ps0KtP6geW/wfkcjI38mT5wVWrJgXJYxP6fp5+PDhEyjn14DVpTKAm81PzJzddNq3bFt7MTfX/7stW7YEuy4pOwrkWo1c54p1RVka2yZfr63d9mJ2XD56auQ8lcZZQ0korp476upomjN/x44dP+s5t39aDkZZ2YiVhqGfL/0Dt9eywmft2rUrNpaGDq04zjTDWOFGK5AegWYovxrbtn9WV1fnmuqSMROX1TKhse/NJQf73fhCkx7adKjoxuqD/W4aKYX8/oL+20N2eIIdCr5l5ngTh6NZrc1iIVzgz++7js3A0z2B5jmOccmGgz82TN9Sen9MEbgWhWmiCNruiFcEMy6pOnX6+KWvFjSNenvauKrbWJRkrcQAU2Qt4JebwU86/EA67In4UzB0ngsGQxuGDSs/g3BvuwR5aERfbxfQG/wsSx8CHgPglSBvD8J5tq2P6Q3Z/ll5nHLKKZzOOcfEYTycNhjo1VesAdO0xHJehyLYgL/Rsoz7mbz2M8E8QDir072slEGbLzy3UAtcwYp+cK7uG8c89tKB/rNGiHBrzx1Yq4fCE5xQ2xtiEXhOLAXd8B2rBwLVl7126KrK1bUD3vjD4SfNnNwfsjegO5blkooi0EyfgyK4vXrsgHu8/NMql5xNwirTMMcCzAmGbt434+LFZ3rpmfqAGCkoLgNx7myczhdt7P0gPd00neqysnLk6FWXE18+nLNqk16VpBNm1D0YL2eq5+TsqWgS47Ret7SSZfi0h8ErfhllERar1XUseW+mX/6CgExe92BVnmkY9nMoht9hxTUPH15+YYQys7/ZzUK2VmgZzPiaozU5IS1P930u6GgrD/abM7H/oYXbVl1QvLPytb0T/SHteRTCGVZrxErBYtB009ffduwnfbl9d+i+wBgvTcR0FYFhOk44dGv1eQOlcq6bWll1nm7oK6AYGLYiqwOfmaNZdqjEo+mBvxsNej9gJigJxr9o4wHEHw/vcwC/f7tC0PvRFlWsnU/r7rosWV6MpHtt2yknnmIczTCcjck0n4Rwa2vr63l5ebPocIUia7JMdECxujBXtT6SJoOeei0m/u8p9JtOVW3HsVYJrXLZI4DV4N+zZ89Qllmvjhgx4pv0nc1AvgBcL3McczjhlaZpj4WzLO0yclkpAxrv0SYtdI1fM4aEWFa3OGFXITBMVx4q+sbEfocf2brmnJIPJ6yrv8LS9OfM3PyzZKkgTk4I6CEBzTTH2G2tMeF0wxRlYFuh4C3V5xUv8BKmVi660GeYz5BngG1HlKPPDGjhcPB/A5a93qPrji8dFdD21NVtFxMrrWNzDzPWpEPrY2Wgyo/nUaZpXkumh9JmzCIBGZ7Lgvyoke7du1fWo4vTCXDiiSf2bWg4chMt7SqDCJ39ZG1tnZivyvUyAihnWSqLUqZbMjNrxvF0z2rCb9bVbdvAHs9F0l2zKTYrZVDcuOjv+4pmTzF0/29NR+8jFoIoBJYMnw1p1qo9hbMnDmp49P2XLiytH/9a4yQtGBSFcLanEGQEukohKmFUEVi2HZqLIljoCT5l3KJKnxF4Gu3Rz1MEpuHTbCtc43P8X1+0buZhj7b7viPmuNQ/3gxLYMcGzJahQ0ddxwaNzNbFXiLVmMBzZ8rAjyL5IorkNGbG0bRXATNqgI2goG1b98D3fY9XWdnIK1khcbnKYbbUD/l8xt3pdoy9PCUlJX3z8/NPY4V1CnqNvRCnLzK560NZbcGnEcX9E8rZ6eWJ90eOHDmEDnQBvzPAeDhpYmZixum1yPsnTlP+e9u2bbvj83T1TOeER+JeDopT9heycnJigbJFLu0kLLfB1C8AH9rJ+SOz4CLiXatE6AzD/AHpfaGLOcI2v4NYXJvZfvof6vFOLDH6QNtMIu8kkResWP4498P7vWS6dGHZ4W9ubp1HvkFgJhbQu3V1O+6HvkNfGjKkYqTPZ19AWV9CrjLaSTBp4nk7+L/OHtG67liZmzZtahs+fMQeBj3LaLGynP9DIdwrliuWQj/qdTF94dV0dUgVn5UyEAbFhx99eV+/OXOZ4h+nDdgSdrTWiEI4HjW0cm/BDRNKGhf/o/qcgr2T1jVMatNbf4NCGBtTCFEpRBFgE4etcNuN1WMHxmacqRc/fpmpG0/CttCGrzj2CUT9HQw71td+tXZ6TZTFx+Lt2rV1x4gRI8XU+hdkEDmk3DGlpaV9UhwB6gzua+kkt0AjA5WhKcpZ9ib4KxGGsYmI/+LnOpYFt7LZc5awRVlwtGhLAz4fTU7whg4dmm+a/m8QOYsOeKzQR1yEv0cs8WzO/YHwci9O/PLy8lzLcr5DWTciDQPNyx9Ppd9oWfaHw4ePvK+2drtsQsUNtXi63n1mgB6j6+YdDJorkM3dJBPFJk7kZMCfM2bMmKXeiQ60Y8FyjrSHRxehjvylDSRPiIGxHpq7a2trX/PSyTsVnld4mIOlLHGu8dK78puaWr5mmsZtkh/pkM+pLy4ufmTfvn2NXl5RGC0tLXehlGdSFhuvEazjISf+m/DZxonBf6BMYmPA49G172BJ67SR8xQy3MMp1Bsohx8Qdwd1Lti+fUfGSwQpKwp318XGUxQfWrg05DjzcxmknhOF4NPN4wJG4LcohGMl/oULC/eHmxuvZFmwzpdf4N49YO+AI8d8lgZmm93WNjteEUwbt/irPtN4GthQBJGlPAOFuultthae8as116P9joZzZADHHA3KoDTbd0lJGTVqVBED6Ck65q9p5C8RZUhH9X5eZsLtoBFJGCgT6FKNUI1OPcrn83EMqaH9NRffpHxeEa5PWkLbSudkkD9L3vmkDZa8qVw0npMD7V5mnoehSeCTKk9P46jbZGbqP1LmTHgNFBm8n8ebNIcNsxg2kKSVy8uLzzGyfhF9aA1lXOfxQr8t9WiYmSX6EsG3Pb2zp7E+ZLney08LCvHyeEVQWjp6UHNzyyrK/g5pA4Q22bXn18pR3osYxHcn06QI6/S72ASOgvvAcewfUs44mmkq/e9e/AkUZ/fpk3c7+TtYKil4xqJijGMxGT481DBo/tzC+vI+hv+6Zlmy4KIWwmc0I7ByX8GsibKsWDNu+IHJr+6Z1Nyq34ZynAB2BZwYbOZewS9Wn1/8slfctHGPTebUYRnpeZ4ioJJYBQY2kPXtpdXXH82z9w6zIwM+1sLMWDnBYNsTxF3uNTzPbtUIH+FhH7MP58BaG7rgTa/OmfqDBlWUMsO8BM8TU/A/BB/MYld7SrKFtdFE3F/i+OvNzc0PMiiS5HPegeZNTEqWXU4Jz2dSxjHCRH48z6GT/plOF1vCxfHslUcG4eW0869hFpAyxQl2PAvme4hqQDaWks4fduzYHtlFJsHn09+ARPpPfmQ88wQhWeUeiVyOkqWQWw/CuYQfwvp4XZZoDKg1KMa/EvfZaD1ZajhToZvHr1NXVrblKwy4Uz1Z8VsptsrLNHbsWN+WLVur4H2ORyNpPMtE9ja1oz/YKFv9LJ5HejS0zQ+wKt/uYg9JR3klKEHqI+08bdiwUViKoSKspTosgg+lzGxdt5XBPG2ePSdvyk2trX2Hc6pwvuwdiHMVguY71jH8q/b3nT1h4JFH31tx3iAZEPPnzXPuXn/utsD68yoAsN2hCL5m6L4lAJPLiUMswWf6NU4Rfr509cwFscij8ECf+kx0bHsdtYEbijLgXNfaGroB7Z4w0KjLRvL9F3sAf8Rc3EtHbOM2o/T2rLS1FJCba/2YjpOgCBgIslm0kM7xFjPmQeSxsRycgQMHWpQjZlUMSAa0zI7T2jueHiT/rYWFhVWy9pQyxBUXH1eQm9tyJ3X5ntDyD6f/kOXJb7josk9CvelY98tyQPZeYoqA5xbKfYTh8zQyfJCbm9tUVFRkJ98E3b59+9+gvZBfB4e8xWBxPQng5u4LCQ1LAUM2fv+dfYRWlJAMYFkGuY72ncJezL3RjVIvuoPPIL6Bn2v14csgX4uy/KtHuGVLzdXgx45+TLE1oqzmMMifgUbaxXVSd6yhn8BjdjutMx/Zq8E67Y1dL3+yv3Pn1n8kx2Ub7rYykIIG1z/RtDd/7hTdb63L0czjg9G6ttLfURCjW0xj+f4BN5898MBDaHfU7jz3jDRBEUy/+PEvs2ZYQnJuZDIQSlqQI8SwFXx6W2vtnZGYo/OXgTSUksd7DRaV4j3pUPIsewds2nyTjuYmSQdhllmFmXbt5s2bY2tI9heiWbPzUCJj6Ff/wuCI8Sf8AB3wdiJiA97jilzeY5yvz/UCIh+8HiQ/Ay7R7dvnyvt91rAn0aEvlTpDP8ww/KzjtccSqXseYu1OvdpnRzgewUq5moEjiq7bLqq4fsYy54vIf3Vc253qMcU6WB4O29K3SqL1HIXiuZSwDNqUrhxH28YGuhAh76NxxCwBHbCOWIWCNRDOpz5PxdG4j9FNw7n0r8+jW86IynACcomCeymZ/uMIJ5gc3SmwpHnBLttwvmbpzi4OCWMsxFLI5x6CHrYviUWmejCcWzgpwCKIKU02hLiObLdtMALarPXr52U9k6YqpmOcq5g65V1RUTGShn2cRi2Nz08bP+eFmZHpYIZrWkscjXogEPDdFK8IPNru+eY4ys+XvNHO9ZecHP/3CXZQBKn4038HE/8Vb0DgN/GL78AdstHBExQFK6LLOhD1PILqOOzoR1y0bgt6qgg8fuJTj4RBRbv189Jr3Dv+sofS3mehn0V6e4RHHPXZfP1XHgslGJX3z1xTfzmarIH1MbA7pR1ruT4cWuqlp/Cl/y2Oj0eGjwLr+CLSPvfIMvC4Djj4yF/2Fc7+SZ4ReNiKLhckjX1W5jPrOI8u2Z83b55R8yfnM1xDjCVFQLYPhS1r+hPVN8Rm1hhBLzxEG6sErXwb5SUMKtKkM/Qn/nhmjnPx3ZlDihXZmBneDoWCMWWAIjidaJkBvA6yduvWrTt6QUyXBev/MwRJzzGAnvZ21L24znzqcwL5+yNhVD5nc23t6O2aVpt2ImBP8y+O4xNrTl58EfYnyUlGd8zXdLJhJrOxpiNbxOKhnDbZfE1Hny5eNkaDwWCfaLslkHH0GmGeENseYIn1GLv5NxDjLlOwhs5m+XAyx4wb26kiT2IBIuvUSPeIxFFmVXxb8Oq+zPIuL+krpL8DZvuh7gRr5034hqCRvQ7aSDsZDwujfUkh8R+H6xVlsIt3FKjJTeG49X5MeN2oiz0nPaAMbN412AuAsRQAxDIwCwxTn0PkHbGE3n9gE8e4L5mtNGK789bNkYFO/Iemqc+sq2t/q5AOPKadXp7sNxLDPQux3qygjHhH58nckb/cU1aCLW5MWdnW/9O0EWmZgAElOvF3BEpzcnJkgxEl0jsOc5gTDQ0lFcEW2Xay91GTCXdZuiGjDMxxnPdXIGtfYZOcl3r7I1VOTomEMdXfgdcrVHdcFBvo9ZmkdlAG1H88RcQsQMqsD4X8CUsKZBrllRTldwpLrg68PBrxRUY8t4WjeYaxNCxiY/BAPF30WU4TRFF8JK7HyoB3E4rY7l/u180TWuL2xthD4MpyuC4n7F/VqeSOswwL4oJIW0YUORdzTJ/h/+70ysdql6y5fkGn+XuQGAU/LYd4xQDta+xpzKWR3k3KMCAxbB5MDHc/FLlyurfI44AMspm81wtn4qNIOKpLoCxkBvxiQkyKQFKePhQ8ELJeUwZgW0R93NkwUrzewCwaTCFKQhSDt5Kxs4jIEREll1i5BOLMAgsh42iO4e1WWr+qoqLi3yPLiBgDuT49S8oTJ/0CC/GZ+voP9kRiIn/JLy+3xUcNAOuk/hGfHMsXH1nEeJc2T6UMRMbEqSE+Zw+fe8QYfPyObVbl6eaZ3mmCyMN1ZWwc54Cth6f2bfpFpztn5WfWPRmygg/KyUFEIQgHubhua7zVeO+08Y9dLjEfgaMXyZom1U8L0ajc4nPe5/cM7XvVoEElF6ZQBLI8iOvQImXiLbyeyC1XTuGfMBPw3krCsqYr/sifrPCJcuTGWqc/+Eo58pPRdjgUMmKnJ4R77HgLMpAtE25OHk8fWU57jEB+d/DGDT5RJN5PNndlR75L5cLAWwOv9zw+8C4Jhaxr4mUbOnTkF0g/V8qMOvjaj3kBz8eqSMa6U4zh56aTPx7r/YRF/o/dJQuflQAHCs172CS8sjlun4CbifQeZ1+bbk0uObRovccw/GLp5RifE7ECuCrpvBNsDi3L//qBWlkqQP+t6ZWP26Yv8C02XMgiDW3z18zRNd+S6y5eXLls7Q2dmlteOZn40vDI8AGz3VWcy7ZvWEQzS3I4bDYaRtteFECLRHOUlQnrXqXh2M9mR1w2mTyHuIa7geVFZOC78gud1JsZ7RWWOt/JIJ9HgkJyDu3aVbPVizhaPnJwWzPy4lhUBo4h7Yexlp5nwozNpJy329KuxJ3DHssS8qUVWU6FuKxTBcF97UTODN61WOgdu/LCzwysEVd5RTFcl2piwApLwJpyn8My+Gk73y6fdD5Qsre2dteHaSh5yTf6mm8agp5Ed1sZ7O8351aWAt+KtwiiimB32HYmlzQu/KMI5jyrmXag9AHDp9/sznHoQN3Ur8zVA9e3vVh6TeCK+v9FQTi8W3Xr9Moq1geBb3sKQU4YTMM/kJeblnMEecGStTNqe1LZpLwtrBn/nBT3SQuiCHSZKVwXUWL2GAIuttHorrxd8QTog0IU29vxcZ+G53L3KrV9HvK7LorF7RyR/jKd/CNGVIyOGDbpKCLxDOLlKEk5ZnRvP8L7cw0NDecRXjN48Bg2kNuujucDvSwtOjj2Z3aRFufcK8G9inWqySuuwB49JoieKaf9RTdeFXDM/wy5s3ckl+wR2BwvNjnWxAGNv4wognmaD0XwsBHQb7bD8rETfm38WtjZNrSRpqH9NvTioC975S5ZM/N2y2q7P37JwOvKohCOYaX05IwJVQUebS/40q18vcDnI2XB7LI5sQCdjazMHRPJZni41k90hvxsxNzOnMcngZJ6yNq7OFoHsewOMouu6Ew2BnjCEisdLcqRmViPHTOiDLCG9BuE3u8PclVaHyTPEQXkvBsIBP5bwskOi2uTFxeRUz9NXgrz4nrB5wVfI2lZ2gtcoyyyVgb7C278sk/TH2OC94shL04UAZ8pqWtz7AlDGhayU40efVXzWZ8vXYgimG3z0YNkS80KcWrA+T1AvxhaUXK25BHHbcNvy61DeUsR+N04+ZYBry+fbYf0hZMn987/zeAy/gT+oRNF576IcHTLVzwxowPhMnabT/fiuvK5ByHK5B/SkcXh84Uh7XY38Cn6IzuniJvQX/v2lUOE9I4qp18fJGVjRpc9gDaJFpzJOy6iNI2p8aTEJxwnxqeR721+MRMfrAdwN+GWeJpsngsKChLklzZEKX4hGx7Z0CaA21XGw4U3jmHwLjc0vYivHbnkriLQnB1tmnN5yeFH3HW98+YpfuvQoMWs+Ge6iiDKWLpj7EU7ni2sBRRCiZljPh9aMXhslEwUwnd5dfmniQohqHHCcG2fI43zoUsYMF6+fwY/WfPTAdbSwXi1ODaY8zkKe4qOen4m9Y3clNSf8milo9Pdp3Oe/j0eMpo5vbxH0+doT+6cHPFkAI/+fIpushdO7be/PxJNT9tvZOkENus9nKHnBMVZBlZfjGAmSkLby4bj06nL0jTuJxyEii8NRYqRfIyX28B6dro8ncVHPxYbWyYKP9r+Dtpe7iL0usvKTOadsdsLdN+IRsdVoLyYbWohzakJGvYVpVw8Eunef3BMjl23c7GZa0yRZYHnBB9RBJatBU2fniOKQJz4hIvNXO250IrSq/2T61+W+CWrZ9x53bjHLZ/h+6F80wDbAq0YcsGYUVn17ONrZn7S1/tSjR476WB0prth9Kgwi3QIfRQddS3xGwi/jjGxA/0ou+cuqNyjOkKnXetdmQa3habpm0YnHS35cbSE/lM2J1lyONK5/2rbpnvBi9cb5N69HCPK8eNWyv8N/lF3W7duPYy8G6nDUKlDtB68WTmS+wrWE1wl3sEFIOmYJu8Y5PJVJk4c7G/EC062Q/HhpGdY2gu5In2xF09ZX4qWQ7+TTWf7mZqa7fVeeiof3O+jba6GflBUTjHrFyL7JNppBbcE/sbLl9JWbIXZfN7LKabMk6nDW+x/rOnIU3+NuK/ExZdTj/Xwq6btXsVS+FVvXQbLShkgkF90nvz4oIkWcqzNYTt8RenhRX8XYZ0l5bl2/8YqI0e/Nl4RyDUJjF+HfYMfYe2tNmynCquBT6snKIQBRo72m9CKQdf4J+9ZK/yWrZ5x17TxVQ57BnfJZqKAy7cOzLChd24fSuZ/IseAXMzllROZZeTzVt5AAFX9fDodv+TK8qanbU8gdqWk0Fn2sbTgvXpT3nx0bxUKH57Pgcc58szHplxTj79EC0dh6jSWl5e/iVLZJnyOtuMbDQ9wI1P2TLx+m4ekdzHN3NbW1lbLAJFBJheNqKNWShrvu0T6mMjOUsDFQ57TuNXQ/43q89WgGM4eaZCXn2Qp0akTrLgLMZsZ/Gn45Hjl81yJPJUoCuSI3M6jCKKxjYEaRbGLW5mfi76zECtDli+0ibwc5d6EjfIrIIzCMXgpypBl4KuxDD14MLLKa+r3N2nhd/2aGUQR/CnoWJfLa8rCw3m2LM8a0LKsgyLwoQV0rQFFcJ05ac+PA1fufoOYi9gz+J2ZyzlCtCNHlgxaP+KeDb9QIg3uuqXVM/+NL6D8iIq3UHHbckJPFWoNb3rpmfqAmFzXXjGRadAEPoSTy+lKxKT8HeSU/A4fv7iFTnEznWCP9B75eR022ZcMKIOEI0iOwjYwmCYin7t/EJ9f6HEit/zkKBHerlYogE/C7CqEqRx00pIJdU+BeSxrx7SuN/t27tz+KoNmDrIfiscAplwV1nmzVD+Zn7xkNZI49wZlHN0yNv6WxgRI8RA5RtYfT04SHuDxSk1NTUbWKDP8i9BfQx3d5V0kf4JyScAa7FAQxlCsiinJZbN8qWHf6Fp41Xl1ERrC7o+87pFncr7uhBMarysGAw88vKk1p+FMv21+fs9hc6x84kzyOCuH5luB8BNmQL86wSLglgBCb7VC9qW+SXue8PjrE/fUm/6CK8NB50GOHEVju879TKKjFRmm8Uz4+dLLPPpla2bMt8P2yaSfVn563ZSHqr8Z9NIy9QFbbotFNjrcTLqYe3HhTDkl0tFQu9tjHL5ea2d1Q5AGFjlcB1Zhwvu8cLLPSzwLuPxzGv2AL+i478eL2duhDsInWt8EFnV129azZDiLFruThHf4tXgdLNknLSxtx+9vCUzSBOjInLEnyM63G3ysoVM76iBprV4q4y2pfbyURB8MqsLhEG/5ua8fy0REX4jOKFFSZJanJrx/UNcnGGuX1tbumBb/HkGUNIVnP0X+hONYiPgAjZHVTVgUwm9RCGfyuxt+csoQTMa4PSz85YOmujuekoXCMlxnWeEvU597I3VKuM/QlEzf3XAiit3g4jxb0tfyG08wo1/hKQJ3imDWt4LaH0JtznV5V9fXpGNtvTB4LiuOn6M2ckUZiHOXFWwWsVXwdd9X61dFYnv2V6728jWak1ivBfjPmsSJRRc3kLvHn3V7f5Y+x/o4YqFBg2jyd+EUrUnXPOXde2asMUIZCmmtO3duk/xW1zk1PrFWJrOJrE3lRR1RqQ5KQD4IxB2KbbKH00FReHwFj927D1bwfcfhTOpFKDEXFTqkTPK8SmzsCgYbt8Z/wcfLm84vx5E2WNIRopmZ9q88ppNBx5xm6WOIySujN+v2oP55mO78T0LOEIxGzxLiIpJzGFz2sJ7eHZntRaLMHRt0FchUSntwtCjtEurRnRT5+A08RiFXGdgWArG7zKEMLkY5jUi2E/+DTGSVOkNfxkqJV8tt2R95LflbD5nXNJGyx8rAeqF0lpGvP2o1u9rY3STU2VlgIC8zNL7rPrH9u3CJRbeHwi8MHm/4nCqWT0PkHoI40+Xh1BzQrc+XZMCjnZt6UggoBLqDQFbLhNQF6P3cvWkSZUZnBc03Tp27jIm7p2eiCISnb9Lu6lBQv4g9Qv5HJtbDEok9xLK1KMfK6bU1kbBVTiGgEEiNQI+VAZbKU06ztlFe0WBO382FwX/1T6q/mwEdmeJTl9shNueq3Zv0I0Ylp5bPwYt9Ij7NxUc5C7/avf9EskMBKkIhoBDoFIEeLxOEu+wbaIXmCVowXIc1kLz50qkAyYkYBHpo5eBTnbDTkvPVellzKqcQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAgoBBQCCgGFgEJAIaAQUAh8ChH4f83+7IatsdAQAAAAAElFTkSuQmCC"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fmt_n(n):
    return f"{int(n):,}".replace(",", ".")

def fmt_pct(v, decimals=1):
    return f"{v:.{decimals}f}%"

def channel_active(val):
    if val is None or val == '-' or val == '': return False
    try: return float(val) > 0
    except: return False

def channel_intensity(val):
    if val is None or val == '-' or val == '': return 0
    try:
        v = float(val)
        return v if v > 0 else 0
    except: return 0

# ─────────────────────────────────────────────
# 1. CARGA Y LIMPIEZA
# ─────────────────────────────────────────────
print(f"[1/5] Cargando {INPUT_FILE}...")
df_raw = pd.read_excel(INPUT_FILE, header=2, dtype=str)
df_raw['Hotel'] = df_raw['Hotel'].str.strip()
df_raw.columns = list(df_raw.columns[:20]) + ['Expedia_tercero'] + list(df_raw.columns[21:])
df_raw['HtActive'] = pd.to_numeric(df_raw['HtActive'], errors='coerce').fillna(0).astype(int)
df_raw['Prpios']   = pd.to_numeric(df_raw['Prpios'],   errors='coerce').fillna(0).astype(int)
df_raw['Terceros'] = pd.to_numeric(df_raw['Terceros'], errors='coerce').fillna(0).astype(int)

df_sistema = df_raw[df_raw['HtActive'] == 1].copy()
df_sinc    = df_sistema[df_sistema['TipoHotel'] == 'sincontrato'].copy()
df         = df_sistema[df_sistema['TipoHotel'] != 'sincontrato'].copy()

print(f"    Sistema:      {fmt_n(len(df_sistema))}")
print(f"    Con contrato: {fmt_n(len(df))} (universo)")
print(f"    Sin contrato: {fmt_n(len(df_sinc))} (gestión)")

# ─────────────────────────────────────────────
# 2. REGLAS DE NEGOCIO
# ─────────────────────────────────────────────
print("[2/5] Aplicando reglas...")

REGION_MAP = {'MX':'México','CAR':'Caribe','LATAM':'LATAM','CO':'Colombia',
              'USA/CA':'USA/CA','EU':'EMEA','MEA':'EMEA','APAC':'APAC'}
REGION_ORDER = ['México','Caribe','LATAM','Colombia','USA/CA','EMEA','APAC']
TIPO_MAP = {'sólo propio':'Solo Propio','Propio_con_tercero':'Hybrid','sólo terceros':'Third Party'}

df['Region_display'] = df['region_supply'].map(REGION_MAP).fillna('Otros')
df['Tipo_display']   = df['TipoHotel'].map(TIPO_MAP).fillna(df['TipoHotel'])
df_sinc['Region_display'] = df_sinc['region_supply'].map(REGION_MAP).fillna('Otros')

# ─────────────────────────────────────────────
# 3. KPIs
# ─────────────────────────────────────────────
print("[3/5] Calculando KPIs...")

N         = len(df)
N_sistema = len(df_sistema)
N_sinc    = len(df_sinc)

solo_propio = (df['TipoHotel'] == 'sólo propio').sum()
hybrid      = (df['TipoHotel'] == 'Propio_con_tercero').sum()
solo_terc   = (df['TipoHotel'] == 'sólo terceros').sum()
pp          = solo_propio + hybrid
gap         = TARGET_PROPIO - pp
pct_avance  = pp / TARGET_PROPIO * 100
ritmo_nec   = round(gap / SEMANAS_RESTANTES)

# Por región
def region_stats(dataframe):
    rows = []
    for reg in REGION_ORDER:
        sub = dataframe[dataframe['Region_display'] == reg]
        sp  = (sub['TipoHotel'] == 'sólo propio').sum()
        hy  = (sub['TipoHotel'] == 'Propio_con_tercero').sum()
        st  = (sub['TipoHotel'] == 'sólo terceros').sum()
        tot = len(sub)
        pp_r = sp + hy
        rows.append({'region':reg,'total':tot,'solo_propio':sp,'hybrid':hy,
                     'prod_propio':pp_r,'solo_tercero':st,
                     'pct_propio': pp_r/tot*100 if tot else 0,
                     'share': tot/N*100 if N else 0})
    return rows

reg_stats    = region_stats(df)
sinc_by_reg  = df_sinc['Region_display'].value_counts().to_dict()

# Corporativos
corp_grp = df.groupby('Corporativo').agg(
    total       =('IdHotel','count'),
    solo_propio =('TipoHotel', lambda x: (x=='sólo propio').sum()),
    hybrid      =('TipoHotel', lambda x: (x=='Propio_con_tercero').sum()),
    solo_tercero=('TipoHotel', lambda x: (x=='sólo terceros').sum()),
).reset_index()
corp_grp['prod_propio'] = corp_grp['solo_propio'] + corp_grp['hybrid']
corp_grp['pct_propio']  = corp_grp['prod_propio'] / corp_grp['total'] * 100
corp_grp = corp_grp.sort_values('total', ascending=False).reset_index(drop=True)
# JSON para filtrado dinámico (top 200)
corp_json = corp_grp.head(200).to_dict('records')
for r in corp_json:
    for k,v in r.items():
        if hasattr(v,'item'): r[k] = v.item()

# Destinos
dest_grp = df.groupby(['Destino','Region_display']).agg(
    total       =('IdHotel','count'),
    prod_propio =('TipoHotel', lambda x: ((x=='sólo propio')|(x=='Propio_con_tercero')).sum()),
    solo_tercero=('TipoHotel', lambda x: (x=='sólo terceros').sum()),
).reset_index()
dest_grp['pct_propio'] = dest_grp['prod_propio'] / dest_grp['total'] * 100
dest_grp = dest_grp.sort_values('total', ascending=False).reset_index(drop=True)
dest_json = dest_grp.head(500).to_dict('records')
for r in dest_json:
    for k,v in r.items():
        if hasattr(v,'item'): r[k] = v.item()

# ── MARKET PENETRATION (proxy: dataset como mercado) ──────────────────────
# Third Party = sin contratación directa
df_tp = df[df['TipoHotel']=='sólo terceros'].copy()
df_pp = df[df['TipoHotel'].isin(['sólo propio','Propio_con_tercero'])].copy()

# KPIs globales
market_total   = N                          # 309,509
market_pp      = pp                         # 52,491 con directo
market_tp      = int(solo_terc)             # 257,018 sin directo
market_share   = pp / N * 100              # 17.0%

# Sin contratación directa por corporativo (excluye AA-Independent)
corp_tp = df_tp[df_tp['Corporativo']!='AA-Independent'].groupby('Corporativo').size().reset_index(name='sin_directo')
corp_total_mkt = df[df['Corporativo']!='AA-Independent'].groupby('Corporativo').size().reset_index(name='total')
corp_pp_mkt    = df[(df['TipoHotel'].isin(['sólo propio','Propio_con_tercero'])) &
                    (df['Corporativo']!='AA-Independent')].groupby('Corporativo').size().reset_index(name='con_directo')
corp_mkt = (corp_total_mkt
            .merge(corp_tp, on='Corporativo', how='left')
            .merge(corp_pp_mkt, on='Corporativo', how='left')
            .fillna(0))
corp_mkt['sin_directo'] = corp_mkt['sin_directo'].astype(int)
corp_mkt['con_directo'] = corp_mkt['con_directo'].astype(int)
corp_mkt['pct_penetracion'] = corp_mkt['con_directo'] / corp_mkt['total'] * 100
corp_mkt = corp_mkt[corp_mkt['sin_directo'] > 0].sort_values('sin_directo', ascending=False).reset_index(drop=True)
corp_mkt_json = corp_mkt.head(200).to_dict('records')
for r in corp_mkt_json:
    for k,v in r.items():
        if hasattr(v,'item'): r[k] = v.item()

# Sin contratación directa por región
reg_tp    = df_tp.groupby('Region_display').size().reset_index(name='sin_directo')
reg_total_mkt = df.groupby('Region_display').size().reset_index(name='total')
reg_pp_mkt    = df[df['TipoHotel'].isin(['sólo propio','Propio_con_tercero'])].groupby('Region_display').size().reset_index(name='con_directo')
reg_mkt = (reg_total_mkt
           .merge(reg_tp, on='Region_display', how='left')
           .merge(reg_pp_mkt, on='Region_display', how='left')
           .fillna(0))
reg_mkt['sin_directo']    = reg_mkt['sin_directo'].astype(int)
reg_mkt['con_directo']    = reg_mkt['con_directo'].astype(int)
reg_mkt['pct_penetracion'] = reg_mkt['con_directo'] / reg_mkt['total'] * 100
reg_mkt['pct_sin']         = reg_mkt['sin_directo']  / reg_mkt['total'] * 100
REGION_ORDER_SET = set(REGION_ORDER)
reg_mkt = reg_mkt[reg_mkt['Region_display'].isin(REGION_ORDER_SET)]
reg_mkt['_order'] = reg_mkt['Region_display'].map({r:i for i,r in enumerate(REGION_ORDER)})
reg_mkt = reg_mkt.sort_values('sin_directo', ascending=False)

# Sin contratación directa por destino (top 500, excluye AA-Independent)
dest_tp = (df_tp[df_tp['Corporativo']!='AA-Independent']
           .groupby(['Destino','Region_display']).size().reset_index(name='sin_directo'))
dest_mkt = (dest_grp[['Destino','Region_display','total','prod_propio']]
            .merge(dest_tp, on=['Destino','Region_display'], how='left')
            .fillna(0))
dest_mkt['sin_directo']     = dest_mkt['sin_directo'].astype(int)
dest_mkt['con_directo']     = dest_mkt['prod_propio'].astype(int)
dest_mkt['pct_penetracion'] = dest_mkt['con_directo'] / dest_mkt['total'] * 100
dest_mkt = dest_mkt[dest_mkt['sin_directo'] > 0].sort_values('sin_directo', ascending=False).reset_index(drop=True)
dest_mkt_json = dest_mkt.head(500).to_dict('records')
for r in dest_mkt_json:
    for k,v in r.items():
        if hasattr(v,'item'): r[k] = v.item()

# Sin contratación directa — Independientes (AA-Independent) por región y destino
df_tp_indep = df_tp[df_tp['Corporativo']=='AA-Independent'].copy()
df_pp_indep = df_pp[df_pp['Corporativo']=='AA-Independent'].copy()

# Independientes por región
indep_reg_tp  = df_tp_indep.groupby('Region_display').size().reset_index(name='sin_directo')
indep_reg_tot = df[df['Corporativo']=='AA-Independent'].groupby('Region_display').size().reset_index(name='total')
indep_reg_pp  = df_pp_indep.groupby('Region_display').size().reset_index(name='con_directo')
indep_reg = (indep_reg_tot
             .merge(indep_reg_tp, on='Region_display', how='left')
             .merge(indep_reg_pp, on='Region_display', how='left')
             .fillna(0))
indep_reg['sin_directo']     = indep_reg['sin_directo'].astype(int)
indep_reg['con_directo']     = indep_reg['con_directo'].astype(int)
indep_reg['pct_penetracion'] = indep_reg['con_directo'] / indep_reg['total'] * 100
indep_reg['pct_sin']         = indep_reg['sin_directo'] / indep_reg['total'] * 100
indep_reg = indep_reg[indep_reg['Region_display'].isin(set(REGION_ORDER))]
indep_reg = indep_reg.sort_values('sin_directo', ascending=False)

# Independientes por destino (top 500)
indep_dest_tp  = df_tp_indep.groupby(['Destino','Region_display']).size().reset_index(name='sin_directo')
indep_dest_tot = df[df['Corporativo']=='AA-Independent'].groupby(['Destino','Region_display']).size().reset_index(name='total')
indep_dest_pp  = df_pp_indep.groupby(['Destino','Region_display']).size().reset_index(name='con_directo')
indep_dest = (indep_dest_tot
              .merge(indep_dest_tp, on=['Destino','Region_display'], how='left')
              .merge(indep_dest_pp, on=['Destino','Region_display'], how='left')
              .fillna(0))
indep_dest['sin_directo']     = indep_dest['sin_directo'].astype(int)
indep_dest['con_directo']     = indep_dest['con_directo'].astype(int)
indep_dest['pct_penetracion'] = indep_dest['con_directo'] / indep_dest['total'] * 100
indep_dest = indep_dest[indep_dest['sin_directo'] > 0].sort_values('sin_directo', ascending=False).reset_index(drop=True)
indep_dest_json = indep_dest.head(500).to_dict('records')
for r in indep_dest_json:
    for k,v in r.items():
        if hasattr(v,'item'): r[k] = v.item()

n_indep_sin = int(df_tp_indep['IdHotel'].count()) if 'IdHotel' in df_tp_indep.columns else int(len(df_tp_indep))
print(f"    Independientes sin directo: {fmt_n(n_indep_sin)} | Destinos: {len(indep_dest)}")
def channel_stats_propio(dataframe):
    channels = [('DerbySoft','DerbySoft'),('HBSI','HBSI'),('Internal','Internal'),
                ('Omnibees','Omnibees'),('Siteminder','Siteminder'),('SynXis','SynXis'),('Travelclick','Travelclick')]
    rows = []
    for name, col in channels:
        mask = dataframe[col].apply(channel_active)
        sub  = dataframe[mask]
        n    = len(sub)
        if n == 0:
            rows.append({'channel':name,'hoteles':0,'avg_contratos':0,'destinos':0,'avg_destinos':0,'top_region':'—','top_corp':'—','residual':False})
            continue
        avg_ctr  = sub[col].apply(channel_intensity).mean()
        destinos = sub['Destino'].nunique()
        top_reg  = sub['Region_display'].value_counts().index[0] if n else '—'
        top_corp = sub['Corporativo'].value_counts().index[0] if n else '—'
        avg_dest = round(destinos/n, 1) if n else 0
        rows.append({'channel':name,'hoteles':n,'avg_contratos':round(avg_ctr,1),
                     'destinos':destinos,'avg_destinos':avg_dest,'top_region':top_reg,'top_corp':top_corp,'residual':False})
    return rows

def channel_stats_tercero(dataframe):
    channels = [('Expedia','Expedia_tercero',False),('HotelBeds Apitude','HotelBeds Apitude',False),
                ('Hotel Unico V2','Hotel Unico V2',False),('Travelgate','Travelgate',False),
                ('RateFox','RateFox',False),('Tourico','Tourico',True)]
    rows = []
    for name, col, residual in channels:
        mask = dataframe[col].apply(channel_active)
        sub  = dataframe[mask]
        n    = len(sub)
        if n == 0:
            rows.append({'channel':name,'hoteles':0,'destinos':0,'top_region':'—','top_destino':'—','top_corp':'—','residual':residual})
            continue
        destinos = sub['Destino'].nunique()
        top_reg  = sub['Region_display'].value_counts().index[0] if n else '—'
        top_dest = sub['Destino'].value_counts().index[0] if n else '—'
        top_corp = sub['Corporativo'].value_counts().index[0] if n else '—'
        rows.append({'channel':name,'hoteles':n,'destinos':destinos,
                     'top_region':top_reg,'top_destino':top_dest,'top_corp':top_corp,'residual':residual})
    return rows

ch_propio   = channel_stats_propio(df)
ch_tercero  = channel_stats_tercero(df)
max_avg_ctr = max((r['avg_contratos'] for r in ch_propio), default=1) or 1

# Channel stats por tipo (para pills de filtro en la vista Channel)
df_sp = df[df['TipoHotel'] == 'sólo propio'].copy()
df_hy = df[df['TipoHotel'] == 'Propio_con_tercero'].copy()
ch_propio_sp = channel_stats_propio(df_sp)
ch_propio_hy = channel_stats_propio(df_hy)
ch_tercero_sp = channel_stats_tercero(df_sp)
ch_tercero_hy = channel_stats_tercero(df_hy)
max_avg_ctr_sp = max((r['avg_contratos'] for r in ch_propio_sp), default=1) or 1
max_avg_ctr_hy = max((r['avg_contratos'] for r in ch_propio_hy), default=1) or 1

# Datos a nivel hotel por channel para drill
CHANNEL_COL_MAP = {
    'DerbySoft':       'DerbySoft',
    'HBSI':            'HBSI',
    'Internal':        'Internal',
    'Omnibees':        'Omnibees',
    'Siteminder':      'Siteminder',
    'SynXis':          'SynXis',
    'Travelclick':     'Travelclick',
    'Expedia':         'Expedia_tercero',
    'HotelBeds':       'HotelBeds Apitude',
    'Hotel Unico V2':  'Hotel Unico V2',
    'Travelgate':      'Travelgate',
}
ch_drill_data = {}
for ch_label, ch_col in CHANNEL_COL_MAP.items():
    if ch_col not in df.columns: continue
    sub = df[df[ch_col].apply(channel_active)][['Hotel','Region_display','Corporativo','Destino']].copy()
    sub = sub.rename(columns={'Region_display':'region'})
    ch_drill_data[ch_label] = sub.head(500).to_dict('records')

# Histórico — acumulado global continuo
df_hist = df[df['FechaCreación'].notna() & (df['FechaCreación'] != '-')].copy()
df_hist['fecha_dt'] = pd.to_datetime(df_hist['FechaCreación'], errors='coerce')
df_hist = df_hist[df_hist['fecha_dt'].notna()]
df_hist = df_hist[df_hist['TipoHotel'].isin(['sólo propio','Propio_con_tercero'])]
df_hist = df_hist[df_hist['fecha_dt'].dt.year >= 2021].copy()
df_hist['year']  = df_hist['fecha_dt'].dt.year
df_hist['month'] = df_hist['fecha_dt'].dt.month
df_hist['yw']    = df_hist['year'].astype(str) + '-W' + df_hist['fecha_dt'].dt.isocalendar().week.astype(int).astype(str).str.zfill(2)
df_hist['ym']    = df_hist['year'].astype(str) + '-' + df_hist['month'].astype(str).str.zfill(2)
df_hist['ch']    = df_hist['TipoHotel'].map({'sólo propio':'Solo Propio','Propio_con_tercero':'Hybrid'})

# Acumulado por año (global)
by_year_g = df_hist.groupby('year').size().reset_index(name='netnew').sort_values('year')
cum = 0; acum_years = []
for _, r in by_year_g.iterrows():
    cum += int(r['netnew'])
    acum_years.append({'year': int(r['year']), 'netnew': int(r['netnew']), 'acum': cum})

# Acumulado por mes (global) — con fill hasta el mes del snapshot
by_month_g = df_hist.groupby('ym').size().reset_index(name='netnew').sort_values('ym')
month_netnew = {r['ym']: int(r['netnew']) for _, r in by_month_g.iterrows()}

# Mes del snapshot derivado de WEEK_NUM (aprox: cada 4.33 semanas = 1 mes)
import math
snapshot_month = min(12, math.ceil(WEEK_NUM / 52 * 12))
snapshot_ym    = f"{YEAR_ACTUAL}-{snapshot_month:02d}"
first_ym       = by_month_g['ym'].iloc[0]
if snapshot_ym < by_month_g['ym'].iloc[-1]:
    snapshot_ym = by_month_g['ym'].iloc[-1]

all_months = []
yr, mo = int(first_ym.split('-')[0]), int(first_ym.split('-')[1])
s_yr, s_mo = int(snapshot_ym.split('-')[0]), int(snapshot_ym.split('-')[1])
while (yr, mo) <= (s_yr, s_mo):
    ym = f"{yr}-{mo:02d}"
    all_months.append({'ym': ym, 'netnew': month_netnew.get(ym, 0)})
    mo += 1
    if mo > 12: mo = 1; yr += 1

cum = 0; acum_months = []
for r in all_months:
    cum += r['netnew']
    acum_months.append({'ym': r['ym'], 'netnew': r['netnew'], 'acum': cum})

# Acumulado por semana ISO (global) — con fill de semanas sin datos
by_week_g = df_hist.groupby(['yw','ym']).size().reset_index(name='netnew').sort_values('yw')
# Construir lookup de netnew por yw
week_netnew = {r['yw']: int(r['netnew']) for _, r in by_week_g.iterrows()}
# Determinar yw de la primera semana con dato hasta la última
first_yw = by_week_g['yw'].iloc[0]
# Extender hasta la semana del snapshot (WEEK_NUM del año actual)
last_yw_data = by_week_g['yw'].iloc[-1]
snapshot_yw  = f"{YEAR_ACTUAL}-W{WEEK_NUM:02d}"
last_yw = snapshot_yw if snapshot_yw > last_yw_data else last_yw_data
# Generar todas las semanas ISO entre first y last
from datetime import date, timedelta
def yw_to_date(yw):
    yr, wk = int(yw.split('-W')[0]), int(yw.split('-W')[1])
    return date.fromisocalendar(yr, wk, 1)  # lunes de esa semana
def date_to_yw(d):
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"
def date_to_ym(d):
    return f"{d.year}-{d.month:02d}"

all_weeks = []
cur = yw_to_date(first_yw)
end = yw_to_date(last_yw)
while cur <= end:
    yw = date_to_yw(cur)
    all_weeks.append({'yw': yw, 'ym': date_to_ym(cur), 'netnew': week_netnew.get(yw, 0)})
    cur += timedelta(weeks=1)

cum = 0; acum_weeks = []
for r in all_weeks:
    cum += r['netnew']
    acum_weeks.append({'yw': r['yw'], 'ym': r['ym'], 'netnew': r['netnew'], 'acum': cum})

# Sets de años y meses — from acum_weeks (includes fill weeks)
years_available = sorted(set(int(r['yw'][:4]) for r in acum_weeks))
months_by_year  = {}
for r in acum_weeks:
    yr  = int(r['yw'][:4])
    mo  = int(r['ym'].split('-')[1])
    if yr not in months_by_year:
        months_by_year[yr] = set()
    months_by_year[yr].add(mo)
months_by_year = {yr: sorted(mos) for yr, mos in months_by_year.items()}
# Cap current year to snapshot month
if YEAR_ACTUAL in months_by_year:
    months_by_year[YEAR_ACTUAL] = [m for m in months_by_year[YEAR_ACTUAL] if m <= snapshot_month]

# weeks_by_ym — todas las semanas del mes, incluyendo vacías
weeks_by_ym = {}
for r in acum_weeks:
    ym = r['ym']
    if ym not in weeks_by_ym:
        weeks_by_ym[ym] = []
    weeks_by_ym[ym].append({
        'week': int(r['yw'].split('-W')[1]),
        'yw':   r['yw'],
        'netnew': r['netnew'],
        'acum':   r['acum']
    })

# ── ÍNDICE DIMENSIONAL para filtros combinables ──────────────────────────────
top_corps = df_hist['Corporativo'].value_counts().head(50).index.tolist()

CHANNELS_PROPIO  = ['DerbySoft','HBSI','Internal','Omnibees','Siteminder','SynXis','Travelclick']
CHANNELS_TERCERO = ['Expedia_tercero','HotelBeds Apitude','Hotel Unico V2','Travelgate']
CHANNEL_LABELS   = {
    'DerbySoft':'DerbySoft','HBSI':'HBSI','Internal':'Internal',
    'Omnibees':'Omnibees','Siteminder':'Siteminder','SynXis':'SynXis','Travelclick':'Travelclick',
    'Expedia_tercero':'Expedia','HotelBeds Apitude':'HotelBeds',
    'Hotel Unico V2':'Hotel Unico','Travelgate':'Travelgate'
}
ALL_CHANNELS = CHANNELS_PROPIO + CHANNELS_TERCERO

def ch_active(val):
    if val is None or str(val) in ('nan','-',''): return False
    try: return float(val) > 0
    except: return False

# Índice dimensional nivel hotel (para filtros de región/corp/tipo — sin duplicar por channel)
df_dim_hotel = df_hist[['yw','ym','Region_display','Corporativo','TipoHotel','Hotel']].copy()
df_dim_hotel['corp']    = df_dim_hotel['Corporativo'].where(df_dim_hotel['Corporativo'].isin(top_corps), 'Otros')
df_dim_hotel['ch_tipo'] = df_dim_hotel['TipoHotel'].map(
    {'sólo propio':'Solo Propio','Propio_con_tercero':'Hybrid'}).fillna('—')

dim_hotel_idx = (df_dim_hotel.groupby(['yw','ym','Region_display','corp','ch_tipo'])
                 .size().reset_index(name='n').sort_values('yw'))
dim_hotel_idx = dim_hotel_idx.rename(columns={'Region_display':'region'})

# Índice dimensional nivel channel (para filtro de channel específico)
rows_with_ch = []
for ch_col in ALL_CHANNELS:
    if ch_col not in df_hist.columns: continue
    sub = df_hist[df_hist[ch_col].apply(ch_active)][
        ['yw','ym','Region_display','Corporativo','TipoHotel','Hotel']].copy()
    sub['channel'] = CHANNEL_LABELS[ch_col]
    rows_with_ch.append(sub)

df_dim_raw = pd.concat(rows_with_ch, ignore_index=True)
df_dim_raw['corp']    = df_dim_raw['Corporativo'].where(df_dim_raw['Corporativo'].isin(top_corps), 'Otros')
df_dim_raw['ch_tipo'] = df_dim_raw['TipoHotel'].map(
    {'sólo propio':'Solo Propio','Propio_con_tercero':'Hybrid'}).fillna('—')

dim_index = (df_dim_raw.groupby(['yw','ym','Region_display','corp','ch_tipo','channel'])
             .size().reset_index(name='n').sort_values('yw'))
dim_index = dim_index.rename(columns={'Region_display':'region'})

hist_regions          = sorted(df_hist['Region_display'].unique().tolist())
hist_corps            = sorted(top_corps)
hist_channels_propio  = ['DerbySoft','HBSI','Internal','Omnibees','Siteminder','SynXis','Travelclick']
hist_channels_tercero = ['Expedia','HotelBeds','Hotel Unico','Travelgate']
hist_tipos            = ['Solo Propio','Hybrid']

dim_rows = dim_index.to_dict('records')
for r in dim_rows:
    for k, v in r.items():
        if hasattr(v, 'item'): r[k] = v.item()

dim_hotel_rows = dim_hotel_idx.to_dict('records')
for r in dim_hotel_rows:
    for k, v in r.items():
        if hasattr(v, 'item'): r[k] = v.item()

# Build ch_corp_map from full universe
ch_corp_map = {}
for ch_label, ch_col in CHANNEL_COL_MAP.items():
    if ch_col not in df.columns: continue
    sub = df[df[ch_col].apply(channel_active)][['Corporativo']].copy()
    corps = sorted(sub['Corporativo'].dropna().unique().tolist())
    if corps:
        ch_corp_map[ch_label] = [c for c in corps if c and c != 'Otros']
for ch_col, ch_label in CHANNEL_LABELS.items():
    if ch_col not in df.columns or ch_label in ch_corp_map: continue
    sub = df[df[ch_col].apply(channel_active)][['Corporativo']].copy()
    corps = sorted(sub['Corporativo'].dropna().unique().tolist())
    if corps:
        ch_corp_map[ch_label] = [c for c in corps if c and c != 'Otros']

class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer,)): return int(obj)
        if isinstance(obj, (np.floating,)): return float(obj)
        return super().default(obj)

hist_data = {
    'by_year':  acum_years,
    'by_month': acum_months,
    'by_week':  acum_weeks,
    'weeks_by_ym': weeks_by_ym,
    'years':    years_available,
    'months_by_year': months_by_year,
    'dim':        dim_rows,
    'dim_hotel':  dim_hotel_rows,
    'hist_regions':          hist_regions,
    'hist_corps':            hist_corps,
    'hist_channels_propio':  hist_channels_propio,
    'hist_channels_tercero': hist_channels_tercero,
    'hist_tipos':            hist_tipos,
    'ritmo_mensual': int(ritmo_nec) * 4,
    'ritmo_semanal': int(ritmo_nec),
    'target': int(TARGET_PROPIO),
    'actual': int(pp),
}

print(f"    Universo: {fmt_n(N)} | PP: {fmt_n(pp)} ({fmt_pct(pct_avance)} avance) | Gap: {fmt_n(gap)} | Ritmo: ~{fmt_n(ritmo_nec)}/sem")

# ─────────────────────────────────────────────
# 4. HTML
# ─────────────────────────────────────────────
print("[4/5] Generando HTML...")

# Colores corporativos para TipoHotel
# Solo Propio  → verde  #1A6B4A / #E1F5EE
# Hybrid       → violet #5C469C / #EDE8F7
# Third Party  → dark grey #333132 / #E8E6E3

CSS = """
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'Geist',system-ui,sans-serif;font-size:14px;line-height:1.55;
     background:#F8F4EC;color:#1A1917;-webkit-font-smoothing:antialiased;}
:root{
  --ink:#1A1917;--ink-soft:#4A4A4A;--ink-muted:#8A8377;
  --paper:#FFFFFF;--paper-soft:#F8F4EC;--rule:#C9C1B0;--rule-soft:#D9D2C1;
  --accent:#4FC3F4;--accent-soft:#E0F7FE;
  --green:#1A6B4A;--green-soft:#E1F5EE;
  --violet:#5C469C;--violet-soft:#EDE8F7;
  --dgrey:#333132;--dgrey-soft:#E8E6E3;
  --amber:#A86A1D;--amber-soft:#FFF4E0;
  --muted:#5F5E5A;--muted-soft:#F2EEE6;
  --maxw:1280px;
}
.dark-invert{filter:saturate(0) brightness(0);}
.shell{max-width:var(--maxw);margin:0 auto;padding:0 48px 80px;}
/* MASTHEAD */
.masthead-top-rule{height:3px;background:var(--ink);}
.masthead-sub{display:flex;justify-content:space-between;padding:7px 0;font-size:10px;
  font-weight:500;color:var(--ink-muted);letter-spacing:.09em;text-transform:uppercase;
  border-bottom:3px solid var(--ink);}
/* HERO */
.hero{padding:16px 0 8px;}
.hero h1{font-size:clamp(22px,2.6vw,34px);font-weight:700;letter-spacing:-.02em;margin-bottom:4px;}
.hero h1 .accent{color:var(--accent);}
.hero-sub{font-size:12px;color:var(--ink-muted);text-transform:uppercase;letter-spacing:.1em;margin-bottom:20px;}
/* KPI BAR */
.kpi-bar{display:grid;grid-template-columns:2.2fr 1.4fr 1.5fr 1.5fr;
  border:1px solid var(--rule);margin-bottom:36px;background:var(--paper);
  align-items:stretch;width:100%;}
.kpi-cell{padding:18px 16px;border-right:1px solid var(--rule);display:flex;
  flex-direction:column;min-width:0;}
.kpi-cell:last-child{border-right:none;}
.kpi-label{font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
  color:var(--ink-muted);margin-bottom:8px;}
.kpi-value{font-size:26px;font-weight:700;letter-spacing:-.02em;line-height:1;}
.kpi-sub{font-size:10px;color:var(--ink-muted);margin-top:4px;}
.kpi-split{display:grid;grid-template-columns:1fr 1fr;flex:1;min-width:0;}
.kpi-split-cell{padding:14px 14px;display:flex;flex-direction:column;min-width:0;}
.kpi-split-cell:first-child{border-right:1px solid var(--rule-soft);}
.kpi-split-label{font-size:9px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;margin-bottom:6px;}
.kpi-split-value{font-size:20px;font-weight:700;letter-spacing:-.02em;line-height:1;}
.kpi-split-sub{font-size:10px;color:var(--ink-muted);margin-top:3px;}
.kpi-triple{display:grid;grid-template-columns:1fr 1fr 1fr;flex:1;min-width:0;}
.kpi-triple-cell{padding:14px 10px;display:flex;flex-direction:column;min-width:0;}
.kpi-triple-cell:not(:last-child){border-right:1px solid var(--rule-soft);}
.kpi-triple-label{font-size:9px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  color:var(--ink-muted);margin-bottom:6px;}
.kpi-triple-value{font-size:17px;font-weight:700;letter-spacing:-.02em;line-height:1;}
.prog-bar{height:4px;background:var(--rule-soft);margin-top:6px;border-radius:2px;}
.prog-fill{height:100%;background:var(--accent);border-radius:2px;}
/* Tables */
.data-table{width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed;}
.tab-content{display:none;padding:14px 0;width:100%;}
.tab-content.on{display:block;}
.tabs-nav{display:flex;gap:0;border-bottom:1px solid var(--rule);flex-wrap:wrap;width:100%;}
/* DYNAMIC KPI NOTE */
.kpi-note{font-size:9px;color:var(--ink-muted);margin-top:8px;font-style:italic;
  padding:6px 8px;background:var(--paper-soft);border-left:2px solid var(--accent);display:none;}
.kpi-note.visible{display:block;}
/* SECTION */
.sec-head{display:flex;align-items:baseline;gap:10px;margin:32px 0 14px;
  padding-bottom:8px;border-bottom:1px solid var(--rule);}
.sec-title{font-size:13px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;}
.sec-badge{font-size:9px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  background:var(--accent-soft);color:#0277A8;padding:2px 8px;border-radius:2px;}
/* TABS */
.distrib-tab{display:none;}
.distrib-tab.on{display:block;}
.gap-pill.active{background:#FEE2E2!important;color:#C0392B!important;border-color:#C0392B!important;}
/* Metric pills use their own accent color when active */
.metric-pill.on{background:var(--pill-on-bg,#E0F4FD)!important;color:var(--pill-on-fg,#1A6B8A)!important;border-color:var(--pill-on-bd,#4FC3F4)!important;}
/* Active row filter */
tr.ud-filter-active td{font-weight:700;}
tr.ud-filter-active td:first-child::after{content:' ×';color:#C0392B;font-size:10px;cursor:pointer;}
/* Column visibility by active pill — table gets col-show-XX class */
/* col-show-pp no longer hides other columns — all columns always visible */
.col-show-pp-DISABLED .th-sp,
.col-show-all .th-pp,.col-show-all .td-pp,
.col-show-all .th-sp,.col-show-all .td-sp,
.col-show-all .th-hy,.col-show-all .td-hy,
.col-show-all .th-tp,.col-show-all .td-tp{display:table-cell;}
.col-show-all .th-pp,.col-show-all .td-pp,
.col-show-all .th-sp,.col-show-all .td-sp,
.col-show-all .th-hy,.col-show-all .td-hy,
.col-show-all .th-tp,.col-show-all .td-tp{display:table-cell;}
/* tp always visible */
/* col-show-sp no longer hides columns */
/* col-show-hy no longer hides columns */
/* col-show-tp no longer hides columns */
/* Active/selected row */
tr.ud-filter-active > td { background:var(--accent-soft,#E0F4FD) !important; }
tr.ud-filter-active > td:first-child { border-left:3px solid var(--accent,#4FC3F4); }
.data-table tbody tr:not(.global-row):hover > td { background:var(--paper-soft); cursor:pointer; }
/* chart area */.t-btn{padding:8px 14px;font-size:10px;font-weight:700;letter-spacing:.07em;
  text-transform:uppercase;color:var(--ink-muted);cursor:pointer;border:none;background:none;
  border-bottom:3px solid transparent;margin-bottom:-1px;transition:all .15s;
  font-family:inherit;white-space:nowrap;}
.t-btn:hover{color:var(--ink-soft);}
.t-btn.on{color:var(--accent);border-bottom-color:var(--accent);}
.data-table th{font-size:9px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  color:var(--ink-muted);padding:8px 10px;text-align:right;border-bottom:2px solid var(--ink);}
.data-table th:first-child{text-align:left;}
.data-table td{padding:9px 10px;border-bottom:1px solid var(--rule-soft);text-align:right;
  font-variant-numeric:tabular-nums;}
.data-table td:first-child{text-align:left;font-weight:600;color:var(--ink);}
.data-table tr.global-row td{font-weight:700;background:var(--paper-soft);border-bottom:2px solid var(--rule);}
.data-table tr:not(.global-row):hover td{background:rgba(79,195,244,.04);}
.th-pp{background:var(--accent-soft);color:#0277A8!important;}
.th-sp{background:var(--green-soft);color:var(--green)!important;}
.th-hy{background:var(--violet-soft);color:var(--violet)!important;}
.th-tp{background:var(--dgrey-soft);color:var(--dgrey)!important;}
.td-pp{color:#0277A8;font-weight:700;}
.td-sp{color:var(--green);font-weight:600;}
.td-hy{color:var(--violet);font-weight:600;}
.td-tp{color:var(--dgrey);}
.ud-th-active{border-bottom:2px solid currentColor!important;}
.td-pos{color:var(--green);font-size:11px;font-weight:700;text-align:right;}
tr.sel > td{background:var(--accent-soft,#E0F4FD)!important;font-weight:700;}
tr.sel > td:first-child{border-left:3px solid var(--accent,#4FC3F4);}
.td-neg{color:#C0392B;font-size:11px;font-weight:700;text-align:right;}
.td-pp{color:#4FC3F4;font-weight:600;}
.td-tp{color:var(--ink);font-weight:400;}
.td-tot{color:var(--ink);font-weight:400;font-variant-numeric:tabular-nums;}
/* PCT BAR */
.pct-wrap{display:flex;align-items:center;gap:6px;justify-content:flex-end;}
.pct-bar{height:3px;background:var(--rule-soft);width:50px;border-radius:2px;flex-shrink:0;}
.pct-fill{height:100%;border-radius:2px;}
.pct-val{font-size:11px;font-weight:700;min-width:38px;text-align:right;}
/* SEARCH + AUTOCOMPLETE */
.sb-wrap{display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;position:relative;}
.sb-input{border:1px solid var(--rule);background:var(--paper);padding:7px 12px;
  font-size:11px;font-family:inherit;color:var(--ink);border-radius:3px;width:240px;}
.sb-input:focus{outline:none;border-color:var(--accent);}
.sb-select{border:1px solid var(--rule);background:var(--paper);padding:7px 10px;
  font-size:11px;font-family:inherit;color:var(--ink);border-radius:3px;cursor:pointer;}
.sb-select:focus{outline:none;border-color:var(--accent);}
.sb-count{font-size:10px;color:var(--ink-muted);}
.autocomplete-list{position:absolute;top:calc(100% + 2px);left:0;z-index:100;
  background:var(--paper);border:1px solid var(--rule);border-radius:3px;
  max-height:200px;overflow-y:auto;min-width:240px;box-shadow:0 4px 12px rgba(0,0,0,.08);}
.autocomplete-item{padding:7px 12px;font-size:11px;cursor:pointer;
  border-bottom:1px solid var(--rule-soft);}
.autocomplete-item:last-child{border-bottom:none;}
.autocomplete-item:hover,.autocomplete-item.active{background:var(--accent-soft);}
.autocomplete-item mark{background:transparent;color:var(--accent);font-weight:700;}
.more-hint{text-align:center;padding:10px;color:var(--ink-muted);font-size:10px;
  font-style:italic;border-top:1px dashed var(--rule-soft);}
/* CHANNEL TABLES */
.ch-block-label{font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
  color:var(--ink-muted);margin:14px 0 0;padding:6px 0;border-bottom:1px solid var(--rule-soft);}
.ch-table{width:100%;border-collapse:collapse;font-size:11px;}
.ch-table th{font-size:9px;font-weight:700;letter-spacing:.09em;text-transform:uppercase;
  color:var(--ink-muted);padding:7px 8px;text-align:right;border-bottom:1px solid var(--rule);}
.ch-table th:first-child{text-align:left;}
.ch-table td{padding:7px 8px;border-bottom:1px solid var(--rule-soft);text-align:right;
  font-variant-numeric:tabular-nums;}
.ch-table td:first-child{text-align:left;font-weight:600;color:var(--ink);}
.ch-table tr:hover td{background:rgba(79,195,244,.04);}
.ch-table td.dim{color:var(--ink-muted);font-weight:400;}
.ch-table tr.residual td{opacity:.55;}
.int-wrap{display:flex;flex-direction:column;align-items:flex-end;gap:2px;}
.int-bar{height:3px;background:var(--rule-soft);border-radius:2px;width:52px;}
.int-fill{height:100%;background:var(--accent);border-radius:2px;opacity:.8;}
.int-val{font-size:9px;color:var(--ink-muted);}
/* HISTORICAL */
.hist-controls{display:flex;justify-content:space-between;align-items:flex-start;
  margin-bottom:14px;gap:12px;flex-wrap:wrap;}
.drill-toggle{display:inline-flex;border:1.5px solid var(--ink);border-radius:4px;overflow:hidden;}
.dt-btn{padding:6px 16px;font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  cursor:pointer;border:none;background:transparent;font-family:inherit;color:var(--ink-muted);transition:all .15s;}
.dt-btn+.dt-btn{border-left:1.5px solid var(--ink);}
.dt-btn.on{background:var(--ink);color:#fff;}
.filters-row{display:flex;gap:8px;align-items:flex-start;flex-wrap:wrap;margin-top:10px;}
.f-group{display:flex;flex-direction:column;gap:0;}
.f-label{font-size:9px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  color:var(--ink-muted);white-space:nowrap;margin-bottom:3px;display:block;}
.f-select,.f-input{
  border:1.5px solid var(--rule);
  background:var(--paper);
  padding:6px 28px 6px 10px;
  font-size:10px;font-weight:600;font-family:inherit;
  color:var(--ink);border-radius:4px;
  cursor:pointer;
  appearance:none;-webkit-appearance:none;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M1 1l4 4 4-4' stroke='%238A8377' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");
  background-repeat:no-repeat;
  background-position:right 9px center;
  transition:border-color .15s, box-shadow .15s;
  max-width:160px;
}
.f-input{width:130px;background-image:none;padding:6px 10px;}
.f-select:hover{border-color:var(--accent);}
.f-select:not([disabled]){border-color:var(--accent)!important;color:var(--accent)!important;}
.f-select:focus,.f-input:focus{outline:none;border-color:var(--accent);box-shadow:0 0 0 2px rgba(79,195,244,.15);}
/* Pills — Región */
.pills-wrap{display:flex;gap:4px;flex-wrap:wrap;}
.pill{padding:4px 10px;font-size:9px;font-weight:700;letter-spacing:.07em;text-transform:uppercase;
  border:1px solid var(--rule);border-radius:20px;cursor:pointer;background:var(--paper);
  color:var(--ink-muted);transition:all .12s;white-space:nowrap;font-family:inherit;}
.pill:hover{border-color:var(--accent);color:var(--accent);}
.pill.on{background:var(--ink);color:#fff;border-color:var(--ink);}
/* Toggle tipo producto */
.tipo-toggle{display:inline-flex;border:1px solid var(--rule);border-radius:4px;overflow:hidden;}
.tipo-btn{padding:5px 10px;font-size:9px;font-weight:700;letter-spacing:.07em;text-transform:uppercase;
  cursor:pointer;border:none;background:transparent;font-family:inherit;color:var(--ink-muted);
  transition:all .12s;white-space:nowrap;}
.tipo-btn+.tipo-btn{border-left:1px solid var(--rule);}
.tipo-btn.on{background:var(--ink);color:#fff;}
/* Limpiar link */
.clear-link{font-size:9px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;
  color:var(--ink-muted);cursor:pointer;text-decoration:none;border:none;background:none;
  font-family:inherit;padding:0;opacity:.7;}
.clear-link:hover{opacity:1;color:var(--accent);}
.breadcrumb{display:flex;align-items:center;gap:6px;margin-bottom:10px;
  font-size:11px;color:var(--ink-muted);min-height:20px;}
.bc-link{cursor:pointer;color:var(--accent);font-weight:600;}
.bc-link:hover{text-decoration:underline;}
.bc-sep{color:var(--ink-muted);}
.bc-cur{color:var(--ink);font-weight:600;}
.chart-area{background:transparent;border:1px solid var(--rule);padding:16px 16px 10px;border-radius:2px;}
.chart-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;}
.chart-title{font-size:11px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;
  color:var(--ink-muted);}
.chart-legend{display:flex;gap:14px;}
.leg-item{display:flex;align-items:center;gap:5px;font-size:10px;color:var(--ink-muted);}
.leg-line{width:14px;height:2px;border-radius:1px;}
.leg-cyan{background:var(--accent);}
.leg-amber{background:#FCB000;}
.leg-dashed{background:repeating-linear-gradient(90deg,#FCB000 0,#FCB000 4px,transparent 4px,transparent 7px);height:1.5px;}
canvas{display:block;}
"""

JS = f"""
const HIST = {json.dumps(hist_data, cls=NpEncoder)};
const CORP_DATA     = {json.dumps(corp_json, cls=NpEncoder)};
const DEST_DATA     = {json.dumps(dest_json, cls=NpEncoder)};
const CORP_MKT_DATA = {json.dumps(corp_mkt_json, cls=NpEncoder)};
const CH_DRILL_DATA = {json.dumps(ch_drill_data, cls=NpEncoder)};
const CH_CORP_MAP = {json.dumps(ch_corp_map, cls=NpEncoder)};
const CH_DATA = {{
  'todos': {{ propio:{json.dumps(ch_propio, cls=NpEncoder)}, tercero:{json.dumps(ch_tercero, cls=NpEncoder)}, maxAvg:{max_avg_ctr} }},
  'sp':    {{ propio:{json.dumps(ch_propio_sp, cls=NpEncoder)}, tercero:{json.dumps(ch_tercero_sp, cls=NpEncoder)}, maxAvg:{max_avg_ctr_sp} }},
  'hy':    {{ propio:{json.dumps(ch_propio_hy, cls=NpEncoder)}, tercero:{json.dumps(ch_tercero_hy, cls=NpEncoder)}, maxAvg:{max_avg_ctr_hy} }},
}};
const DEST_MKT_DATA = {json.dumps(dest_mkt_json, cls=NpEncoder)};
const MARKET_TOTAL  = {market_total};
const MARKET_SHARE  = {market_share:.2f};
const GLOBAL_PP   = {int(pp)};
const GLOBAL_TERC = {int(solo_terc)};
const GLOBAL_N    = {int(N)};
const TARGET      = {int(TARGET_PROPIO)};
const GAP_GLOBAL  = {int(gap)};

/* ── TABS ── */
function showTab(groupId, id, btn) {{
  const group = document.getElementById(groupId);
  if (!group) return;
  // Only direct child tab-content elements to avoid affecting nested tabs
  group.querySelectorAll(':scope > .tab-content, :scope > div > .tab-content').forEach(t=>t.classList.remove('on'));
  group.querySelectorAll(':scope > .tabs-nav .t-btn').forEach(b=>b.classList.remove('on'));
  const target = document.getElementById(groupId+'-'+id);
  if (target) target.classList.add('on');
  if (btn) btn.classList.add('on');
}}

/* ── DYNAMIC CARD UPDATE ── */
function updateCards(filter) {{
  // filter: {{type:'all'}} | {{type:'corp',name:'...'}} | {{type:'dest',name:'...'}} | {{type:'region',name:'...'}}
  let subset, label;
  if (filter.type === 'all') {{
    document.getElementById('kpi-note').classList.remove('visible');
    document.getElementById('kpi-note').textContent = '';
    renderCardValues(GLOBAL_PP, GLOBAL_TERC, GLOBAL_N);
    return;
  }}
  if (filter.type === 'corp') {{
    subset = CORP_DATA.find(r => r.Corporativo === filter.name);
    if (!subset) return;
    label = 'Corporativo: '+filter.name;
    renderCardValues(subset.prod_propio, subset.solo_tercero, subset.total);
  }} else if (filter.type === 'dest') {{
    subset = DEST_DATA.find(r => r.Destino === filter.name);
    if (!subset) return;
    label = 'Destino: '+filter.name;
    renderCardValues(subset.prod_propio, subset.solo_tercero, subset.total);
  }} else if (filter.type === 'region') {{
    // sum from CORP_DATA by region — use dest_data grouped
    let pp=0, tp=0, tot=0;
    DEST_DATA.forEach(r => {{ if(r.Region_display===filter.name){{pp+=r.prod_propio;tp+=r.solo_tercero;tot+=r.total;}} }});
    label = 'Región: '+filter.name;
    renderCardValues(pp, tp, tot);
  }}
  const note = document.getElementById('kpi-note');
  note.textContent = '▸ Filtrando por '+label+' — target global 70K como referencia';
  note.classList.add('visible');
}}

function renderCardValues(pp, tp, tot) {{
  const pctPP  = tot>0 ? (pp/tot*100).toFixed(1)+'%' : '—';
  const pctTP  = tot>0 ? (tp/tot*100).toFixed(1)+'%' : '—';
  document.getElementById('card-total').textContent   = tot.toLocaleString('es-MX');
  document.getElementById('card-pp').textContent      = pp.toLocaleString('es-MX');
  document.getElementById('card-pp-pct').textContent  = 'Solo Propio + Hybrid · '+pctPP;
  document.getElementById('card-tp').textContent      = tp.toLocaleString('es-MX');
  document.getElementById('card-tp-pct').textContent  = 'Sin canal propio · '+pctTP;
  // Gap vs TARGET — Opción C: target global fijo
  const gap = TARGET - GLOBAL_PP;  // siempre vs global
  document.getElementById('card-gap').textContent  = gap.toLocaleString('es-MX');
  document.getElementById('card-avance').textContent = (GLOBAL_PP/TARGET*100).toFixed(1)+'%';
  document.getElementById('prog-fill').style.width = Math.min(100, GLOBAL_PP/TARGET*100).toFixed(1)+'%';
}}

/* ── AUTOCOMPLETE FACTORY ── */
function makeAutocomplete(inputId, listId, data, labelKey, onSelect) {{
  const inp  = document.getElementById(inputId);
  const list = document.getElementById(listId);
  if (!inp || !list) return;
  let activeIdx = -1;

  function norm(s) {{ return s.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,''); }}
  function highlight(text, q) {{
    const ni = norm(text).indexOf(norm(q));
    if (ni < 0 || !q) return text;
    return text.slice(0,ni)+'<mark>'+text.slice(ni,ni+q.length)+'</mark>'+text.slice(ni+q.length);
  }}

  function buildList(q) {{
    if (!q) {{ list.innerHTML=''; list.style.display='none'; return; }}
    const matches = data.filter(r => norm(r[labelKey]).includes(norm(q))).slice(0,10);
    if (!matches.length) {{ list.innerHTML=''; list.style.display='none'; return; }}
    list.innerHTML = matches.map((r,i) =>
      `<div class="autocomplete-item" data-idx="${{i}}" data-val="${{r[labelKey].replace(/"/g,'&quot;')}}">`+
      highlight(r[labelKey], q)+`</div>`).join('');
    list.style.display = 'block';
    activeIdx = -1;
    list.querySelectorAll('.autocomplete-item').forEach(el => {{
      el.addEventListener('mousedown', e => {{
        e.preventDefault();
        selectItem(el.dataset.val);
      }});
    }});
  }}

  function selectItem(val) {{
    inp.value = val;
    list.innerHTML = ''; list.style.display = 'none';
    filterTable(inp.dataset.tbodyId, val, 0);
    if (onSelect) onSelect(val);
  }}

  inp.addEventListener('input', () => {{ buildList(inp.value.trim()); filterTable(inp.dataset.tbodyId, inp.value, 0); }});
  inp.addEventListener('keydown', e => {{
    const items = list.querySelectorAll('.autocomplete-item');
    if (e.key==='ArrowDown') {{ activeIdx=Math.min(activeIdx+1,items.length-1); items.forEach((el,i)=>el.classList.toggle('active',i===activeIdx)); e.preventDefault(); }}
    else if (e.key==='ArrowUp') {{ activeIdx=Math.max(activeIdx-1,0); items.forEach((el,i)=>el.classList.toggle('active',i===activeIdx)); e.preventDefault(); }}
    else if (e.key==='Enter' && activeIdx>=0) {{ selectItem(items[activeIdx].dataset.val); }}
    else if (e.key==='Escape') {{ list.innerHTML=''; list.style.display='none'; }}
  }});
  inp.addEventListener('blur', () => setTimeout(()=>{{ list.style.display='none'; }}, 150));
}}

function filterGapTable(tbodyId, q, colIdx, verMasRowId) {{
  const rows = document.getElementById(tbodyId).querySelectorAll('tr[data-row-idx]');
  const ql = q.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').trim();
  const verMasRow = verMasRowId ? document.getElementById(verMasRowId) : null;
  rows.forEach(r => {{
    const cell = r.cells[colIdx];
    if (!cell) return;
    const t = cell.textContent.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'');
    if (!ql) {{
      const idx = parseInt(r.dataset.rowIdx||'0');
      r.style.display = idx < 10 ? '' : 'none';
    }} else {{
      r.style.display = t.includes(ql) ? '' : 'none';
    }}
  }});
  if (verMasRow) verMasRow.style.display = ql ? 'none' : '';
}}

function filterGapDestRegion(val, tbodyId, verMasRowId) {{
  const rows = document.getElementById(tbodyId).querySelectorAll('tr[data-row-idx]');
  rows.forEach(r => {{
    const cell = r.cells[1];
    if (!cell) return;
    const idx = parseInt(r.dataset.rowIdx||'0');
    if (!val) {{
      r.style.display = idx < 10 ? '' : 'none';
    }} else {{
      r.style.display = cell.textContent.trim()===val ? '' : 'none';
    }}
  }});
  const vmr = document.getElementById(verMasRowId);
  if (vmr) vmr.style.display = val ? 'none' : '';
}}

/* ── SINGLE TABLE CONTROLLER ── */
let udCurrentContent = 'pp';

// ── UNIFIED BADGE SYSTEM ──
// Single function renders all active selection badges in ud-active-badge-container
// Sources: hFRegion, hFChannel, hFCorp, udCurrentContent, udActiveFilter (row click)
function udSyncBadges() {{
  const container = document.getElementById('ud-active-badge-container');
  if (!container) return;
  container.innerHTML = '';

  function makePill(label, onX) {{
    const pill = document.createElement('span');
    pill.style.cssText = 'display:inline-flex;align-items:center;gap:4px;padding:3px 10px;'
      + 'border:1.5px solid #4FC3F4;border-radius:20px;font-size:9px;font-weight:700;'
      + 'letter-spacing:.06em;text-transform:uppercase;color:#4FC3F4;'
      + 'background:#E0F4FD;white-space:nowrap;';
    const x = document.createElement('span');
    x.textContent = '\u00d7';
    x.style.cssText = 'cursor:pointer;font-size:11px;line-height:1;margin-left:2px;color:#4FC3F4;';
    x.addEventListener('click', onX);
    pill.textContent = label;
    pill.appendChild(x);
    return pill;
  }}

  // Collect all active values (no duplicates)
  // Priority: h* filter vars first, then udActiveFilter only if different
  const entries = [];
  const seen = new Set();

  function addEntry(label, onX) {{
    const key = label.toUpperCase();
    if (seen.has(key)) return;
    seen.add(key);
    entries.push({{ label, onX }});
  }}

  if (hFRegion) addEntry(hFRegion, function() {{
    hFRegion = '';
    document.querySelectorAll('#hf-region-pills .pill').forEach(p=>p.classList.remove('on'));
    const t = document.querySelector('#hf-region-pills .pill[data-region=""]');
    if(t) t.classList.add('on');
    // Also deactivate table row if it was the same value
    if (udActiveFilter && udActiveFilter.type==='region') {{
      udActiveFilter = {{type:null, value:null}};
      document.querySelectorAll('#ud-tbody tr').forEach(r=>r.classList.remove('ud-filter-active'));
    }}
    hRenderActivePills(); hApplyFilter();
  }});

  if (hFChannel) addEntry(hFChannel, function() {{
    hFChannel=''; document.getElementById('hf-channel').value='';
    hUpdateComboStyle('hf-channel'); hFilterCorpByChannel('');
    if (udActiveFilter && udActiveFilter.type==='channel') {{
      udActiveFilter = {{type:null, value:null}};
      document.querySelectorAll('#ud-tbody tr').forEach(r=>r.classList.remove('ud-filter-active'));
    }}
    hRenderActivePills(); hApplyFilter();
  }});

  if (hFCorp) addEntry(hFCorp, function() {{
    hFCorp=''; document.getElementById('hf-corp').value='';
    hUpdateComboStyle('hf-corp');
    if (udActiveFilter && udActiveFilter.type==='corp') {{
      udActiveFilter = {{type:null, value:null}};
      document.querySelectorAll('#ud-tbody tr').forEach(r=>r.classList.remove('ud-filter-active'));
    }}
    hRenderActivePills(); hApplyFilter();
  }});

  if (udActiveFilter && udActiveFilter.value) {{
    addEntry(udActiveFilter.value, function() {{
      udActiveFilter = {{type:null, value:null}};
      document.querySelectorAll('#ud-tbody tr').forEach(r=>r.classList.remove('ud-filter-active'));
      hFRegion=''; hFCorp=''; hApplyFilter();
      udSyncBadges();
    }});
  }}

  entries.forEach(e => container.appendChild(makePill(e.label, e.onX)));
}}

// Legacy no-ops for compatibility
function udRenderMetricPill(id, label) {{ udSyncBadges(); }}

function udContent(id, btn) {{
  // Second click on active metric pill → reset to Todos
  if (id !== 'all' && id !== 'gap' && btn && btn.classList.contains('on')) {{
    udContent('all', document.querySelector('.distrib-pills .pill[data-col="all"]'));
    return;
  }}
  if (id === 'all') {{
    document.querySelectorAll('.distrib-pills .pill').forEach(p=>p.classList.remove('on'));
    if (btn) btn.classList.add('on');
    hFTipo = ''; if (typeof hRender === 'function') hRender();
    udCurrentContent = 'all';
    udRenderMetricPill(null, null);
    document.querySelectorAll('#ud-tbody').forEach(tb => {{
      const tbl = tb.closest('table');
      if (tbl) tbl.querySelectorAll('th').forEach(th => {{ th.classList.remove('ud-th-active'); th.style.color=''; }});
    }});
    const tbl = document.querySelector('#ud-tbody')?.closest('table');
    if (tbl) {{
      tbl.className = tbl.className.replace(/\bcol-show-\S+/g,'').trim();
      tbl.classList.add('col-show-all');
    }}
    const gb = document.getElementById('btn-gap');
    if (gb && gb.classList.contains('active')) {{
      gb.classList.remove('active');
      document.getElementById('ud-gap-content').style.display = 'none';
      document.getElementById('ud-main-content').style.display = '';
    }}
    return;
  }}
  // Handle gap/sin contrat directa
  if (id === 'gap') {{
    document.querySelectorAll('.distrib-pills .pill').forEach(p=>p.classList.remove('on'));
    if (btn) btn.classList.add('on');
    udToggleGap(btn);
    hFTipo = ''; if (typeof hRender === 'function') hRender();
    return;
  }}
  // Sync tipo filter to historical chart
  if (id === 'sp') {{ hFTipo = 'Solo Propio'; }}
  else if (id === 'hy') {{ hFTipo = 'Hybrid'; }}
  else {{ hFTipo = ''; }}
  if (typeof hRender === 'function') hRender();
  udCurrentContent = id;
  document.querySelectorAll('.distrib-pills .pill').forEach(p=>p.classList.remove('on'));
  if (btn) btn.classList.add('on');
  const gapBtn = document.getElementById('btn-gap');
  if (gapBtn && gapBtn.classList.contains('active')) {{
    gapBtn.classList.remove('active');
    document.getElementById('ud-gap-content').style.display = 'none';
    document.getElementById('ud-main-content').style.display = '';
  }}
  // Apply a class to the table so CSS can hide columns
  const tbl = document.querySelector('#ud-tbody')?.closest('table');
  if (tbl) {{
    tbl.className = tbl.className.replace(/\bcol-show-\S+/g, '').trim();
    tbl.classList.add('col-show-' + (id||'all'));
  }}
  // Render metric pill badge
  const metricLabels = {{pp:'Producto Propio', sp:'Solo Propio', hy:'Hybrid'}};
  udRenderMetricPill(id, metricLabels[id] || null);
  // Highlight active column header
  const colColors = {{pp:'#1A6B4A', sp:'#1A6B4A', hy:'#1A6B4A'}};
  document.querySelectorAll('#ud-tbody').forEach(tb => {{
    const tbl = tb.closest('table');
    if (!tbl) return;
    tbl.querySelectorAll('th').forEach(th => {{
      th.classList.remove('ud-th-active');
      th.style.color = '';
    }});
    const thMap = {{pp:'.th-pp', sp:'.th-sp', hy:'.th-hy', tp:'.th-tp'}};
    const sel = thMap[id];
    if (sel && colColors[id]) {{
      const th = tbl.querySelector(sel);
      if (th) {{ th.classList.add('ud-th-active'); th.style.color = colColors[id]; }}
    }}
  }});
  // Update % header label
  const pctLabel = {{pp:'% P.Propio', sp:'% Solo P.', hy:'% Hybrid', tp:'% Third P.'}};
  document.querySelectorAll('.th-pct-label').forEach(el => el.textContent = pctLabel[id]||'% Propio');
}}

function udToggleGap(btn) {{
  const gapDiv  = document.getElementById('ud-gap-content');
  const mainDiv = document.getElementById('ud-main-content');
  const isActive = btn ? btn.classList.contains('active') : false;
  if (isActive) {{
    if (btn) btn.classList.remove('active');
    gapDiv.style.display  = 'none';
    mainDiv.style.display = '';
    udRenderMetricPill(null, null);
  }} else {{
    if (btn) btn.classList.add('active');
    gapDiv.style.display  = '';
    mainDiv.style.display = 'none';
    document.getElementById('ud-ch-content').style.display = 'none';
    gapSyncDim();
    udRenderMetricPill('gap', 'Sin Contrat. Directa');
  }}
}}

function gapSyncDim() {{
  // Get active filter (supports both singular udActiveFilter and array udActiveFilters)
  const _af = (typeof udActiveFilters !== 'undefined' && udActiveFilters.length)
    ? udActiveFilters
    : (udActiveFilter && udActiveFilter.type ? [udActiveFilter] : []);
  const activeCorp = _af.filter(f=>f.type==='corp').map(f=>f.value);
  const activeReg  = _af.filter(f=>f.type==='region').map(f=>f.value);

  document.querySelectorAll('.gap-reg-row').forEach(r => {{
    if (udDim !== 'reg') {{ r.style.display = 'none'; return; }}
    const reg = r.querySelector('td strong')?.textContent || '';
    r.style.display = (!activeReg.length || activeReg.includes(reg)) ? '' : 'none';
  }});

  document.querySelectorAll('.gap-corp-row').forEach(r => {{
    if (r.id==='gap-corp-ver-mas2') {{ r.style.display = udDim==='corp' ? '' : 'none'; return; }}
    if (udDim !== 'corp') {{ r.style.display = 'none'; return; }}
    const idx  = parseInt(r.dataset.rowIdx??'999');
    const corp = r.querySelector('td strong')?.textContent || '';
    if (activeCorp.length) {{
      r.style.display = activeCorp.includes(corp) ? '' : 'none';
    }} else {{
      r.style.display = idx < 10 ? '' : 'none';
    }}
  }});

  document.querySelectorAll('.gap-dest-row').forEach(r => {{
    if (r.id==='gap-dest-ver-mas2') {{ r.style.display = udDim==='dest' ? '' : 'none'; return; }}
    if (udDim !== 'dest') {{ r.style.display = 'none'; return; }}
    const idx    = parseInt(r.dataset.rowIdx??'999');
    const rowReg = r.dataset.region || '';
    if (activeReg.length) {{
      r.style.display = activeReg.includes(rowReg) ? '' : 'none';
    }} else {{
      r.style.display = idx < 10 ? '' : 'none';
    }}
  }});

  const th = document.getElementById('gap-dim-th');
  if (th) th.textContent = udDim==='reg' ? 'Región' : udDim==='corp' ? 'Corporativo' : 'Destino';
  const gfr = document.getElementById('gap-f-region');
  if (gfr) gfr.style.display = udDim==='dest' ? '' : 'none';
  const gs = document.getElementById('gap-search');
  if (gs) gs.style.display = udDim!=='reg' ? '' : 'none';
}}

function gapToggleCorp(btn) {{
  const open=btn.dataset.open==='1';
  document.querySelectorAll('.gap-corp-row[data-row-idx]').forEach(r=>{{
    const idx=parseInt(r.dataset.rowIdx); if(idx>=10&&idx<20) r.style.display=open?'none':'';
  }});
  btn.dataset.open=open?'0':'1'; btn.textContent=open?'Ver 10 más':'Ver menos';
}}

function gapToggleDest(btn) {{
  const open=btn.dataset.open==='1';
  document.querySelectorAll('.gap-dest-row[data-row-idx]').forEach(r=>{{
    const idx=parseInt(r.dataset.rowIdx); if(idx>=10&&idx<20) r.style.display=open?'none':'';
  }});
  btn.dataset.open=open?'0':'1'; btn.textContent=open?'Ver 10 más':'Ver menos';
}}

function gapFilter() {{
  const q=document.getElementById('gap-search').value.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').trim();
  const reg=document.getElementById('gap-f-region').value;
  if (udDim==='corp') {{
    document.querySelectorAll('.gap-corp-row[data-row-idx]').forEach(r=>{{
      const txt=r.cells[0]?.textContent.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'')||'';
      const idx=parseInt(r.dataset.rowIdx);
      r.style.display=!q?(idx<10?'':'none'):(txt.includes(q)?'':'none');
    }});
    const vm=document.getElementById('gap-corp-ver-mas2'); if(vm) vm.style.display=q?'none':'';
  }} else if (udDim==='dest') {{
    document.querySelectorAll('.gap-dest-row[data-row-idx]').forEach(r=>{{
      const rReg=r.dataset.region||'';
      const txt=r.cells[0]?.textContent.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'')||'';
      const idx=parseInt(r.dataset.rowIdx);
      r.style.display=((!reg||rReg===reg)&&(!q||txt.includes(q)))?(!q&&!reg?(idx<10?'':'none'):''):'none';
    }});
    const vm=document.getElementById('gap-dest-ver-mas2'); if(vm) vm.style.display=(q||reg)?'none':'';
  }}
}}

let udDim = 'reg';
let udActiveFilters = [];

/* ── PERSISTENT ROW FILTER ── */
let udActiveFilter = {{type:null, value:null}};

function udRowClick(type, value, el) {{
  const tbody = document.getElementById('ud-tbody');
  // Toggle off if clicking the same
  if (udActiveFilter.type===type && udActiveFilter.value===value) {{
    udActiveFilter = {{type:null, value:null}};
    tbody.querySelectorAll('tr').forEach(r=>r.classList.remove('ud-filter-active'));
    udShowActiveFilterBadge(null);
    return;
  }}
  // Set new filter
  udActiveFilter = {{type, value}};
  tbody.querySelectorAll('tr').forEach(r=>r.classList.remove('ud-filter-active'));
  el.classList.add('ud-filter-active');
  udShowActiveFilterBadge(value);
  // Apply to chart if relevant
  if (type==='region')  {{ hFRegion=value; hApplyFilter(); }}
  else if (type==='corp') {{ hFCorp=value;   hApplyFilter(); }}
  else {{ hFRegion=''; hFCorp=''; hApplyFilter(); }}
  // Re-sync gap view if active
  const _gd = document.getElementById('ud-gap-content');
  if (_gd && _gd.style.display !== 'none') gapSyncDim();
  udSyncBadges();
}}

function udShowActiveFilterBadge(value) {{
  let badge = document.getElementById('ud-active-filter-badge');
  if (!badge) {{
    badge = document.createElement('span');
    badge.id = 'ud-active-filter-badge';
    badge.style.cssText = 'display:inline-flex;align-items:center;gap:4px;padding:3px 10px;'
      + 'border:1px solid var(--accent,#4FC3F4);border-radius:20px;font-size:9px;font-weight:700;'
      + 'letter-spacing:.06em;text-transform:uppercase;color:var(--accent,#4FC3F4);'
      + 'background:var(--accent-soft,#E0F4FD);cursor:pointer;margin-left:6px;';
    badge.onclick = function() {{
      udActiveFilter = {{type:null, value:null}};
      document.querySelectorAll('#ud-tbody tr').forEach(r=>r.classList.remove('ud-filter-active'));
      udShowActiveFilterBadge(null);
      hFRegion=''; hFCorp=''; hApplyFilter();
    }};
    document.getElementById('ud-dim-pills').appendChild(badge);
  }}
  if (value) {{
    badge.innerHTML = value + ' <span style="font-size:11px;line-height:1;">×</span>';
    badge.style.display = 'inline-flex';
  }} else {{
    badge.style.display = 'none';
  }}
}}

function udSetDim(dim, btn) {{
  // Don't clear active filters when switching dimension — preserve cross-dim context
  // Only reset hFRegion/hFCorp from current udActiveFilters
  const lastReg  = udActiveFilters.filter(f=>f.type==='region').map(f=>f.value);
  const lastCorp = udActiveFilters.filter(f=>f.type==='corp').map(f=>f.value);
  hFRegion = lastReg.length  === 1 ? lastReg[0]  : '';
  hFCorp   = lastCorp.length === 1 ? lastCorp[0] : '';
  if (dim === 'gap') {{
    document.getElementById('ud-main-content').style.display='none';
    document.getElementById('ud-gap-content').style.display='';
    document.getElementById('ud-ch-content').style.display='none';
    const mp=document.getElementById('ud-metric-pills'); if(mp) mp.style.display='none';
    document.getElementById('ud-f-region').style.display='none';
    document.getElementById('ud-search').style.display='none';
    document.querySelectorAll('#ud-dim-pills .pill').forEach(p=>p.classList.remove('on'));
    document.querySelectorAll('.distrib-pills .pill').forEach(p=>p.classList.remove('on'));
    if (btn) btn.classList.add('on');
    return;
  }}
  if (dim === 'ch') {{
    document.getElementById('ud-main-content').style.display='none';
    document.getElementById('ud-gap-content').style.display='none';
    document.getElementById('ud-ch-content').style.display='';
    const mp=document.getElementById('ud-metric-pills'); if(mp) mp.style.display='none';
    // In channel view, show only h* filter badges (not row-click dimension badge)
    const bc = document.getElementById('ud-active-badge-container');
    if (bc) bc.style.display='flex';
    udSyncBadges();
    document.getElementById('ud-f-region').style.display='none';
    document.getElementById('ud-search').style.display='none';
    document.querySelectorAll('#ud-dim-pills .pill').forEach(p=>p.classList.remove('on'));
    if (btn) btn.classList.add('on');
    // If there's an active filter, show it in the channel filter area
    udApplyFilterToChannel();
    return;
  }}
  udDim = dim;
  document.getElementById('ud-main-content').style.display='';
  document.getElementById('ud-gap-content').style.display='none';
  document.getElementById('ud-ch-content').style.display='none';
  const mp=document.getElementById('ud-metric-pills'); if(mp) mp.style.display='';
  document.querySelectorAll('#ud-dim-pills .pill').forEach(p=>p.classList.remove('on'));
  if (btn) btn.classList.add('on');
  // Show first 10 rows of active dimension
  document.querySelectorAll('.ud-reg-row').forEach(r => r.style.display = dim==='reg' ? '' : 'none');
  document.querySelectorAll('.ud-corp-row').forEach(r => {{
    if (r.id === 'ud-corp-ver-mas') {{ r.style.display = dim==='corp' ? '' : 'none'; return; }}
    const idx = parseInt(r.dataset.rowIdx ?? '999');
    r.style.display = dim==='corp' ? (idx < 10 ? '' : 'none') : 'none';
  }});
  document.querySelectorAll('.ud-dest-row').forEach(r => {{
    if (r.id === 'ud-dest-ver-mas') {{ r.style.display = dim==='dest' ? '' : 'none'; return; }}
    const idx = parseInt(r.dataset.rowIdx ?? '999');
    r.style.display = dim==='dest' ? (idx < 10 ? '' : 'none') : 'none';
  }});
  // Reset ver-mas buttons
  ['ud-corp-ver-mas','ud-dest-ver-mas'].forEach(id => {{
    const vm=document.getElementById(id); if(!vm) return;
    const b=vm.querySelector('button'); if(!b) return;
    b.dataset.open='0'; b.textContent='Ver 10 más';
  }});
  // Update header
  const thEl=document.getElementById('ud-dim-th');
  if(thEl) thEl.textContent = dim==='reg' ? 'Región' : dim==='corp' ? 'Corporativo' : 'Destino';
  // For Destino: no Solo P./Hybrid breakdown — auto-switch to PP and hide those cols
  const tbl = document.querySelector('#ud-tbody')?.closest('table');
  if (tbl && dim==='dest') {{
    tbl.className = tbl.className.replace(/\bcol-show-\S+/g,'').trim();
    tbl.classList.add('col-show-pp');
    document.querySelectorAll('.distrib-pills .pill:not(.gap-pill)').forEach(p=>p.classList.remove('on'));
    const ppBtn = document.querySelector('.distrib-pills .pill:not(.gap-pill)');
    if(ppBtn) ppBtn.classList.add('on');
    udCurrentContent = 'pp';
  }}
  // Filters
  document.getElementById('ud-f-region').style.display = dim==='dest' ? '' : 'none';
  document.getElementById('ud-search').style.display   = dim!=='reg'  ? '' : 'none';
  document.getElementById('ud-search').placeholder = dim==='corp' ? 'Buscar corporativo...' : 'Buscar destino...';
  document.getElementById('ud-search').value='';
  document.getElementById('ud-f-region').value='';
}}

function udFilter() {{
  const q   = document.getElementById('ud-search').value.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').trim();
  const reg = document.getElementById('ud-f-region').value;
  if (udDim === 'corp') {{
    const rows = document.querySelectorAll('.ud-corp-row[data-row-idx]');
    const vm   = document.getElementById('ud-corp-ver-mas');
    rows.forEach(r => {{
      const txt = r.cells[0]?.textContent.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'') || '';
      const idx = parseInt(r.dataset.rowIdx);
      r.style.display = !q ? (idx<10 ? '' : 'none') : (txt.includes(q) ? '' : 'none');
    }});
    if (vm) vm.style.display = q ? 'none' : '';
  }} else if (udDim === 'dest') {{
    const rows = document.querySelectorAll('.ud-dest-row[data-row-idx]');
    const vm   = document.getElementById('ud-dest-ver-mas');
    rows.forEach(r => {{
      const rReg = r.dataset.region || '';
      const txt  = r.cells[0]?.textContent.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'') || '';
      const idx  = parseInt(r.dataset.rowIdx);
      const matchR = !reg || rReg===reg;
      const matchQ = !q   || txt.includes(q);
      r.style.display = (matchR && matchQ) ? (!q && !reg ? (idx<10 ? '' : 'none') : '') : 'none';
    }});
    if (vm) vm.style.display = (q||reg) ? 'none' : '';
  }}
}}


function udToggleCorp(btn) {{
  const open = btn.dataset.open === '1';
  document.querySelectorAll('.ud-corp-row[data-row-idx]').forEach(r => {{
    const idx = parseInt(r.dataset.rowIdx);
    if (idx >= 10 && idx < 20) r.style.display = open ? 'none' : '';
  }});
  btn.dataset.open = open ? '0' : '1';
  btn.textContent  = open ? 'Ver 10 más' : 'Ver menos';
}}

function udToggleDest(btn) {{
  const open = btn.dataset.open === '1';
  document.querySelectorAll('.ud-dest-row[data-row-idx]').forEach(r => {{
    const idx = parseInt(r.dataset.rowIdx);
    if (idx >= 10 && idx < 20) r.style.display = open ? 'none' : '';
  }});
  btn.dataset.open = open ? '0' : '1';
  btn.textContent  = open ? 'Ver 10 más' : 'Ver menos';
}}

/* ── TABLE SORT ── */
let udSortDir = -1;
function udSortCol(colClass) {{
  const tbody = document.getElementById('ud-tbody');
  if (!tbody) return;
  const rows = Array.from(tbody.querySelectorAll('tr.ud-reg-row,tr.ud-corp-row[data-row-idx],tr.ud-dest-row[data-row-idx]'));
  const asc = tbody.dataset['sort_'+colClass] === 'asc';
  rows.sort((a,b) => {{
    const getVal = r => {{
      const td = r.querySelector('.'+colClass);
      return parseInt((td?.textContent||'0').replace(/[^0-9]/g,'')) || 0;
    }};
    return asc ? getVal(a)-getVal(b) : getVal(b)-getVal(a);
  }});
  rows.forEach(r => tbody.appendChild(r));
  tbody.dataset['sort_'+colClass] = asc ? 'desc' : 'asc';
}}

function udSortTotal() {{
  udSortDir *= -1;
  const tbody = document.getElementById('ud-tbody');
  // Only sort rows of the ACTIVE dimension
  const selector = udDim==='reg' ? '.ud-reg-row' : udDim==='corp' ? '.ud-corp-row[data-row-idx]' : '.ud-dest-row[data-row-idx]';
  const rows = Array.from(tbody.querySelectorAll(selector));
  rows.sort((a,b) => {{
    const av = parseInt((a.cells[1]?.textContent||'0').replace(/[\.\s]/g,'').replace(',','')) || 0;
    const bv = parseInt((b.cells[1]?.textContent||'0').replace(/[\.\s]/g,'').replace(',','')) || 0;
    return udSortDir * (bv - av);
  }});
  // Re-insert sorted rows, keeping visibility rules
  rows.forEach((r,i) => {{
    r.dataset.rowIdx = i;
    tbody.appendChild(r);
    if (udDim !== 'reg') r.style.display = i < 10 ? '' : 'none';
  }});
  // Move ver-mas after sorted rows
  const verMas = tbody.querySelector('#ud-corp-ver-mas, #ud-dest-ver-mas');
  if (verMas) tbody.appendChild(verMas);
  const th = document.getElementById('ud-sort-total');
  if (th) th.textContent = 'Total ' + (udSortDir === -1 ? '↓' : '↑');
}}

/* ── DEST TABLE SEARCH (unified table) ── */
let _destActiveRegion = '';
let _destActiveName   = '';
function destFilterRegion(val) {{
  _destActiveRegion = val.trim().toLowerCase();
  _destApplyFilter();
}}
function destFilterName(val) {{
  _destActiveName = val.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').trim();
  _destApplyFilter();
}}
function _destApplyFilter() {{
  // Works on BOTH the standalone dest-tbody and ud-dest-rows
  const rows = document.querySelectorAll('.ud-dest-row[data-row-idx], #dest-tbody tr[data-row-idx]');
  const vmUd   = document.getElementById('ud-dest-ver-mas');
  const vmDest = document.getElementById('dest-ver-mas-row');
  const reg = _destActiveRegion;
  const q   = _destActiveName;
  const hasFilter = !!(reg || q);
  rows.forEach(r => {{
    const rReg = (r.dataset.region||'').toLowerCase();
    const txt  = (r.cells[0]?.textContent||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'');
    const mR   = !reg || rReg.startsWith(reg);
    const mQ   = !q   || txt.includes(q);
    const idx  = parseInt(r.dataset.rowIdx??'999');
    r.style.display = (mR && mQ) ? (hasFilter ? '' : (idx<10 ? '' : 'none')) : 'none';
  }});
  if (vmUd)   vmUd.style.display   = hasFilter ? 'none' : '';
  if (vmDest) vmDest.style.display = hasFilter ? 'none' : '';
}}

/* ── APPLY ACTIVE FILTER TO CHANNEL VIEW ── */
function udApplyFilterToChannel() {{
  const filterBar = document.getElementById('ch-active-filter-bar');
  const f = udActiveFilter;
  if (!f.value) {{
    if (filterBar) filterBar.style.display = 'none';
    return;
  }}
  if (filterBar) {{
    filterBar.style.display = '';
    const label = filterBar.querySelector('#ch-filter-label');
    if (label) label.textContent = (f.type==='region' ? 'Región: ' : 'Corporativo: ') + f.value;
  }}
}}

/* ── CHANNEL DRILL ── */
let chCurrentData = [];
let chTipo = 'todos';

function chSetTipo(btn, tipo) {{
  document.querySelectorAll('#ch-tipo-pills .pill').forEach(p=>p.classList.remove('on'));
  btn.classList.add('on');
  chTipo = tipo;
  chRenderOverview();
}}

function chRenderOverview() {{
  const d = CH_DATA[chTipo];
  if (!d) return;
  const maxAvg = d.maxAvg || 1;

  // Render Producto Propio rows
  const pTbody = document.getElementById('ch-propio-tbody');
  if (pTbody) {{
    pTbody.innerHTML = d.propio.map(r => {{
      if (r.hoteles === 0) return '';
      const w = Math.min(100, r.avg_contratos / maxAvg * 100);
      return `<tr style="cursor:pointer" onclick="chDrill('${{r.channel}}',this)">
        <td><strong>${{r.channel}}</strong></td>
        <td>${{r.hoteles.toLocaleString('es-MX')}}</td>
        <td><div class="int-wrap"><div class="int-bar"><div class="int-fill" style="width:${{w.toFixed(0)}}%"></div></div>
          <span class="int-val">${{r.avg_contratos}} ctr/hotel</span></div></td>
        <td>${{r.destinos.toLocaleString('es-MX')}}</td></tr>`;
    }}).join('');
  }}

  // Render Third Party rows
  const tTbody = document.getElementById('ch-tercero-tbody');
  if (tTbody) {{
    tTbody.innerHTML = d.tercero.map(r => {{
      if (r.hoteles === 0) return '';
      const cls = r.residual ? ' class="residual"' : '';
      const badge = r.residual ? ' <span style="font-size:9px;color:var(--ink-muted)">Residual</span>' : '';
      return `<tr${{cls}}>
        <td><strong>${{r.channel}}</strong>${{badge}}</td>
        <td>${{r.hoteles.toLocaleString('es-MX')}}</td>
        <td>${{r.destinos.toLocaleString('es-MX')}}</td></tr>`;
    }}).join('');
  }}
}}

function chDrill(channel, row) {{
  // Update historical filter
  hFChannel = channel;
  hFilterCorpByChannel(channel);
  document.getElementById('hf-channel').value = channel;
  hRenderActivePills();
  hApplyFilter();
  // Highlight selected row
  document.querySelectorAll('#ch-propio-tbody tr, #ch-tercero-tbody tr').forEach(r=>r.classList.remove('sel'));
  row.classList.add('sel');
  // Add channel badge to ud-active-badge-container
  const container = document.getElementById('ud-active-badge-container');
  if (container) {{
    const existing = container.querySelector('[data-channel-pill]');
    if (existing) existing.remove();
    const pill = document.createElement('span');
    pill.dataset.channelPill = channel;
    pill.style.cssText = 'display:inline-flex;align-items:center;gap:4px;padding:3px 10px;'
      + 'border:1.5px solid #4FC3F4;border-radius:20px;font-size:9px;font-weight:700;'
      + 'letter-spacing:.06em;text-transform:uppercase;color:#4FC3F4;'
      + 'background:#E0F4FD;white-space:nowrap;';
    const x = document.createElement('span');
    x.textContent = '\u00d7';
    x.style.cssText = 'cursor:pointer;font-size:11px;line-height:1;margin-left:2px;';
    x.addEventListener('click', function() {{
      pill.remove();
      hFChannel = ''; hFilterCorpByChannel('');
      document.getElementById('hf-channel').value = '';
      document.querySelectorAll('#ch-propio-tbody tr, #ch-tercero-tbody tr').forEach(r=>r.classList.remove('sel'));
      hRenderActivePills(); hApplyFilter();
    }});
    pill.textContent = channel;
    pill.appendChild(x);
    container.appendChild(pill);
  }}
  // Scroll to chart
  const hist = document.querySelector('.sec-head');
  if (hist) hist.scrollIntoView({{behavior:'smooth', block:'start'}});
}}
function chBack() {{
  document.getElementById('ch-overview').style.display = '';
  document.getElementById('ch-drill').style.display = 'none';
}}
function chFilter() {{
  const reg  = document.getElementById('ch-f-region').value.trim().toLowerCase();
  const corp = document.getElementById('ch-f-corp').value;
  const dest = document.getElementById('ch-f-dest').value.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'');
  const filtered = chCurrentData.filter(r => {{
    const rDest = (r.Destino||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'');
    return (!reg  || (r.region||'').toLowerCase().startsWith(reg)) &&
           (!corp || r.Corporativo===corp) &&
           (!dest || rDest.includes(dest));
  }});
  document.getElementById('ch-drill-tbody').innerHTML = filtered.slice(0,200).map(r=>
    `<tr><td><strong>${{r.Hotel}}</strong></td><td class="dim">${{r.region}}</td><td class="dim">${{r.Corporativo}}</td><td class="dim">${{r.Destino}}</td></tr>`
  ).join('');
  document.getElementById('ch-drill-count').textContent =
    `Mostrando ${{Math.min(filtered.length,200)}} de ${{filtered.length}} hoteles`;
}}

function toggleCorpRows(btn) {{
  const open = btn.dataset.open === '1';
  const rows = document.querySelectorAll('#corp-tbody tr[data-row-idx]');
  let shown = 0;
  rows.forEach(r => {{
    const idx = parseInt(r.dataset.rowIdx);
    if (open) {{
      if (idx >= 10 && idx < 20) {{ r.style.display='none'; r.classList.add('rows-more'); }}
    }} else {{
      if (idx >= 10 && idx < 20) {{ r.style.display=''; r.classList.remove('rows-more'); shown++; }}
    }}
  }});
  btn.dataset.open = open ? '0' : '1';
  btn.textContent  = open ? 'Ver 10 más' : 'Ver menos';
}}

function toggleDestRows(btn) {{
  const open = btn.dataset.open === '1';
  const rows = document.querySelectorAll('#dest-tbody tr[data-row-idx]');
  rows.forEach(r => {{
    const idx = parseInt(r.dataset.rowIdx);
    if (open) {{
      if (idx >= 10 && idx < 20) {{ r.style.display='none'; r.classList.add('rows-more'); }}
    }} else {{
      if (idx >= 10 && idx < 20) {{ r.style.display=''; r.classList.remove('rows-more'); shown++; }}
    }}
  }});
  btn.dataset.open = open ? '0' : '1';
  btn.textContent  = open ? 'Ver 10 más' : 'Ver menos';
}}
function filterTable(tbodyId, q, colIdx) {{
  const rows = document.getElementById(tbodyId).querySelectorAll('tr[data-row-idx]');
  const ql = q.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').trim();
  const verMasRow = document.getElementById(tbodyId.replace('-tbody','-ver-mas-row'));
  rows.forEach(r => {{
    const cell = r.cells[colIdx];
    if (!cell) return;
    const t = cell.textContent.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'');
    if (!ql) {{
      const idx = parseInt(r.dataset.rowIdx||'0');
      r.style.display = idx < 10 ? '' : 'none';
      if (idx>=10) r.classList.add('rows-more'); else r.classList.remove('rows-more');
    }} else {{
      r.style.display = t.includes(ql) ? '' : 'none';
      r.classList.remove('rows-more');
    }}
  }});
  if (verMasRow) verMasRow.style.display = ql ? 'none' : '';
}}

function filterByRegion(val, tbodyId, colIdx) {{
  const rows = document.getElementById(tbodyId).querySelectorAll('tr:not(.more-hint-row)');
  rows.forEach(r => {{
    const cell = r.cells[colIdx];
    if (!cell) return;
    r.style.display = (!val || cell.textContent.trim()===val) ? '' : 'none';
  }});
  if (val) updateCards({{type:'region', name:val}});
  else updateCards({{type:'all'}});
}}

/* ── HISTORICAL ── */
let hLevel = 'sem';
let hYear  = 2026;
let hMonth = null;
let hChart = null;
const MN_HIST = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'];

// Filter state
let hFRegion  = '';
let hFCorp    = '';
let hFChannel = '';
let hFTipo    = '';

function hDestroyChart() {{ if(hChart){{ hChart.destroy(); hChart=null; }} }}

// Returns filtered + aggregated data array — uses hotel-level index to avoid
// double-counting hotels with multiple channels (unless channel filter is active)
function hGetDim() {{
  const src = hFChannel ? HIST.dim : HIST.dim_hotel;
  return src.filter(r =>
    (!hFRegion  || r.region  === hFRegion)  &&
    (!hFCorp    || r.corp    === hFCorp)    &&
    (!hFChannel || r.channel === hFChannel) &&
    (!hFTipo    || r.ch_tipo === hFTipo)
  );
}}

// Aggregate dim rows to weekly netnew, then compute running acum within window
function hAggrByYw(rows) {{
  const map = {{}};
  rows.forEach(r => {{
    if (!map[r.yw]) map[r.yw] = {{yw:r.yw, ym:r.ym, netnew:0}};
    map[r.yw].netnew += r.n;
  }});
  const arr = Object.values(map).sort((a,b)=>a.yw.localeCompare(b.yw));
  let c=0; arr.forEach(r=>{{ c+=r.netnew; r.acum=c; }});
  return arr;
}}

function hAggrByYm(rows) {{
  const map = {{}};
  rows.forEach(r => {{
    if (!map[r.ym]) map[r.ym] = {{ym:r.ym, netnew:0}};
    map[r.ym].netnew += r.n;
  }});
  const arr = Object.values(map).sort((a,b)=>a.ym.localeCompare(b.ym));
  let c=0; arr.forEach(r=>{{ c+=r.netnew; r.acum=c; }});
  return arr;
}}

function hIsFiltered() {{
  return !!(hFRegion || hFCorp || hFChannel || hFTipo);
}}

// Called when any filter changes
function hUpdateComboStyle(selId) {{
  const el = document.getElementById(selId);
  if (!el) return;
  if (el.value) el.classList.add('has-value');
  else el.classList.remove('has-value');
}}

function hFilterCorpByChannel(ch) {{
  const sel = document.getElementById('hf-corp');
  const current = sel.value;
  sel.innerHTML = '<option value="">Todos</option>';
  const corps = ch ? (CH_CORP_MAP[ch] || []) : {json.dumps(hist_corps, cls=NpEncoder)};
  corps.forEach(c => {{
    const opt = document.createElement('option');
    opt.value = c; opt.textContent = c;
    sel.appendChild(opt);
  }});
  // Keep current selection if still valid
  sel.value = corps.includes(current) ? current : '';
  hFCorp = sel.value;
}}

function hApplyFilter() {{
  hBreadcrumb();
  hRenderActivePills();
  hRender();
  const _gd=document.getElementById('ud-gap-content'); if(_gd&&_gd.style.display!=='none') gapSyncDim();
}}

function hPillRegion(btn) {{
  const val = btn.dataset.region;
  document.querySelectorAll('#hf-region-pills .pill').forEach(p=>p.classList.remove('on'));
  btn.classList.add('on');
  hFRegion = val;
  // Highlight matching row in distribution table
  document.querySelectorAll('#ud-tbody tr').forEach(r => r.classList.remove('ud-filter-active'));
  if (val) {{
    document.querySelectorAll('.ud-reg-row').forEach(r => {{
      const name = r.querySelector('td strong')?.textContent || '';
      if (name === val) r.classList.add('ud-filter-active');
    }});
  }}
  hApplyFilter();
}}

function hTipo(btn, val) {{
  document.querySelectorAll('#htipo-pills .pill').forEach(b=>b.classList.remove('on'));
  btn.classList.add('on');
  hFTipo = val;
  hApplyFilter();
}}

function hClearFilters() {{
  hFRegion=''; hFCorp=''; hFChannel=''; hFTipo='';
  document.querySelectorAll('#hf-region-pills .pill').forEach(p=>p.classList.remove('on'));
  const todasBtn = document.querySelector('#hf-region-pills .pill[data-region=""]');
  if (todasBtn) todasBtn.classList.add('on');
  document.querySelectorAll('#htipo-pills .pill').forEach(b=>b.classList.remove('on'));
  document.getElementById('htipo-all').classList.add('on');
  document.getElementById('hf-channel').value=''; hUpdateComboStyle('hf-channel');
  hFilterCorpByChannel('');
  document.getElementById('hf-corp').value=''; hUpdateComboStyle('hf-corp');
  hRenderActivePills();
  hApplyFilter();
}}

function hDestroyChart() {{ if(hChart){{ hChart.destroy(); hChart=null; }} }}

function hRenderActivePills() {{
  // Show/hide Limpiar button
  const limpiar = document.getElementById('btn-limpiar');
  if (limpiar) limpiar.style.display = hIsFiltered() ? '' : 'none';
  // Update hf-active-pills (historical section above chart only)
  const container = document.getElementById('hf-active-pills');
  if (container) {{
    container.innerHTML = '';
    const defs = [
      {{ key:'region', val:hFRegion }},
      {{ key:'channel', val:hFChannel }},
      {{ key:'corp', val:hFCorp }},
    ];
    defs.forEach(d => {{
      if (!d.val) return;
      const pill = document.createElement('span');
      pill.style.cssText = 'display:inline-flex;align-items:center;gap:4px;padding:3px 10px;'
        + 'border:1.5px solid var(--accent);border-radius:20px;font-size:9px;font-weight:700;'
        + 'letter-spacing:.06em;text-transform:uppercase;color:var(--accent);'
        + 'background:var(--accent-soft,#E0F4FD);white-space:nowrap;';
      const x = document.createElement('span');
      x.textContent = '\u00d7';
      x.dataset.filterKey = d.key;
      x.style.cssText = 'cursor:pointer;font-size:11px;line-height:1;margin-left:2px;';
      x.addEventListener('click', function() {{ hClearFilter(this.dataset.filterKey); }});
      pill.textContent = d.val;
      pill.appendChild(x);
      container.appendChild(pill);
    }});
  }}
  // udSyncBadges handles ud-active-badge-container independently
  if (typeof udSyncBadges === 'function') udSyncBadges();
}}

function hClearFilter(key) {{
  if (key==='region') {{
    hFRegion='';
    document.querySelectorAll('#hf-region-pills .pill').forEach(p=>p.classList.remove('on'));
    const todas = document.querySelector('#hf-region-pills .pill[data-region=""]');
    if (todas) todas.classList.add('on');
  }}
  else if (key==='channel') {{ hFChannel=''; document.getElementById('hf-channel').value=''; hUpdateComboStyle('hf-channel'); hFilterCorpByChannel(''); }}
  else if (key==='corp') {{ hFCorp=''; document.getElementById('hf-corp').value=''; hUpdateComboStyle('hf-corp'); }}
  else if (key==='tipo') {{
    hFTipo='';
    document.querySelectorAll('#htipo-pills .pill').forEach(b=>b.classList.remove('on'));
    document.getElementById('htipo-all').classList.add('on');
  }}
  hRenderActivePills();
  hApplyFilter();
}}

function hBreadcrumb() {{
  const bc = document.getElementById('drill-bc');
  bc.innerHTML = '';
  function mk(txt, cls, fn) {{
    const s = document.createElement('span');
    s.className = cls; s.textContent = txt;
    if (fn) {{ s.style.cursor='pointer'; s.addEventListener('click', fn); }}
    bc.appendChild(s);
  }}
  function sep() {{ mk(' \u203a ', 'bc-sep'); }}
  if (hLevel === 'anio') {{
    mk('Todos los a\u00f1os', 'bc-cur');
  }} else if (hLevel === 'mes') {{
    mk('Todos los a\u00f1os', 'bc-link', function(){{ hGoLevel('anio'); }});
    sep(); mk(String(hYear), 'bc-cur');
  }} else {{
    mk('Todos los a\u00f1os', 'bc-link', function(){{ hGoLevel('anio'); }});
    sep(); mk(String(hYear), 'bc-link', function(){{ hGoLevel('mes'); }});
    sep(); mk(MN_HIST[hMonth] || '', 'bc-cur');
  }}
}}

function hPopulateWeeks(yr, mo) {{
  const prefix = yr ? String(yr)+'-W' : '';
  let weeks = HIST.by_week.filter(r => r.yw.startsWith(prefix));
  if (mo) {{
    const ymKey = String(yr)+'-'+(mo<10?'0':'')+mo;
    weeks = weeks.filter(r => r.ym === ymKey);
  }}
  const sel = document.getElementById('sel-week');
  if (!sel) return;
  sel.innerHTML = '<option value="">Todas</option>'
    + weeks.map(r => {{
        const wn = r.yw.split('-W')[1];
        return '<option value="'+r.yw+'">W'+wn+'</option>';
      }}).join('');
  sel.value = '';
  sel.disabled = false;
}}

function hGoLevel(level) {{
  if (level === 'anio') {{ hMonth = null; }}
  hLevel = level;
  if (level === 'sem' && hYear) hPopulateWeeks(hYear, hMonth);
  hSyncUI();
  hBreadcrumb();
  hRender();
}}

function hSetLevel(level) {{
  hGoLevel(level);
}}

function hSyncUI() {{
  ['anio','mes','sem'].forEach(l => {{
    document.getElementById('dt-'+l).classList.toggle('on', l===hLevel);
  }});
  document.getElementById('sel-year').disabled  = false;
  document.getElementById('sel-month').disabled = hLevel==='anio';
}}

function hSelYear(y) {{
  if (!y) return;
  hYear  = parseInt(y);
  hMonth = null;
  hUpdateComboStyle('sel-year');
  const avail = HIST.months_by_year[hYear] || [];
  const msel = document.getElementById('sel-month');
  msel.innerHTML = '<option value="">Todos</option>'
    + avail.map(m=>'<option value="'+m+'">'+MN_HIST[m]+'</option>').join('');
  msel.value = '';
  if (hLevel === 'sem') {{
    hPopulateWeeks(hYear, null);
    hBreadcrumb(); hRender();
  }} else {{
    hGoLevel('mes');
  }}
}}

function hSelMonth(m) {{
  if (!m) return;
  hMonth = parseInt(m);
  hUpdateComboStyle('sel-month');
  if (hLevel === 'sem') hPopulateWeeks(hYear, hMonth);
  hGoLevel('sem');
}}

function hRender() {{
  hDestroyChart();
  const ctx = document.getElementById('canvas-hist').getContext('2d');

  if (hLevel === 'sem' && !hYear) {{
    ctx.clearRect(0,0,ctx.canvas.width,ctx.canvas.height);
    ctx.font='13px system-ui'; ctx.fillStyle='#8A8377';
    ctx.fillText('Seleccion\u00e1 un a\u00f1o primero', 20, 90);
    return;
  }}

  const filtered = hIsFiltered();
  let labels, acum, netnew, title, onClickFn = null;

  // Helper: compute true global cumulative acum for a series of rows
  // For filtered data, we must compute a running sum starting from 0
  // (filtered acum can't be pre-computed globally since it's a subset)
  // For unfiltered data, use the pre-computed global acum from HIST

  if (hLevel === 'anio') {{
    // ── Por Año: un punto por año ──
    let d;
    if (filtered) {{
      const byYm = hAggrByYm(hGetDim());
      const byYr = {{}};
      byYm.forEach(r => {{ const yr=r.ym.split('-')[0]; byYr[yr]=(byYr[yr]||0)+r.netnew; }});
      let c=0; d=Object.keys(byYr).sort().map(yr=>{{c+=byYr[yr];return{{yr,netnew:byYr[yr],acum:c}};}});
    }} else {{
      d = HIST.by_year;
    }}
    labels = d.map(r=>String(r.year||r.yr));
    netnew = d.map(r=>r.netnew);
    acum   = d.map(r=>r.acum);
    title  = 'Acumulado — Por A\u00f1o';
    onClickFn = (_,els)=>{{
      if (!els.length) return;
      const yr = parseInt(labels[els[0].index]);
      if (!isNaN(yr)) {{ hYear=yr; document.getElementById('sel-year').value=yr; hSelYear(yr); }}
    }};

  }} else if (hLevel === 'mes') {{
    // ── Por Mes: meses del año seleccionado, con fill de meses vacíos ──
    const prefix = String(hYear)+'-';
    // Referencia: todos los meses de este año en HIST.by_month (tiene fill hasta snapshot)
    const refMonths = HIST.by_month.filter(r=>r.ym.startsWith(prefix));
    let d;
    if (filtered) {{
      const all = hAggrByYm(hGetDim());
      const mMap = {{}};
      all.forEach(r=>{{ mMap[r.ym]=r.netnew; }});
      const before = all.filter(r=>r.ym < prefix);
      let c = before.length ? before[before.length-1].acum : 0;
      d = refMonths.map(r=>{{ const n=mMap[r.ym]||0; c+=n; return {{ym:r.ym,netnew:n,acum:c}}; }});
    }} else {{
      d = refMonths;
    }}
    if (!d.length) {{
      ctx.clearRect(0,0,ctx.canvas.width,ctx.canvas.height);
      ctx.font='13px system-ui'; ctx.fillStyle='#8A8377';
      ctx.fillText('Sin datos para '+hYear, 20, 90); return;
    }}
    labels = d.map(r=>MN_HIST[parseInt(r.ym.split('-')[1])]);
    netnew = d.map(r=>r.netnew);
    acum   = d.map(r=>r.acum);
    title  = 'Acumulado — '+hYear;
    onClickFn = (_,els)=>{{
      if (!els.length) return;
      hMonth = parseInt(d[els[0].index].ym.split('-')[1]);
      document.getElementById('sel-month').value = hMonth;
      hGoLevel('sem');
    }};

  }} else {{
    // ── Por Semana: TODAS las semanas del período incluyendo vacías ──
    // Referencia: HIST.by_week tiene fill completo hasta W21 para todas las semanas
    let refWeeks;
    if (hMonth) {{
      const ymKey = String(hYear)+'-'+String(hMonth).padStart(2,'0');
      refWeeks = HIST.by_week.filter(r=>r.ym===ymKey);
    }} else if (hYear) {{
      refWeeks = HIST.by_week.filter(r=>r.yw.startsWith(String(hYear)+'-W'));
    }} else {{
      refWeeks = HIST.by_week;
    }}

    let d;
    if (filtered) {{
      // Mapa sparse de netnew filtrado (solo hoteles con FechaCreación)
      const dimRows = hGetDim();
      const sparseMap = {{}};
      dimRows.forEach(r=>{{ sparseMap[r.yw]=(sparseMap[r.yw]||0)+r.n; }});
      // Acum base = hoteles del subconjunto antes del rango de semanas visible
      const firstYw = refWeeks.length ? refWeeks[0].yw : '';
      const allDim = hAggrByYw(dimRows);
      const before = allDim.filter(r=>r.yw < firstYw);
      // Si no hay historia antes del rango, el acum arranca en 0 para netnew relativo
      // El acumulado absoluto lo obtenemos sumando todo el dim_hotel del subconjunto
      const totalInSubset = allDim.length ? allDim[allDim.length-1].acum : 0;
      const inRange = allDim.filter(r=>r.yw >= firstYw).reduce((s,r)=>s+r.netnew,0);
      // Base = total acumulado - lo que cae dentro del rango visible
      let c = before.length ? before[before.length-1].acum : (totalInSubset - inRange);
      // Fill todas las semanas de referencia con netnew=0 donde no hay datos
      d = refWeeks.map(w=>{{ const n=sparseMap[w.yw]||0; c+=n; return {{yw:w.yw,ym:w.ym,netnew:n,acum:c}}; }});
    }} else {{
      d = refWeeks; // ya tiene fill completo y acum global
    }}

    if (!d.length) {{
      ctx.clearRect(0,0,ctx.canvas.width,ctx.canvas.height);
      ctx.font='13px system-ui'; ctx.fillStyle='#8A8377';
      ctx.fillText('Sin datos de contratos para este per\u00edodo', 20, 90); return;
    }}
    labels = d.map(r=>r.yw ? 'W'+r.yw.split('-W')[1] : 'W'+(r.week||'?'));
    netnew = d.map(r=>r.netnew);
    acum   = d.map(r=>r.acum);
    title  = hMonth
      ? 'Acumulado — '+MN_HIST[hMonth]+' '+hYear
      : hYear ? 'Acumulado — Semanas '+hYear
      : 'Acumulado — Todas las semanas';
  }}

  const filterTag = filtered
    ? ' · ' + [hFRegion,hFChannel,hFTipo,hFCorp].filter(Boolean).join(' / ')
    : '';

  const yMin = Math.floor(Math.min(...acum) * 0.995 / 10) * 10;
  const yMax = Math.ceil(Math.max(...acum) * 1.005 / 10) * 10;

  hChart = new Chart(ctx, {{
    type:'bar',
    data:{{
      labels,
      datasets:[
        {{label:'Acumulado',data:acum,
          backgroundColor:'rgba(79,195,244,0.25)',borderColor:'#4FC3F4',borderWidth:1,
          yAxisID:'y',order:2}},
        {{label:'Net new',data:netnew,type:'line',
          borderColor:'#1A6B4A',borderWidth:2,
          backgroundColor:'rgba(26,107,74,0.08)',fill:false,
          pointRadius:3,pointBackgroundColor:'#1A6B4A',tension:0.3,
          yAxisID:'y2',order:1}}
      ]
    }},
    options:{{
      responsive:true,maintainAspectRatio:false,
      onClick: onClickFn || undefined,
      plugins:{{
        title:{{display:false}},
        legend:{{
          display:true,
          position:'bottom',
          labels:{{font:{{size:10,family:'Geist,system-ui,sans-serif'}},
            boxWidth:10,padding:16,color:'#5F5E5A'}}
        }},
        tooltip:{{mode:'index',intersect:false,padding:8,
          backgroundColor:'rgba(26,25,23,0.85)',titleColor:'#fff',bodyColor:'#E2DDD6',
          titleFont:{{size:11,weight:'700'}},bodyFont:{{size:10}},
          callbacks:{{label:c=>c.dataset.label+': '+c.parsed.y.toLocaleString('es-MX')}}}}
      }},
      scales:{{
        x:{{
          grid:{{color:'rgba(226,221,214,0.4)'}},
          ticks:{{font:{{size:10,family:'Geist,system-ui,sans-serif'}},color:'#8A8377',maxRotation:0}}
        }},
        y:{{
          position:'left',
          grid:{{color:'rgba(226,221,214,0.4)'}},
          min:yMin, max:yMax,
          ticks:{{font:{{size:10,family:'Geist,system-ui,sans-serif'}},color:'#4FC3F4',
            callback:v=>v>=1000?(v/1000).toFixed(1)+'K':v}},
          title:{{display:true,text:'Acumulado',font:{{size:9}},color:'#4FC3F4'}}
        }},
        y2:{{
          position:'right',
          grid:{{display:false}},
          beginAtZero:true,
          ticks:{{font:{{size:10,family:'Geist,system-ui,sans-serif'}},color:'#1A6B4A',
            callback:v=>v>=1000?(v/1000).toFixed(1)+'K':v}},
          title:{{display:true,text:'Net new',font:{{size:9}},color:'#1A6B4A'}}
        }}
      }}
    }}
  }});
}}

/* ── INIT — se ejecuta después de Chart.js (ver script al final del body) ── */
function hInit() {{
  // Check if 2026 has weekly data — if not, fallback to Por Año
  const weeks2026 = HIST.by_week.filter(r=>r.yw.startsWith('2026-W'));
  if (weeks2026.length > 0) {{
    document.getElementById('sel-year').value = '2026';
    const avail = HIST.months_by_year[2026] || [];
    const sel = document.getElementById('sel-month');
    sel.innerHTML = '<option value="">Todos</option>'
      + avail.map(m=>'<option value="'+m+'">'+MN_HIST[m]+'</option>').join('');
    sel.disabled = false;
    hLevel = 'sem'; hYear = 2026; hMonth = null;
    hPopulateWeeks(2026, null);
  }} else {{
    // No weekly data for 2026 — show Por Año instead
    hLevel = 'anio'; hYear = null; hMonth = null;
  }}
  hSyncUI();
  hBreadcrumb();
  hRender();
}}

"""

# ── HTML HELPERS ──────────────────────────────
def pct_bar_html(pct, color='var(--accent)', max_pct=100):
    w = min(100, pct/max_pct*100) if max_pct else 0
    return (f'<div class="pct-wrap">'
            f'<div class="pct-bar"><div class="pct-fill" style="width:{w:.1f}%;background:{color}"></div></div>'
            f'<span class="pct-val" style="color:{color}">{pct:.1f}%</span></div>')

def vs_bar_html(vs):
    color = 'var(--green)' if vs >= 0 else '#C0392B'
    sign  = '+' if vs >= 0 else ''
    w = min(100, abs(vs) / 30 * 100)
    return (f'<div class="pct-wrap">'
            f'<div class="pct-bar"><div class="pct-fill" style="width:{w:.1f}%;background:{color}"></div></div>'
            f'<span class="pct-val" style="color:{color}">{sign}{vs:.1f}pp</span></div>')

# ── ZONA 3: REGIÓN ──
def build_unified_distrib():
    region_opts_ud = ''.join(f'<option value="{r}">{r}</option>' for r in REGION_ORDER)

    reg_rows = ''
    for r in reg_stats:
        reg=r['region']; tot=r['total']; sp=r['solo_propio']; hy=r['hybrid']; pp_=r['prod_propio']; tp=r['solo_tercero']
        pct_pp = pp_/tot*100 if tot else 0; vs = pct_pp - pp/N*100
        reg_rows += (
            f'<tr class="ud-reg-row" style="cursor:pointer"'
            f' onclick="udRowClick(\'region\',\'{reg}\',this);updateCards({{type:\'region\',name:\'{reg}\'}})">'
            f'<td><strong>{reg}</strong></td>'
            f'<td class="td-tot">{fmt_n(tot)}</td>'
            f'<td class="td-pp">{fmt_n(pp_)}</td>'
            f'<td class="td-sp" style="color:var(--green);font-size:11px;">{fmt_n(sp)}</td>'
            f'<td class="td-hy" style="color:var(--violet,#5C469C);font-size:11px;">{fmt_n(hy)}</td>'
            f'<td class="td-tp">{fmt_n(tp)}</td>'
            f'<td>{pct_bar_html(pct_pp,"#4FC3F4")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )

    corp_rows = ''
    for i,(_, r) in enumerate(corp_grp.head(200).iterrows()):
        cls = 'rows-more' if i>=10 else ''; sty = 'display:none'
        corp=str(r['Corporativo']).replace("'",""); tot=r['total']
        sp=int(r['solo_propio']); hy=int(r['hybrid']); pp_=sp+hy; tp=int(r['solo_tercero'])
        pct_pp = pp_/tot*100 if tot else 0; vs = pct_pp - pp/N*100
        corp_rows += (
            f'<tr class="ud-corp-row {cls}" style="display:none;cursor:pointer" data-row-idx="{i}"'
            f' onclick="udRowClick(\'corp\',\'{corp}\',this);updateCards({{type:\'corp\',name:\'{corp}\'}})">'
            f'<td><strong>{r["Corporativo"]}</strong></td>'
            f'<td class="td-tot">{fmt_n(tot)}</td>'
            f'<td class="td-pp">{fmt_n(pp_)}</td>'
            f'<td class="td-sp" style="color:var(--green);font-size:11px;">{fmt_n(sp)}</td>'
            f'<td class="td-hy" style="color:var(--violet,#5C469C);font-size:11px;">{fmt_n(hy)}</td>'
            f'<td class="td-tp">{fmt_n(tp)}</td>'
            f'<td>{pct_bar_html(pct_pp,"#4FC3F4")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )
    ver_mas_corp = ('<tr class="ud-corp-row" style="display:none" id="ud-corp-ver-mas"><td colspan="8" style="text-align:center;padding:10px;">'
                    '<button onclick="udToggleCorp(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
                    'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
                    'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
                    'Ver 10 más</button></td></tr>')

    dest_rows = ''
    for i,(_, r) in enumerate(dest_grp.head(500).iterrows()):
        cls = 'rows-more-dest' if i>=10 else ''; sty = 'display:none'
        dest_name = str(r['Destino']).replace("'", '')
        pp_=r['prod_propio']; tp=r['solo_tercero']; tot=r['total']
        pct_pp = pp_/tot*100 if tot else 0; vs = pct_pp - pp/N*100
        dest_rows += (
            f'<tr class="ud-dest-row {cls}" style="display:none;cursor:pointer" onclick="udRowClick(&quot;dest&quot;,&quot;{dest_name}&quot;,this)" data-row-idx="{i}" data-region="{r['Region_display']}">'
            f'<td><strong>{r["Destino"]}</strong>'
            f'<div style="font-size:10px;color:var(--ink-muted);line-height:1.3;margin-top:1px;">{r["Region_display"]}</div></td>'
            f'<td>{fmt_n(int(tot))}</td>'
            f'<td class="td-pp">{fmt_n(int(pp_))}</td>'
            f'<td class="td-sp" style="color:var(--green);font-size:11px;opacity:.4;">—</td>'
            f'<td class="td-hy" style="color:var(--violet,#5C469C);font-size:11px;opacity:.4;">—</td>'
            f'<td class="td-tp">{fmt_n(int(tp))}</td>'
            f'<td>{pct_bar_html(pct_pp,"#4FC3F4")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )
    ver_mas_dest = ('<tr class="ud-dest-row" style="display:none" id="ud-dest-ver-mas"><td colspan="8" style="text-align:center;padding:10px;">'
                    '<button onclick="udToggleDest(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
                    'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
                    'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
                    'Ver 10 más</button></td></tr>')

    return f"""
    <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;">
    </div>
    <table class="data-table">
      <thead><tr>
        <th id="ud-dim-th" style="text-align:left;width:220px;">Región</th>
        <th id="ud-sort-total" onclick="udSortTotal()" style="cursor:pointer;user-select:none;" title="Ordenar por total">Total ↓</th>
        <th class="th-pp" onclick="udSortCol('td-pp')" style="cursor:pointer;user-select:none;" title="Ordenar">P. Propio ↕</th>
        <th class="th-sp" onclick="udSortCol('td-sp')" style="font-size:8px;cursor:pointer;user-select:none;" title="Ordenar">Solo P. ↕</th>
        <th class="th-hy" onclick="udSortCol('td-hy')" style="font-size:8px;cursor:pointer;user-select:none;" title="Ordenar">Hybrid ↕</th>
        <th class="th-tp">Third P.</th>
        <th class="th-pct-label" style="min-width:120px;">% Propio</th>
        <th>vs Global</th>
      </tr></thead>
      <tbody id="ud-tbody">
        <tr class="global-row">
          <td>GLOBAL</td>
          <td class="td-pp">{fmt_n(pp)}</td>
          <td class="td-sp" style="opacity:.55;">{fmt_n(solo_propio)}</td>
          <td class="td-hy" style="opacity:.55;">{fmt_n(hybrid)}</td>
          <td class="td-tp">{fmt_n(solo_terc)}</td>
          <td>{fmt_n(N)}</td>
          <td>{pct_bar_html(pp/N*100,"#4FC3F4")}</td>
          <td>—</td>
        </tr>
        {reg_rows}
        {corp_rows}{ver_mas_corp}
        {dest_rows}{ver_mas_dest}
      </tbody>
    </table>"""

def build_region_tabs():
    t = build_unified_distrib()
    return {"pp": t, "sp": t, "hy": t, "tp": t}


# ── ZONA 4: CORPORATIVO ──
def build_corp_tab():
    rows = ''
    for i, (_, r) in enumerate(corp_grp.head(200).iterrows()):
        cls = 'rows-more' if i >= 10 else ''
        style = 'display:none' if i >= 10 else ''
        onclick = f"onclick=\"updateCards({{type:'corp',name:'{str(r['Corporativo']).replace(chr(39), '')}'}})\"style=\"cursor:pointer\""
        rows += (f'<tr class="{cls}" style="{style}" data-row-idx="{i}" {onclick}>'
                 f'<td><strong>{r["Corporativo"]}</strong></td>'
                 f'<td>{fmt_n(r["total"])}</td>'
                 f'<td class="td-sp">{fmt_n(r["solo_propio"])}</td>'
                 f'<td class="td-hy">{fmt_n(r["hybrid"])}</td>'
                 f'<td class="td-tp">{fmt_n(r["solo_tercero"])}</td>'
                 f'<td>{pct_bar_html(r["pct_propio"],"#4FC3F4")}</td></tr>')
    ver_mas = (f'<tr id="corp-ver-mas-row"><td colspan="6" style="text-align:center;padding:10px;">'
               f'<button onclick="toggleCorpRows(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
               f'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
               f'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
               f'Ver 10 más</button></td></tr>')
    return f'''
      <div class="sb-wrap" style="position:relative;">
        <input class="sb-input" id="corp-search" placeholder="Buscar entre {fmt_n(len(corp_grp))} corporativos…"
          data-tbody-id="corp-tbody" type="text" autocomplete="off">
        <div class="autocomplete-list" id="corp-ac-list" style="display:none;"></div>
        <span class="sb-count" id="corp-tbody-count">Mostrando 10 · {fmt_n(len(corp_grp))} total</span>
      </div>
      <table class="data-table">
        <thead><tr>
          <th style="text-align:left;width:180px;">Corporativo</th>
          <th>Total</th>
          <th class="th-sp" style="font-size:8px;">Solo Propio</th>
          <th class="th-hy" style="font-size:8px;">Hybrid</th>
          <th class="th-tp" style="font-size:8px;">Third Party</th>
          <th style="min-width:120px;">% Propio</th>
        </tr></thead>
        <tbody id="corp-tbody">{rows}{ver_mas}</tbody>
      </table>'''

# ── ZONA 4: DESTINO ──
def build_dest_tab():
    rows = ''
    for i, (_, r) in enumerate(dest_grp.head(500).iterrows()):
        cls = 'rows-more' if i >= 10 else ''
        style = 'display:none' if i >= 10 else ''
        rows += (f'<tr class="{cls}" style="{style}" data-row-idx="{i}" data-region="{r["Region_display"]}">'
                 f'<td><strong>{r["Destino"]}</strong>'
                 f'<div style="font-size:10px;color:var(--ink-muted);font-weight:400;line-height:1.3;margin-top:1px;">{r["Region_display"]}</div></td>'
                 f'<td>{fmt_n(r["total"])}</td>'
                 f'<td class="td-pp">{fmt_n(r["prod_propio"])}</td>'
                 f'<td class="td-tp">{fmt_n(r["solo_tercero"])}</td>'
                 f'<td>{pct_bar_html(r["pct_propio"],"#4FC3F4")}</td></tr>')
    ver_mas = (f'<tr id="dest-ver-mas-row"><td colspan="5" style="text-align:center;padding:10px;">'
               f'<button onclick="toggleDestRows(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
               f'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
               f'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
               f'Ver 10 más</button></td></tr>')
    return f'''
      <div class="sb-wrap" style="position:relative;display:flex;gap:8px;align-items:center;">
        <input class="sb-input" id="dest-region-search" placeholder="Región…"
          type="text" autocomplete="off" oninput="destFilterRegion(this.value)"
          style="max-width:130px;" list="dest-region-list">
        <datalist id="dest-region-list">{''.join(f"<option value='{r}'>" for r in REGION_ORDER)}</datalist>
        <input class="sb-input" id="dest-search" placeholder="Buscar entre {fmt_n(len(dest_grp))} destinos…"
          data-tbody-id="dest-tbody" type="text" autocomplete="off"
          oninput="destFilterName(this.value)">
        <span class="sb-count">Mostrando 10 · {fmt_n(len(dest_grp))} total</span>
      </div>
      <table class="data-table" style="margin-top:8px;">
        <thead><tr>
          <th style="text-align:left;">Destino</th>
          <th>Total</th>
          <th class="th-pp" style="font-size:8px;">P. Propio</th>
          <th class="th-tp" style="font-size:8px;">Third Party</th>
          <th style="min-width:110px;">% Propio</th>
        </tr></thead>
        <tbody id="dest-tbody">{rows}{ver_mas}</tbody>
      </table>'''

# ── ZONA 4: CHANNEL ──
def build_channel_tab():
    region_opts = ''.join(f'<option value="{r}">{r}</option>' for r in REGION_ORDER)
    corp_opts_ch = ''.join(f'<option value="{r["Corporativo"]}">{r["Corporativo"]}</option>'
                           for _, r in corp_grp.head(50).iterrows())

    max_avg_dest = max((r.get('avg_destinos',0) for r in ch_propio), default=1) or 1
    p_rows = ''
    for r in ch_propio:
        pct_bar = min(100, r['avg_contratos']/max_avg_ctr*100) if max_avg_ctr else 0
        p_rows += (f'<tr style="cursor:pointer" data-channel="{r["channel"]}" onclick="chDrill(\'{r["channel"]}\',this)">'
                   f'<td><strong>{r["channel"]}</strong></td>'
                   f'<td>{fmt_n(r["hoteles"])}</td>'
                   f'<td><div class="int-wrap"><div class="int-bar">'
                   f'<div class="int-fill" style="width:{pct_bar:.0f}%"></div></div>'
                   f'<span class="int-val">{r["avg_contratos"]} ctr/hotel</span></div></td>'
                   f'<td>{fmt_n(r["destinos"])}</td>'
                   f'<td><div style="display:flex;align-items:center;gap:6px;">'
                   f'<div style="flex:1;height:3px;background:var(--rule,#E8E4DF);border-radius:2px;max-width:60px;">'
                   f'<div style="height:100%;width:{min(100, r.get("avg_destinos",0)/max_avg_dest*100 if max_avg_dest else 0):.0f}%;background:#4FC3F4;border-radius:2px;"></div></div>'
                   f'<span class="int-val">{r.get("avg_destinos",0)} dest/h</span></div></td></tr>')

    t_rows = ''
    for r in ch_tercero:
        cls = ' class="residual"' if r['residual'] else ''
        badge = ' <span style="font-size:9px;color:var(--ink-muted)">Residual</span>' if r['residual'] else ''
        pct_str = f'{r["hoteles"]/N*100:.1f}%' if r["hoteles"] > 0 else '—'
        t_rows += (f'<tr{cls}><td><strong>{r["channel"]}</strong>{badge}</td>'
                   f'<td>{fmt_n(r["hoteles"])}</td>'
                   f'<td>{fmt_n(r["destinos"])}</td>'
                   f'<td style="color:#C0392B;font-weight:700;">{pct_str}</td></tr>')

    return f'''
      <!-- Pills Tipo de Contratación para Channel -->
      <div class="pills-wrap" id="ch-tipo-pills" style="margin-bottom:12px;">
        <button class="pill on" data-tipo="todos" onclick="chSetTipo(this,'todos')">Todos</button>
        <button class="pill"    data-tipo="sp"    onclick="chSetTipo(this,'sp')">Solo Propio</button>
        <button class="pill"    data-tipo="hy"    onclick="chSetTipo(this,'hy')">Hybrid</button>
      </div>
      <!-- Canal overview: dos columnas lado a lado -->
      <div id="ch-overview">

        <div style="display:grid;grid-template-columns:1fr 1fr;gap:24px;">

          <!-- Columna izquierda: Producto Propio -->
          <div>
            <div class="ch-block-label" style="margin-bottom:8px;">Producto Propio</div>
            <table class="ch-table" style="width:100%;">
              <thead><tr>
                <th style="text-align:left;">Channel</th>
                <th>Hoteles</th><th>Avg Ctr</th><th>Destinos</th><th>Avg Dest</th>
              </tr></thead>
              <tbody id="ch-propio-tbody">{p_rows}</tbody>
            </table>
          </div>

          <!-- Columna derecha: Third Party -->
          <div>
            <div class="ch-block-label" style="margin-bottom:8px;">Third Party</div>
            <table class="ch-table" style="width:100%;">
              <thead><tr>
                <th style="text-align:left;">Channel</th>
                <th>Hoteles</th><th>Destinos</th><th>% Gap</th>
              </tr></thead>
              <tbody id="ch-tercero-tbody">{t_rows}</tbody>
            </table>
          </div>

        </div>
        <div style="font-size:10px;color:var(--ink-muted);margin-top:12px;font-style:italic;">
          </div>
      </div>

      <!-- Panel de drill por channel -->
      <div id="ch-drill" style="display:none;">
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;">
          <button onclick="chBack()" style="font-family:inherit;font-size:9px;font-weight:700;
            letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);
            background:var(--paper);color:var(--ink-muted);padding:5px 12px;cursor:pointer;border-radius:3px;">
            ← Volver</button>
          <span id="ch-drill-title" style="font-size:13px;font-weight:700;color:var(--ink);"></span>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">
          <input class="f-input" id="ch-f-region" placeholder="Región…" type="text"
            oninput="chFilter()" autocomplete="off" style="max-width:140px;" list="ch-region-list">
          <datalist id="ch-region-list">{f''.join(f"<option value='{r}'>" for r in REGION_ORDER)}</datalist>
          <select class="f-select" id="ch-f-corp" onchange="chFilter()" style="max-width:180px;">
            <option value="">Todos los corporativos</option>{corp_opts_ch}
          </select>
          <input class="f-input" id="ch-f-dest" placeholder="Buscar destino…"
            type="text" oninput="chFilter()" style="max-width:180px;">
        </div>
        <table class="data-table" id="ch-drill-table">
          <thead><tr>
            <th style="text-align:left;">Hotel</th>
            <th>Región</th><th>Corporativo</th><th>Destino</th>
          </tr></thead>
          <tbody id="ch-drill-tbody"></tbody>
        </table>
        <div id="ch-drill-count" style="font-size:10px;color:var(--ink-muted);margin-top:6px;"></div>
      </div>'''

# ── ZONA 4: SIN CONTRATACIÓN DIRECTA ──
def build_gap_tab():
    """Sin Contratación Directa — misma estructura que tabla unificada."""
    # Rows by current dim (reg / corp / dest) — uses corp_mkt and reg_mkt
    # Show: Sin Directo | Con Directo | Total | % Penetración

    def vs_val(pct_pen):
        return pct_pen - (pp/N*100)

    # Reg rows
    reg_rows = ''
    for r in reg_stats:
        reg=r['region']; tot=r['total']
        sin_d = int(r['solo_tercero'])  # sin contrato directo = solo terceros
        con_d = int(r['prod_propio'])
        pct = con_d/tot*100 if tot else 0
        vs = vs_val(pct)
        reg_rows += (
            f'<tr class="gap-reg-row" style="cursor:pointer"'
            f' onclick="udRowClick(\'region\',\'{reg}\',this)">'
            f'<td><strong>{reg}</strong></td>'
            f'<td class="td-tot">{fmt_n(tot)}</td>'
            f'<td style="color:#C0392B;font-weight:700;">{fmt_n(sin_d)}</td>'
            f'<td class="td-pp">{fmt_n(con_d)}</td>'
            f'<td>{pct_bar_html(pct,"var(--green)")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )

    # Corp rows
    corp_rows = ''
    for i,(_, r) in enumerate(corp_mkt.head(200).iterrows()):
        cls='rows-more-gap-corp' if i>=10 else ''; sty='display:none' if i>=10 else ''
        corp=str(r['Corporativo']).replace("'","")
        tot=int(r['total']); sin_d=int(r['sin_directo']); con_d=int(r['con_directo'])
        pct=r['pct_penetracion']
        vs = vs_val(pct)
        corp_rows += (
            f'<tr class="gap-corp-row {cls}" style="{sty};cursor:pointer" data-row-idx="{i}"'
            f' onclick="udRowClick(\'corp\',\'{corp}\',this)">'
            f'<td><strong>{r["Corporativo"]}</strong></td>'
            f'<td class="td-tot">{fmt_n(tot)}</td>'
            f'<td style="color:#C0392B;font-weight:700;">{fmt_n(sin_d)}</td>'
            f'<td class="td-pp">{fmt_n(con_d)}</td>'
            f'<td>{pct_bar_html(pct,"var(--green)")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )
    ver_mas_corp = ('<tr class="gap-corp-row" style="display:none" id="gap-corp-ver-mas2">'
                    '<td colspan="6" style="text-align:center;padding:10px;">'
                    '<button onclick="gapToggleCorp(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
                    'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
                    'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
                    'Ver 10 más</button></td></tr>')

    # Dest rows
    dest_rows = ''
    for i,(_, r) in enumerate(dest_mkt.head(500).iterrows()):
        cls='rows-more-gap-dest' if i>=10 else ''; sty='display:none' if i>=10 else ''
        tot=int(r['total']); sin_d=int(r['sin_directo']); con_d=int(r['con_directo'])
        pct=r['pct_penetracion']
        vs = vs_val(pct)
        dest_rows += (
            f'<tr class="gap-dest-row {cls}" style="{sty}" data-row-idx="{i}" data-region="{r["Region_display"]}">'
            f'<td><strong>{r["Destino"]}</strong>'
            f'<div style="font-size:10px;color:var(--ink-muted);line-height:1.3;margin-top:1px;">{r["Region_display"]}</div></td>'
            f'<td class="td-tot">{fmt_n(tot)}</td>'
            f'<td style="color:#C0392B;font-weight:700;">{fmt_n(sin_d)}</td>'
            f'<td class="td-pp">{fmt_n(con_d)}</td>'
            f'<td>{pct_bar_html(pct,"var(--green)")}</td>'
            f'<td>{vs_bar_html(vs)}</td>'
            f'</tr>'
        )
    ver_mas_dest = ('<tr class="gap-dest-row" style="display:none" id="gap-dest-ver-mas2">'
                    '<td colspan="6" style="text-align:center;padding:10px;">'
                    '<button onclick="gapToggleDest(this)" style="font-family:inherit;font-size:10px;font-weight:700;'
                    'letter-spacing:.08em;text-transform:uppercase;border:1px solid var(--rule);background:var(--paper);'
                    'color:var(--ink-muted);padding:6px 18px;cursor:pointer;border-radius:3px;" data-open="0">'
                    'Ver 10 más</button></td></tr>')

    region_opts_gap = ''.join(f'<option value="{r}">{r}</option>' for r in REGION_ORDER)

    return f"""
    <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;">
      <select class="f-select" id="gap-f-region" onchange="gapFilter()" style="max-width:150px;display:none;">
        <option value="">Todas las regiones</option>{region_opts_gap}
      </select>
      <input class="f-input" id="gap-search" placeholder="Buscar..." type="text" autocomplete="off"
        oninput="gapFilter()" style="max-width:220px;display:none;">
    </div>
    <table class="data-table">
      <thead><tr>
        <th id="gap-dim-th" style="text-align:left;width:220px;">Dimensión</th>
        <th>Total</th>
        <th style="color:#C0392B;">Sin Directo</th>
        <th class="th-pp">Con Directo</th>
        <th style="min-width:120px;">Penetración</th>
        <th>vs Global</th>
      </tr></thead>
      <tbody id="gap-tbody">
        <tr class="global-row">
          <td>GLOBAL</td>
          <td>{fmt_n(N)}</td>
          <td style="color:#C0392B;font-weight:700;">{fmt_n(int(market_tp))}</td>
          <td class="td-pp">{fmt_n(pp)}</td>
          <td>{pct_bar_html(pp/N*100,"var(--green)")}</td>
          <td>—</td>
        </tr>
        {reg_rows}
        {corp_rows}{ver_mas_corp}
        {dest_rows}{ver_mas_dest}
      </tbody>
    </table>"""

# ── ZONA 5: HISTÓRICO ──
def build_historical():
    year_opts       = ''.join(f'<option value="{y}"{"  selected" if y==2026 else ""}>{y}</option>' for y in years_available)
    reg_opts        = ''.join(f'<option value="{r}">{r}</option>' for r in hist_regions)
    corp_opts       = ''.join(f'<option value="{c}">{c}</option>' for c in hist_corps)
    ch_propio_opts  = ''.join(f'<option value="{c}">{c}</option>' for c in hist_channels_propio)
    ch_tercero_opts = ''.join(f'<option value="{c}">{c}</option>' for c in hist_channels_tercero)
    return f'''
    <div class="hist-controls">
      <!-- Fila 1: Toggle Por Año/Mes/Semana -->
      <div style="display:flex;gap:12px;align-items:center;margin-bottom:8px;">
        <div class="drill-toggle">
          <button class="dt-btn" id="dt-anio" onclick="hSetLevel('anio')">Por Año</button>
          <button class="dt-btn" id="dt-mes"  onclick="hSetLevel('mes')">Por Mes</button>
          <button class="dt-btn on" id="dt-sem"  onclick="hSetLevel('sem')">Por Semana</button>
        </div>
      </div>
      <!-- Fila 2: Dropdowns Año/Mes/Semana -->
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:10px;">
        <div class="f-group"><span class="f-label">Año</span>
          <select class="f-select" id="sel-year" onchange="hSelYear(this.value);hUpdateComboStyle('sel-year')">
            <option value="">—</option>{year_opts}
          </select></div>
        <div class="f-group"><span class="f-label">Mes</span>
          <select class="f-select" id="sel-month" disabled onchange="hSelMonth(this.value)">
            <option value="">Todos</option>
          </select></div>
        <div class="f-group"><span class="f-label">Semana</span>
          <select class="f-select" id="sel-week">
            <option value="">—</option>
          </select></div>
      </div>
      <!-- Fila 2: Pills de Región -->
      <div class="filters-row">
        <div style="width:100%;">
          <span class="f-label">Región</span>
          <div class="pills-wrap" id="hf-region-pills">
            <button class="pill on" data-region="" onclick="hPillRegion(this)">Todas</button>
            {''.join(f'<button class="pill" data-region="{r}" onclick="hPillRegion(this)">{r}</button>' for r in hist_regions)}
          </div>
        </div>
        <!-- Fila 3: Tipo, Channel, Corporativo -->
        <div style="display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap;margin-top:8px;width:100%;">

          <div>
            <span class="f-label">Channel</span>
            <select class="f-select" id="hf-channel" onchange="if(this.value===hFChannel){{hFChannel='';this.value='';hUpdateComboStyle('hf-channel');hFilterCorpByChannel('');hApplyFilter();}}else{{hFChannel=this.value;hFilterCorpByChannel(this.value);hUpdateComboStyle('hf-channel');hApplyFilter();}}" style="max-width:150px;">
              <option value="">Todos</option>
              <optgroup label="Producto Propio">{ch_propio_opts}</optgroup>
              <optgroup label="Third Party">{ch_tercero_opts}</optgroup>
            </select>
          </div>
          <div>
            <span class="f-label">Corporativo</span>
            <select class="f-select" id="hf-corp" onchange="if(this.value===hFCorp){{hFCorp='';this.value='';hUpdateComboStyle('hf-corp');hApplyFilter();}}else{{hFCorp=this.value;hUpdateComboStyle('hf-corp');hApplyFilter();}}" style="max-width:160px;">
              <option value="">Todos</option>{corp_opts}
              <option value="Otros">Otros</option>
            </select>
          </div>
          <button class="clear-link" id="btn-limpiar" onclick="hClearFilters()" style="display:none;">✕ Limpiar</button>
        </div>
      </div>
    </div>
    <div id="hf-active-pills" style="display:flex;gap:6px;flex-wrap:wrap;margin:8px 0 4px;min-height:0;"></div>
    <div class="breadcrumb" id="drill-bc"></div>
    <div class="chart-area" style="background:transparent;border:1px solid var(--rule);border-radius:2px;padding:16px 12px 8px;">
      <div style="position:relative;height:220px;width:100%;"><canvas id="canvas-hist"></canvas></div>
    </div>'''

# ── ENSAMBLADO ──
def build_html():
    reg_tabs = build_region_tabs()
    corp_tab = build_corp_tab()
    dest_tab = build_dest_tab()
    ch_tab   = build_channel_tab()
    gap_tab  = build_gap_tab()
    hist_tab = build_historical()

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Hotel Inventory · {WEEK} · {YEAR_ACTUAL}</title>
<link href="https://fonts.googleapis.com/css2?family=Geist:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<div class="shell">

<!-- MASTHEAD — estructura idéntica a CR/RND -->
<div style="height:12px;"></div>
<div class="masthead-top-rule"></div>
<div style="display:table;width:100%;padding:18px 0 14px;border-bottom:1px solid var(--rule);">
  <div style="display:table-cell;vertical-align:middle;">
    <div style="font-size:26px;font-weight:800;letter-spacing:-.02em;color:var(--ink);line-height:1;">Week {WEEK_NUM}</div>
    <div style="font-size:12px;font-weight:400;color:var(--ink-muted);margin-top:3px;">{SNAPSHOT_DATE}</div>
  </div>
  <div style="display:table-cell;vertical-align:middle;text-align:right;white-space:nowrap;">
    <img alt="PriceTravel" src="{LOGO_B64}" class="dark-invert" style="height:44px;width:auto;vertical-align:middle;"/>
    <span style="display:inline-block;width:1px;height:38px;background:var(--rule);margin:0 12px;vertical-align:middle;"></span>
    <span style="display:inline-block;vertical-align:middle;text-align:left;line-height:1.15;">
      <span style="display:block;font-size:20px;font-weight:400;letter-spacing:-.01em;color:var(--ink);">Supply Optimization</span>
    </span>
  </div>
</div>
<div class="masthead-sub">
  <span>{SNAPSHOT_DATE}</span>
  <span>Vol. {VOL_NUM}</span>
</div>

<!-- HERO -->
<div class="hero">
  <h1>State of <span class="accent">PriceTravel Product</span></h1>
  <div class="hero-sub"><strong style="color:var(--accent)">{fmt_n(N)}</strong> hoteles con contrato activo · Target 2026: <strong style="color:var(--accent)">{fmt_n(TARGET_PROPIO)}</strong> con Producto Propio</div>
</div>

<!-- KPI BAR — 4 cards en una fila -->
<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;border:1px solid var(--rule);
  background:var(--paper-soft);margin-bottom:36px;width:100%;">

  <!-- Card 1: Hotel Inventory -->
  <div style="padding:20px 22px;border-right:1px solid var(--rule);display:flex;flex-direction:column;gap:14px;">

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Total Hotel Inventory</div>
      <div style="font-size:28px;font-weight:700;color:var(--accent);letter-spacing:-.02em;" id="card-total">{fmt_n(N)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div style="height:100%;width:100%;background:var(--accent);border-radius:2px;opacity:.3;"></div></div>
      <div style="font-size:9px;color:var(--ink-muted);margin-top:3px;">Hoteles con contratos activos</div>
    </div>

    <div style="border-top:1px solid var(--rule-soft);"></div>

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Producto Propio</div>
      <div style="font-size:28px;font-weight:700;color:var(--ink);letter-spacing:-.02em;" id="card-pp">{fmt_n(pp)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div id="bar-pp" style="height:100%;width:{pp/N*100:.1f}%;background:var(--violet);border-radius:2px;"></div></div>
      <div style="font-size:11px;font-weight:700;color:var(--accent);margin-top:3px;" id="card-pp-pct">{fmt_pct(pp/N*100)}</div>
    </div>

    <div style="border-top:1px solid var(--rule-soft);"></div>

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Third Party</div>
      <div style="font-size:28px;font-weight:700;color:var(--ink);letter-spacing:-.02em;" id="card-tp">{fmt_n(solo_terc)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div id="bar-tp" style="height:100%;width:{solo_terc/N*100:.1f}%;background:var(--dgrey);border-radius:2px;"></div></div>
      <div style="font-size:11px;font-weight:700;color:var(--accent);margin-top:3px;" id="card-tp-pct">{fmt_pct(solo_terc/N*100)}</div>
    </div>

  </div>

  <!-- Card 2: Target 2026 -->
  <div style="padding:20px 22px;border-right:1px solid var(--rule);display:flex;flex-direction:column;gap:14px;">

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Target 2026</div>
      <div style="font-size:28px;font-weight:700;color:var(--accent);letter-spacing:-.02em;">{fmt_n(TARGET_PROPIO)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div style="height:100%;width:100%;background:var(--accent);border-radius:2px;opacity:.3;"></div></div>
      <div style="font-size:9px;color:var(--ink-muted);margin-top:3px;">Producto Propio 2026</div>
    </div>

    <div style="border-top:1px solid var(--rule-soft);"></div>

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Avance</div>
      <div style="font-size:28px;font-weight:700;color:var(--green);letter-spacing:-.02em;" id="card-avance">{fmt_pct(pct_avance)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div id="prog-fill" style="height:100%;width:{pct_avance:.1f}%;background:var(--green);border-radius:2px;"></div></div>
      <div style="font-size:11px;font-weight:700;color:var(--green);margin-top:3px;">{fmt_n(pp)} hoteles</div>
    </div>

    <div style="border-top:1px solid var(--rule-soft);"></div>

    <div>
      <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:4px;">Gap</div>
      <div style="font-size:28px;font-weight:700;color:#C0392B;letter-spacing:-.02em;" id="card-gap">{fmt_n(gap)}</div>
      <div style="height:3px;background:var(--rule-soft);border-radius:2px;margin-top:6px;"><div style="height:100%;width:{(100-pct_avance):.1f}%;background:#C0392B;border-radius:2px;"></div></div>
      <div style="font-size:11px;font-weight:700;color:#C0392B;margin-top:3px;">Pendientes para alcanzar el target</div>
    </div>

  </div>

  <!-- Card 3: Gap por Corporativo -->
  <div style="padding:20px 22px;border-right:1px solid var(--rule);">
    <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:16px;">Gap por Corporativo</div>
    {''.join(
      f'<div style="margin-bottom:10px;">'
      f'<div style="font-size:11px;font-weight:700;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:3px;">{r["Corporativo"]}</div>'
      f'<div style="display:flex;align-items:center;gap:6px;">'
      f'<div style="flex:1;height:3px;background:var(--rule-soft);border-radius:2px;">'
      f'<div style="height:100%;width:{min(100,100-int(r["pct_penetracion"])):.1f}%;background:#C0392B;border-radius:2px;opacity:.7;"></div>'
      f'</div>'
      f'<span style="font-size:13px;font-weight:700;color:var(--ink-muted);flex-shrink:0;min-width:46px;text-align:right;">{100-int(r["pct_penetracion"]):.1f}%</span>'
      f'</div>'
      f'<div style="font-size:11px;font-weight:700;color:#C0392B;margin-top:2px;">{fmt_n(int(r["sin_directo"]))}</div>'
      f'</div>'
      for _, r in corp_mkt[corp_mkt['Corporativo']!='AA-Independent'].head(7).iterrows()
    )}
  </div>

  <!-- Card 4: Gap por Región -->
  <div style="padding:20px 22px;">
    <div style="font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-muted);margin-bottom:16px;">Gap por Región</div>
    {''.join(
      f'<div style="margin-bottom:10px;">'
      f'<div style="font-size:11px;font-weight:700;color:var(--ink);margin-bottom:3px;">{r["Region_display"]}</div>'
      f'<div style="display:flex;align-items:center;gap:6px;">'
      f'<div style="flex:1;height:3px;background:var(--rule-soft);border-radius:2px;">'
      f'<div style="height:100%;width:{min(100,r["pct_sin"]):.1f}%;background:#C0392B;border-radius:2px;opacity:.7;"></div>'
      f'</div>'
      f'<span style="font-size:13px;font-weight:700;color:var(--ink-muted);flex-shrink:0;min-width:46px;text-align:right;">{r["pct_sin"]:.1f}%</span>'
      f'</div>'
      f'<div style="font-size:11px;font-weight:700;color:#C0392B;margin-top:2px;">{fmt_n(int(r["sin_directo"]))}</div>'
      f'</div>'
      for _, r in reg_mkt.iterrows()
    )}
  </div>

</div>
<div class="kpi-note" id="kpi-note"></div>

<!-- ZONA 5: HISTÓRICO -->
<div class="sec-head">
  <span class="sec-title">Crecimiento Histórico · Producto Propio</span>
</div>
{hist_tab}

<!-- DISTRIBUCIÓN Y EXPLORACIÓN — tabla única -->
<div style="margin-top:32px;margin-bottom:36px;">
  <div style="border-top:1px solid var(--rule);padding-top:20px;margin-bottom:14px;">
    <span class="sec-title" style="font-size:11px;letter-spacing:.1em;text-transform:uppercase;color:var(--ink-muted);font-weight:700;">Distribución y Exploración</span>
  </div>

  <!-- Fila 1: Dimensiones -->
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:8px;">
    <div class="pills-wrap" id="ud-dim-pills">
      <button class="pill on" onclick="udSetDim('reg',this)">Región</button>
      <button class="pill"    onclick="udSetDim('corp',this)">Corporativo</button>
      <button class="pill"    onclick="udSetDim('dest',this)">Destino</button>
      <button class="pill"    onclick="udSetDim('ch',this)">Channel</button>
    </div>
  </div>

  <!-- Fila 2: Métricas + pills activas (todo en una línea) -->
  <div id="ud-metric-pills" style="display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:18px;padding-bottom:10px;border-bottom:1px solid var(--rule-soft,#E8E4DF);">
    <div class="pills-wrap distrib-pills" style="border:none;padding:0;">
      <button class="pill metric-pill on" data-col="all" style="font-size:9px;--pill-on-bg:#E8F5EE;--pill-on-fg:#1A6B4A;--pill-on-bd:#1A6B4A;" onclick="udContent('all',this)">Todos</button>
      <button class="pill metric-pill" data-col="pp" style="font-size:9px;--pill-on-bg:#E8F5EE;--pill-on-fg:#1A6B4A;--pill-on-bd:#1A6B4A;" onclick="udContent('pp',this)">Producto Propio</button>
      <button class="pill metric-pill" data-col="sp" style="font-size:9px;--pill-on-bg:#E8F5EE;--pill-on-fg:#1A6B4A;--pill-on-bd:#1A6B4A;" onclick="udContent('sp',this)">Solo Propio</button>
      <button class="pill metric-pill" data-col="hy" style="font-size:9px;--pill-on-bg:#E8F5EE;--pill-on-fg:#1A6B4A;--pill-on-bd:#1A6B4A;" onclick="udContent('hy',this)">Hybrid</button>
      <button class="pill metric-pill gap-pill" data-col="gap" id="btn-gap" style="font-size:9px;--pill-on-bg:#FEE2E2;--pill-on-fg:#C0392B;--pill-on-bd:#C0392B;border-color:#C0392B;color:#C0392B;" onclick="udContent('gap',this)">Sin Contrat. Directa</button>
    </div>
    <div id="ud-active-badge-container" style="display:flex;gap:6px;flex-wrap:wrap;align-items:center;"></div>
  </div>

  <!-- Contenido -->
  <div id="ud-main-content">{build_unified_distrib()}</div>
  <div id="ud-gap-content" style="display:none;">{gap_tab}</div>
  <div id="ud-ch-content"  style="display:none;">{ch_tab}</div>

</div>

</div><!-- .shell -->
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>{JS}</script>
<script>
makeAutocomplete('corp-search','corp-ac-list', CORP_DATA,'Corporativo', name => {{ updateCards({{type:'corp',name}}); }});
makeAutocomplete('dest-search','dest-ac-list', DEST_DATA,'Destino', name => {{ updateCards({{type:'dest',name}}); }});
function _tryInit() {{
  if (typeof Chart === 'undefined') {{ setTimeout(_tryInit, 30); return; }}
  requestAnimationFrame(function() {{
    requestAnimationFrame(function() {{
      // Verify canvas has rendered dimensions before init — retry if not
      const cv = document.getElementById('canvas-hist');
      if (!cv || cv.offsetWidth === 0) {{ setTimeout(_tryInit, 50); return; }}
      hInit();
      setTimeout(function() {{ if (hChart) hChart.resize(); }}, 100);
      ['sel-year','sel-month','sel-week','hf-channel','hf-corp'].forEach(id => hUpdateComboStyle(id));
      const regBtn = document.querySelector('#ud-dim-pills .pill');
      const ppBtn  = document.querySelector('.distrib-pills .pill:not(.gap-pill)');
      if (regBtn) udSetDim('reg', regBtn);
      // Default to showing all columns
      const allBtn = document.querySelector('.distrib-pills .pill[data-col="all"]');
      if (allBtn) udContent('all', allBtn);
    }});
  }});
}}
_tryInit();
</script>
<div style="width:100%;margin:40px 0 0;padding:20px 40px;background:var(--paper);border-top:1px solid var(--rule);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:16px;box-sizing:border-box;">
  <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
    <span style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.12em;color:var(--ink-muted);">Descargas {WEEK}</span>
    <a href="Analisis_Inventory_W{WEEK_NUM}.xlsx" download style="font-size:11px;font-weight:700;color:#fff;text-decoration:none;padding:7px 16px;background:var(--ink);border-radius:3px;white-space:nowrap;">⬇ Excel Inventory</a>
  </div>
</div>
</body>
</html>"""

# ─────────────────────────────────────────────
# 5. OUTPUT
# ─────────────────────────────────────────────
print("[5/5] Escribiendo output...")
html = build_html()
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
Path(OUTPUT_DIR / OUTPUT_FILE).write_text(html, encoding='utf-8')
print(f"\n✅ {OUTPUT_FILE}")
print(f"   Tamaño: {len(html)/1024:.0f} KB")
# ─────────────────────────────────────────────
# 6. EXCEL ANALISIS
# ─────────────────────────────────────────────
print("[6/6] Generando Excel de análisis...")

EXCEL_FILE = OUTPUT_FILE.replace('.html', '.xlsx').replace('INVENTORY_', 'Analisis_Inventory_')

# Base columns for hotel-level detail
COLS_BASE = ['Hotel', 'Region_display', 'Corporativo', 'Destino', 'TipoHotel']

# Build channel columns list (active channels only)
ch_cols_active = [c for c in ALL_CHANNELS if c in df.columns]

# ── Hoja Resumen ──
resumen_data = {
    'Métrica': ['Universo con contrato', 'Producto Propio', 'Solo Propio', 'Hybrid',
                'Third Party', 'Target 2026', 'Avance PP', 'Gap restante',
                'Semanas restantes', 'Ritmo necesario (hoteles/semana)'],
    'Valor': [int(N), int(pp), int(solo_propio), int(hybrid),
              int(solo_terc), int(TARGET_PROPIO), f'{pp/TARGET_PROPIO*100:.1f}%', int(gap),
              int(SEMANAS_RESTANTES), int(math.ceil(gap/SEMANAS_RESTANTES)) if SEMANAS_RESTANTES > 0 else 0]
}
df_resumen = pd.DataFrame(resumen_data)

# ── Hoja Por Región ──
df_region = df[COLS_BASE + ch_cols_active].copy()
df_region = df_region.rename(columns={'Region_display': 'Región'})
# Add channel flags as Yes/No
for c in ch_cols_active:
    df_region[CHANNEL_LABELS.get(c, c)] = df_region[c].apply(lambda v: 'Sí' if channel_active(v) else 'No')
    df_region = df_region.drop(columns=[c])
df_region = df_region.sort_values('Región')

# ── Hoja Por Corporativo ──
df_corp = df_region.copy().sort_values(['Corporativo', 'Hotel'])

# ── Hoja Por Destino ──
df_dest = df_region.copy().sort_values(['Destino', 'Hotel'])

# ── Hoja Por Channel ──
# One row per hotel × active channel
ch_rows = []
for c in ch_cols_active:
    ch_label = CHANNEL_LABELS.get(c, c)
    sub = df[df[c].apply(channel_active)][COLS_BASE].copy()
    sub['Channel'] = ch_label
    ch_rows.append(sub)
df_channel = pd.concat(ch_rows, ignore_index=True) if ch_rows else pd.DataFrame()
df_channel = df_channel.rename(columns={'Region_display': 'Región'})
df_channel = df_channel[['Hotel', 'Channel', 'Región', 'Corporativo', 'Destino', 'TipoHotel']]
df_channel = df_channel.sort_values(['Channel', 'Corporativo', 'Hotel'])

# ── Write Excel ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

def write_sheet(wb, name, df, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    ws.title = name
    # Header row
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='333132')
        c.alignment = Alignment(horizontal='center')
    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    # Column widths
    for ci, col in enumerate(df.columns, 1):
        max_w = max(len(str(col)), max((len(str(v or '')) for v in df.iloc[:,ci-1]), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 45)
    return ws

write_sheet(wb, 'Resumen',         df_resumen, first=True)
write_sheet(wb, 'Por Región',      df_region)
write_sheet(wb, 'Por Corporativo', df_corp)
write_sheet(wb, 'Por Destino',     df_dest)
if not df_channel.empty:
    write_sheet(wb, 'Por Channel', df_channel)

wb.save(OUTPUT_DIR / EXCEL_FILE)

print(f"✅ {EXCEL_FILE}")
print(f"   Hoteles: {len(df_region):,} · Hojas: Resumen + Por Región/Corp/Destino/Channel")

