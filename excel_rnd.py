"""
excel_rnd.py · W26+ · Excel Rates No Dispo — estructura simplificada
5 hojas: Severity | Maestra | Críticos | Bajo Rendimiento | Sin Conversión

Severity: cada sección es una Tabla Excel (filtro propio independiente).
Maestra: una fila por Hotel × Canasta — Banda ND + Banda CV
Bandas: Banda ND Global + Banda CV Global coloreadas; exposición cruzada en texto plano.
IPM eliminado (W26+)
"""
import pickle, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_nodispo, banda_convrate, get_dest_tier

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_RND', f'rnd_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D = pickle.load(f)

TAB_ND  = D['TAB_NoDispo']
CANASTA = D['CANASTA']
M       = D['M']

RND = 'EA0074'

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP   = fill('EAF3DE'); WOW_UP_F   = fnt('2F6C34', bold=True)
WOW_DN   = fill('FCE8E6'); WOW_DN_F   = fnt('C0392B', bold=True)
WOW_NEU  = fill('F2EEE6'); WOW_NEU_F  = fnt('8A8377', bold=True)

RANGOS_ND = {
    'Exitosa':       '< 3%',
    'Aceptable':     '3% – 5%',
    'Revisar':       '5% – 20%',
    'Crítica':       '20% – 60%',
    'Súper Crítica': '> 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0',
    'Crítica':        '< 0,8%',
    'Revisar':        '0,8% – 1,5%',
    'Aceptable':      '1,5% – 2,5%',
    'Exitosa':        '≥ 2,5%',
}

_table_counter = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(RND, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(RND, 11, True)

def mk_hdr_plain(ws, row, cols):
    """Header con color pero SIN auto_filter — para tablas Excel."""
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(RND)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_excel_table(ws, hdr_row, data_row_end, n_cols, name_prefix):
    """Crea una Tabla Excel con filtro propio sobre el rango hdr_row:data_row_end."""
    _table_counter[0] += 1
    tname = f'T_{name_prefix}_{_table_counter[0]}'
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{data_row_end}'
    t = Table(displayName=tname, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center', fmt=None):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if fmt: cell.number_format = fmt
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

# ── Fuentes de datos por canasta ──────────────────────────────────────────────
CANASTAS_DEF = [
    ('global', 'Global',      None),
    ('b2c',    'B2C',         'B2C'),
    ('op',     'Opaco',       'B2B-OP'),
    ('cug',    'Ultra Opaco', 'CUG'),
]

def hotel_src(can_id):
    if can_id is None:
        return D.get('p80_hotel', pd.DataFrame())
    can = CANASTA.get(can_id, {})
    src = can.get('p80_hotel')
    if src is None: src = can.get('p80')
    return src if src is not None else TAB_ND.get('hotel', pd.DataFrame())

def build_banda_lookup():
    lookup = {}
    for _, can_label, can_id in CANASTAS_DEF:
        if can_id is None: continue
        df = hotel_src(can_id)
        if df is None or len(df) == 0: continue
        for _, row in df.iterrows():
            h   = str(row.get('Hotel', ''))
            nd  = sf(row.get('%NoDispo'))
            bk  = int(sf(row.get('Bookings')) or 0)
            trf = int(sf(row.get('Trafico')) or 0)
            conv = bk / trf if trf > 0 else None
            b_nd = banda_nodispo(nd) if nd is not None else '—'
            b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            if h not in lookup: lookup[h] = {}
            lookup[h][can_label] = {'nd': b_nd, 'cv': b_cv}
    return lookup

def band_split_nd(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['%NoDispo'].apply(lambda v: banda_nodispo(sf(v)) if sf(v) is not None else '—')
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Hoja Severity ─────────────────────────────────────────────────────────────
def write_severity(ws):
    title(ws, f'Severity RND W{VOL_NUM}', f'Resumen ejecutivo · {PERIODO}')
    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))

    nd_curr = m_curr.get('pct_nodispo', 0); nd_prev = m_prev.get('pct_nodispo', 0)
    nd_wow  = (nd_curr - nd_prev) * 100 if nd_prev else None

    # KPI Global (sin tabla — es solo 1 fila de datos, no necesita filtro)
    r = 4
    section_title(ws, r, 'KPI Global'); r += 1
    r = mk_hdr_plain(ws, r, ['Métrica', f'W{int(VOL_NUM)-1}', f'W{VOL_NUM}', 'WoW'])
    ws.cell(r, 1, '%NoDispo').font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
    ws.cell(r, 2, round(nd_prev, 4) if nd_prev else None).border = BD
    ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
    ws.cell(r, 3, round(nd_curr, 4) if nd_curr else None).border = BD
    ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
    apply_wow(ws, r, 4, nd_wow, invert=True); r += 2

    # Calcular severities
    df_hotel = D.get('p80_hotel', pd.DataFrame())
    sev_nd, sev_cv = {}, {}
    for _, row in df_hotel.iterrows():
        nd  = sf(row.get('%NoDispo'))
        bk  = int(sf(row.get('Bookings')) or 0)
        trf = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        if nd is not None: b = banda_nodispo(nd); sev_nd[b] = sev_nd.get(b, 0) + 1
        b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else None)
        if b_cv: sev_cv[b_cv] = sev_cv.get(b_cv, 0) + 1
    total_nd = sum(sev_nd.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    # Severity %NoDispo — Tabla Excel con filtro propio
    section_title(ws, r, 'Severity %NoDispo · Hoteles Global'); r += 1
    hdr_r = r; r = mk_hdr_plain(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_nd.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True); ws.cell(r, 1).border = BD
        mk_cell(ws, r, 2, RANGOS_ND.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_nd, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_excel_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevND')
    r += 1

    # Severity %Conv Rate — Tabla Excel con filtro propio
    section_title(ws, r, 'Severity %Conv Rate · Hoteles Global'); r += 1
    hdr_r = r; r = mk_hdr_plain(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True); ws.cell(r, 1).border = BD
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_excel_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    # Top Destinos — Tabla Excel con filtro propio
    DIM_COLS = ['Destino', 'Banda ND', 'Banda CV', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    df_dest = TAB_ND.get('destino', pd.DataFrame())
    if df_dest is not None and len(df_dest) > 0:
        section_title(ws, r, 'Top Destinos · %NoDispo DESC'); r += 1
        hdr_r = r; r = mk_hdr_plain(ws, r, DIM_COLS)
        for _, row in df_dest.sort_values('%NoDispo', ascending=False).head(30).iterrows():
            nd   = sf(row.get('%NoDispo'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            conv = bk / trf if trf > 0 else None
            b_nd = banda_nodispo(nd) if nd is not None else '—'
            b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            mk_cell(ws, r, 1, str(row.get('Destino', '—')), align='left')
            mk_cell(ws, r, 2, b_nd, b_nd, is_sev=True)
            mk_cell(ws, r, 3, b_cv, b_cv, is_sev=True)
            mk_cell(ws, r, 4, trf)
            pct_cell(ws, r, 5, nd)
            apply_wow(ws, r, 6, sf(row.get('NoDispo_WoW_pp')), invert=True)
            pct_cell(ws, r, 7, conv)
            mk_cell(ws, r, 8, bk)
            r += 1
        add_excel_table(ws, hdr_r, r-1, len(DIM_COLS), 'Dest')
        r += 1

    # Top Corp — Tabla Excel con filtro propio
    CORP_COLS = ['Corporativo', 'Banda ND', 'Banda CV', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    df_corp = TAB_ND.get('corp', pd.DataFrame())
    if df_corp is not None and len(df_corp) > 0:
        section_title(ws, r, 'Top Corporativos · %NoDispo DESC'); r += 1
        hdr_r = r; r = mk_hdr_plain(ws, r, CORP_COLS)
        for _, row in df_corp.sort_values('%NoDispo', ascending=False).head(30).iterrows():
            nd   = sf(row.get('%NoDispo'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            conv = bk / trf if trf > 0 else None
            b_nd = banda_nodispo(nd) if nd is not None else '—'
            b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            mk_cell(ws, r, 1, str(row.get('CorpName', '—')), align='left')
            mk_cell(ws, r, 2, b_nd, b_nd, is_sev=True)
            mk_cell(ws, r, 3, b_cv, b_cv, is_sev=True)
            mk_cell(ws, r, 4, trf)
            pct_cell(ws, r, 5, nd)
            apply_wow(ws, r, 6, sf(row.get('NoDispo_WoW_pp')), invert=True)
            pct_cell(ws, r, 7, conv)
            mk_cell(ws, r, 8, bk)
            r += 1
        add_excel_table(ws, hdr_r, r-1, len(CORP_COLS), 'Corp')

    autofit(ws, [28, 16, 16, 10, 10, 10, 10, 10])

# ── Hoja Maestra ──────────────────────────────────────────────────────────────
MAESTRA_COLS = [
    'País', 'Destino', 'Clasificación Destino', 'Corporativo', 'Hotel',
    'Canasta', 'Tráfico', 'WoW Tráfico',
    '%NoDispo', 'WoW ND', 'Banda ND',
    '%Conv', 'WoW CV', 'Banda CV', 'Bookings'
]

def write_maestra(ws):
    title(ws, f'Maestra RND · Rates No Dispo W{VOL_NUM}',
          f'Una fila por Hotel × Canasta · {PERIODO} · Filtrar por cualquier columna')
    hdr_r = 4
    r = mk_hdr_plain(ws, hdr_r, MAESTRA_COLS)

    for _, can_label, can_id in CANASTAS_DEF:
        df = hotel_src(can_id)
        if df is None or len(df) == 0: continue
        for _, row in df.iterrows():
            nd   = sf(row.get('%NoDispo'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            conv = bk / trf if trf > 0 else None
            wow_nd = sf(row.get('NoDispo_WoW_pp'))
            wow_cv = sf(row.get('ConvRate_WoW_pp') if 'ConvRate_WoW_pp' in row.index else None)
            bnd_nd = banda_nodispo(nd) if nd is not None else '—'
            bnd_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            pais    = str(row.get('PaisDestino', row.get('Pais', '—')))
            destino = str(row.get('Destino', '—'))
            corp    = str(row.get('CorpName', row.get('Corp', '—')))
            hotel   = str(row.get('Hotel', '—'))
            tier    = get_dest_tier(destino).capitalize()
            for ci, val in enumerate([pais, destino, tier, corp, hotel], 1):
                mk_cell(ws, r, ci, val, align='left')
            mk_cell(ws, r, 6, can_label)
            mk_cell(ws, r, 7, trf)
            apply_wow(ws, r, 8, sf(row.get('Trafico_WoW_pct')), invert=False)
            pct_cell(ws, r, 9, nd)
            apply_wow(ws, r, 10, wow_nd, invert=True)
            mk_cell(ws, r, 11, bnd_nd, bnd_nd, is_sev=True)
            pct_cell(ws, r, 12, conv)
            apply_wow(ws, r, 13, wow_cv, invert=False)
            mk_cell(ws, r, 14, bnd_cv, bnd_cv, is_sev=True)
            mk_cell(ws, r, 15, bk)
            r += 1

    add_excel_table(ws, hdr_r, r-1, len(MAESTRA_COLS), 'Maestra')
    autofit(ws, [16, 22, 14, 22, 40, 14, 10, 10, 10, 10, 18, 10, 10, 18, 10])

# ── Hojas de banda ────────────────────────────────────────────────────────────
BANDA_COLS = [
    'País', 'Destino', 'Clasificación Destino', 'Corporativo', 'Hotel',
    'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'WoW CV', 'Bookings',
    'Banda ND Global', 'Banda CV Global',
    'ND B2C', 'ND Opaco', 'ND Ultra Opaco',
    'CV B2C', 'CV Opaco', 'CV Ultra Opaco',
]

def write_banda(ws, df_banda, sheet_title, banda_lookup):
    title(ws, sheet_title, f'Hoteles Global · {PERIODO} · Top 500')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return

    df_s = df_banda.sort_values('%NoDispo', ascending=False).head(500)
    hdr_r = 4
    r = mk_hdr_plain(ws, hdr_r, BANDA_COLS)

    for _, row in df_s.iterrows():
        nd   = sf(row.get('%NoDispo'))
        bk   = int(sf(row.get('Bookings')) or 0)
        trf  = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        wow_nd = sf(row.get('NoDispo_WoW_pp'))
        wow_cv = sf(row.get('ConvRate_WoW_pp') if 'ConvRate_WoW_pp' in row.index else None)
        bnd_nd = banda_nodispo(nd) if nd is not None else '—'
        bnd_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
        hotel   = str(row.get('Hotel', '—'))
        pais    = str(row.get('PaisDestino', row.get('Pais', '—')))
        destino = str(row.get('Destino', '—'))
        corp    = str(row.get('CorpName', row.get('Corp', '—')))
        tier    = get_dest_tier(destino).capitalize()
        exp     = banda_lookup.get(hotel, {})

        for ci, val in enumerate([pais, destino, tier, corp, hotel], 1):
            mk_cell(ws, r, ci, val, align='left')
        mk_cell(ws, r, 6, trf)
        pct_cell(ws, r, 7, nd)
        apply_wow(ws, r, 8, wow_nd, invert=True)
        pct_cell(ws, r, 9, conv)
        apply_wow(ws, r, 10, wow_cv, invert=False)
        mk_cell(ws, r, 11, bk)
        # Global: coloreadas
        mk_cell(ws, r, 12, bnd_nd, bnd_nd, is_sev=True)
        mk_cell(ws, r, 13, bnd_cv, bnd_cv, is_sev=True)
        # Exposición cruzada: texto plano (sin color)
        for ci, can_label in enumerate(['B2C', 'Opaco', 'Ultra Opaco'], 14):
            mk_cell(ws, r, ci, exp.get(can_label, {}).get('nd', '—'))
        for ci, can_label in enumerate(['B2C', 'Opaco', 'Ultra Opaco'], 17):
            mk_cell(ws, r, ci, exp.get(can_label, {}).get('cv', '—'))
        r += 1

    add_excel_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [16, 22, 14, 22, 40, 10, 10, 10, 10, 10, 10, 16, 16, 12, 12, 14, 12, 12, 14])

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
banda_lookup = build_banda_lookup()
df_global = hotel_src(None)
crit, bajo, sinc = band_split_nd(df_global)

ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = RND
write_severity(ws)
ws = wb.create_sheet('Maestra');          ws.sheet_properties.tabColor = RND
write_maestra(ws)
ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
write_banda(ws, crit, f'Críticos · RND W{VOL_NUM}', banda_lookup)
ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
write_banda(ws, bajo, f'Bajo Rendimiento · RND W{VOL_NUM}', banda_lookup)
ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
write_banda(ws, sinc, f'Sin Conversión · RND W{VOL_NUM}', banda_lookup)

out = f'{OUTPUTS}/Analisis_RatesNoDispo_W{VOL_NUM}.xlsx'
wb.save(out)
print(f'✅ Excel RND: {out}')
print(f'   {len(wb.sheetnames)} hojas: {" | ".join(wb.sheetnames)}')
