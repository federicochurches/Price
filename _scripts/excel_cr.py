"""
excel_cr.py · W21+ · Excel único consolidado CheckRates
Genera: Analisis_CheckRates_W{NN}.xlsx
Hojas: Global | B2C | B2B-OP | CUG
Cada hoja = filtro del mismo DataFrame base (p80_hotel por DistributionCategory).
Reemplaza: excel_cr.py + excel_cr_canastas.py (4 archivos → 1)
"""
import pickle
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from engine import banda_eficacia, banda_convrate

# ── Config ────────────────────────────────────────────────────────────────────
VOL_NUM   = os.getenv('VOL_NUM', '21')
PERIODO   = os.getenv('PERIODO', '18–24 may 2026')
OUTPUTS   = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_CR', f'cr_w{VOL_NUM}_data.pkl'), 'rb') as _f:
    D = pickle.load(_f)

M            = D['M']
CANASTA      = D['CANASTA']
sev_ef_p80   = D['sev_ef_p80']
sev_cv_p80   = D['sev_cv_p80']
g_hotel      = D['g_hotel']
g_corp       = D['g_corp']
g_channel    = D['g_channel']
p80_hotel    = D['p80_hotel'].copy()
hotel_channel_map = D.get('hotel_channel_map', {})

def clean_hotel_name(name):
    if name and isinstance(name, str):
        return re.sub(r'^\(\d+\)\s*-\s*', '', name).strip()
    return name

p80_hotel['Hotel'] = p80_hotel['Hotel'].apply(clean_hotel_name)

_hcm_clean = {clean_hotel_name(k): v for k, v in hotel_channel_map.items()} if hotel_channel_map else {}
if _hcm_clean and 'Channel' not in p80_hotel.columns:
    p80_hotel['Channel'] = p80_hotel['Hotel'].map(_hcm_clean).fillna('—')

# ── Estilos ───────────────────────────────────────────────────────────────────
CR_COLOR    = '5C469C'
HEADER_FILL = PatternFill(start_color=CR_COLOR, end_color=CR_COLOR, fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT  = Font(name='Arial', size=14, bold=True, color=CR_COLOR)
SUBTITLE_FONT = Font(name='Arial', size=11, color=CR_COLOR)
META_FONT   = Font(name='Arial', size=10, color='666666')
DATA_FONT   = Font(name='Arial', size=10)
TAB_FONT    = Font(name='Arial', size=10, bold=True, color='FFFFFF')
THIN        = Side(border_style='thin', color='DDDDDD')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BAND_FILLS = {
    'Exitosa':        PatternFill(start_color='E1F5EE', end_color='E1F5EE', fill_type='solid'),
    'Aceptable':      PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
    'Revisar':        PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid'),
    'Crítica':        PatternFill(start_color='FCE4F1', end_color='FCE4F1', fill_type='solid'),
    'Súper Crítica':  PatternFill(start_color='161616', end_color='161616', fill_type='solid'),
    'Sin Conversión': PatternFill(start_color='F2EEE6', end_color='F2EEE6', fill_type='solid'),
}
BAND_FONTS = {
    'Súper Crítica':  Font(name='Arial', size=10, bold=True, color='FFFFFF'),
    'Sin Conversión': Font(name='Arial', size=10, color='8A8377'),
}

# Colores de pestaña por canasta
TAB_COLORS = {
    'Global': '5C469C',
    'B2C':    '4472C4',
    'B2B-OP': '2E75B6',
    'CUG':    '1F4E79',
}

DISPLAY_RENAME = {
    'CR_Unicos': 'Checkrates',
    'ConvRate':  'Conv Rate',
}

PP_CHANNELS = ['DerbySoft','Internal','HBSI','SynXis','Siteminder','Travelclick','Omnibees']

# ── Helpers ───────────────────────────────────────────────────────────────────
def add_header(ws, title, subtitle='', canasta_label=''):
    """Cabecera estándar: título + subtítulo + metadata."""
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    if canasta_label:
        ws['A2'] = canasta_label
        ws['A2'].font = SUBTITLE_FONT
    if subtitle:
        ws['A3'] = subtitle
        ws['A3'].font = META_FONT
    ws['A4'] = f'W{VOL_NUM} · {PERIODO} · Generado {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A4'].font = META_FONT

def add_table(ws, df, start_row=6, num_formats=None, banda_col=None, banda_col2=None):
    """Escribe DataFrame como tabla con header y bandas de color."""
    if num_formats is None:
        num_formats = {}
    df = df.rename(columns=DISPLAY_RENAME)
    if banda_col and banda_col in DISPLAY_RENAME:
        banda_col = DISPLAY_RENAME[banda_col]
    if banda_col2 and banda_col2 in DISPLAY_RENAME:
        banda_col2 = DISPLAY_RENAME[banda_col2]
    num_formats = {DISPLAY_RENAME.get(k, k): v for k, v in num_formats.items()}

    cols = list(df.columns)
    # Header row
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    # Data rows
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, c in enumerate(cols, 1):
            val = row[c]
            cell = ws.cell(row=start_row + i, column=j, value=(val if not pd.isna(val) else ''))
            cell.font = DATA_FONT
            cell.border = BORDER
            if c in num_formats:
                cell.number_format = num_formats[c]
            if banda_col and c == banda_col and val in BAND_FILLS:
                cell.fill = BAND_FILLS[val]
                if val in BAND_FONTS:
                    cell.font = BAND_FONTS[val]
            if banda_col2 and c == banda_col2 and val in BAND_FILLS:
                cell.fill = BAND_FILLS[val]
                if val in BAND_FONTS:
                    cell.font = BAND_FONTS[val]
    # Column widths
    for j, c in enumerate(cols, 1):
        vals = [str(c)] + [str(row[c]) for _, row in df.iterrows() if not pd.isna(row[c])]
        max_len = max((len(v) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 52)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

def set_tab_color(ws, canasta):
    ws.sheet_properties.tabColor = TAB_COLORS.get(canasta, CR_COLOR)

# ── Función principal por hoja ─────────────────────────────────────────────────
def build_sheet(wb, canasta_label, df_hotel, sev_ef, sev_cv, g_corp_df, g_channel_df, g_dest_df=None):
    """
    Construye todas las sub-secciones de una hoja de canasta.
    df_hotel: p80_hotel filtrado (o completo para Global).
    """
    is_global = (canasta_label == 'Global')
    tab_title = f'CheckRates · {canasta_label}'
    
    ws = wb.create_sheet(canasta_label)
    set_tab_color(ws, canasta_label)

    # ── Portada / índice de la hoja ──
    add_header(ws,
               title=f'Supply CheckRates · W{VOL_NUM} · {canasta_label}',
               subtitle=f'{len(df_hotel)} hoteles P80 · {PERIODO}',
               canasta_label='' if is_global else f'Canasta {canasta_label}')

    current_row = 6  # Reservar filas de portada, empezar secciones desde fila 6

    # Separador visual entre secciones dentro de la hoja
    def section_sep(ws, row, title):
        cell = ws.cell(row=row, column=1, value=f'▌ {title}')
        cell.font = Font(name='Arial', size=11, bold=True, color=CR_COLOR)
        ws.row_dimensions[row].height = 20
        return row + 1

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 1 · SEVERITY EFICACIA
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Severity · Eficacia  (target ≥ 97%)')
    total_ef = int(sev_ef.sum()) if hasattr(sev_ef, 'sum') else (sum(sev_ef.values()) or 1)
    RANGOS_EF = {'Súper Crítica':'<60%','Crítica':'60–85%','Revisar':'85–93%',
                 'Aceptable':'93–97%','Exitosa':'≥97%'}
    data_ef = [{'Banda': n, 'Rango': RANGOS_EF[n],
                'Hoteles': int(sev_ef.get(n, 0)),
                '%': int(sev_ef.get(n, 0)) / total_ef}
               for n in ['Súper Crítica','Crítica','Revisar','Aceptable','Exitosa']]
    add_table(ws, pd.DataFrame(data_ef), start_row=current_row,
              num_formats={'%': '0.0%'}, banda_col='Banda')
    current_row += len(data_ef) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 2 · SEVERITY CONV RATE
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Severity · Conv Rate  (target ≥ 2,5%)')
    total_cv = int(sev_cv.sum()) if hasattr(sev_cv, 'sum') else (sum(sev_cv.values()) or 1)
    RANGOS_CV = {'Sin Conversión':'BKGS=0','Crítica':'<0,8%','Revisar':'0,8–1,5%',
                 'Aceptable':'1,5–2,5%','Exitosa':'≥2,5%'}
    data_cv = [{'Banda': n, 'Rango': RANGOS_CV[n],
                'Hoteles': int(sev_cv.get(n, 0)),
                '%': int(sev_cv.get(n, 0)) / total_cv}
               for n in ['Sin Conversión','Crítica','Revisar','Aceptable','Exitosa']]
    add_table(ws, pd.DataFrame(data_cv), start_row=current_row,
              num_formats={'%': '0.0%'}, banda_col='Banda')
    current_row += len(data_cv) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 3 · TOP 100 CRÍTICOS (BKGS>0, peor Eficacia)
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Críticos  (BKGS>0 · Eficacia < 85%)')
    mask_c = (df_hotel['Bookings'] > 0) & (df_hotel['BandaEficacia'].isin(['Crítica','Súper Crítica']))
    df_c = (df_hotel[mask_c]
            .sort_values('Eficacia', ascending=True)
            .head(100)
            .reset_index(drop=True))
    df_c.insert(0, 'Rk', range(1, len(df_c)+1))
    cols_c = [col for col in ['Rk','Hotel','CorpName','Channel','Destino',
                               'CR_Unicos','Successful','Bookings','Eficacia','ConvRate','BandaEficacia']
              if col in df_c.columns]
    add_table(ws, df_c[cols_c], start_row=current_row,
              num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                           'CR_Unicos':'#,##0','Bookings':'#,##0','Successful':'#,##0'},
              banda_col='BandaEficacia')
    current_row += len(df_c) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 4 · TOP 100 BAJO RENDIMIENTO (ConvRate Crítica/Revisar)
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Bajo Rendimiento  (BKGS>0 · ConvRate Crítica/Revisar)')
    mask_b = (df_hotel['Bookings'] > 0) & (df_hotel['BandaConvRate'].isin(['Crítica','Revisar']))
    df_b = (df_hotel[mask_b]
            .sort_values('Eficacia', ascending=True)
            .head(100)
            .reset_index(drop=True))
    df_b.insert(0, 'Rk', range(1, len(df_b)+1))
    cols_b = [col for col in ['Rk','Hotel','CorpName','Channel','Destino',
                               'CR_Unicos','Bookings','Eficacia','ConvRate','BandaConvRate']
              if col in df_b.columns]
    add_table(ws, df_b[cols_b], start_row=current_row,
              num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                           'CR_Unicos':'#,##0','Bookings':'#,##0'},
              banda_col='BandaConvRate')
    current_row += len(df_b) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 5 · SIN CONVERSIÓN (BKGS=0)
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Sin Conversión  (BKGS=0 · cohorte estructural)')
    df_sc = (df_hotel[df_hotel['Bookings'] == 0]
             .sort_values('Eficacia', ascending=True)
             .head(100)
             .reset_index(drop=True))
    df_sc.insert(0, 'Rk', range(1, len(df_sc)+1))
    cols_sc = [col for col in ['Rk','Hotel','CorpName','Channel','Destino',
                                'CR_Unicos','Successful','Eficacia','BandaEficacia']
               if col in df_sc.columns]
    add_table(ws, df_sc[cols_sc], start_row=current_row,
              num_formats={'Eficacia':'0.00%','CR_Unicos':'#,##0','Successful':'#,##0'},
              banda_col='BandaEficacia')
    current_row += len(df_sc) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 6 · POR CORPORATIVO
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Por Corporativo  (Top 100 · ordenado por Eficacia ↑)')
    df_co = (g_corp_df
             .sort_values('Eficacia', ascending=True, na_position='last')
             .head(100)
             .reset_index(drop=True))
    df_co.insert(0, 'Rk', range(1, len(df_co)+1))
    if _hcm_clean:
        corp_ch = (df_hotel.groupby('CorpName')
                   .apply(lambda x: ', '.join(sorted(set(x['Channel'].fillna('—').tolist()))))
                   .reset_index().rename(columns={0: 'Channels'}))
        df_co = df_co.merge(corp_ch, on='CorpName', how='left')
    cols_co = [col for col in ['Rk','CorpName','Channels','CR_Unicos','Successful',
                                'Bookings','Eficacia','ConvRate','BandaEficacia','BandaConvRate']
               if col in df_co.columns]
    add_table(ws, df_co[cols_co], start_row=current_row,
              num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                           'CR_Unicos':'#,##0','Bookings':'#,##0','Successful':'#,##0'},
              banda_col='BandaEficacia')
    current_row += len(df_co) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 7 · POR DESTINO
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Por Destino  (Top 100 · ordenado por Eficacia ↑)')
    if g_dest_df is not None:
        df_de = (g_dest_df
                 .sort_values('Eficacia', ascending=True, na_position='last')
                 .head(100)
                 .reset_index(drop=True))
    else:
        # Calcular desde df_hotel filtrado
        df_de = (df_hotel.groupby('Destino')
                 .agg(CR_Unicos=('CR_Unicos','sum'), Successful=('Successful','sum'),
                      Bookings=('Bookings','sum'), Hoteles=('Hotel','nunique'))
                 .reset_index())
        df_de['Eficacia']   = df_de['Successful'] / df_de['CR_Unicos'].replace(0, 1)
        df_de['ConvRate']   = df_de['Bookings']   / df_de['CR_Unicos'].replace(0, 1)
        df_de['BandaEficacia'] = df_de['Eficacia'].apply(banda_eficacia)
        df_de['BandaConvRate'] = df_de.apply(lambda r: banda_convrate(r['ConvRate'], r['Bookings']), axis=1)
        df_de = df_de.sort_values('Eficacia', ascending=True, na_position='last').head(100).reset_index(drop=True)
    df_de.insert(0, 'Rk', range(1, len(df_de)+1))
    if _hcm_clean and 'Channel' in df_hotel.columns:
        dest_ch = (df_hotel.groupby('Destino')
                   .apply(lambda x: ', '.join(sorted(set(x['Channel'].fillna('—').tolist()))))
                   .reset_index().rename(columns={0: 'Channels'}))
        df_de = df_de.merge(dest_ch, on='Destino', how='left')
    cols_de = [col for col in ['Rk','Destino','Channels','Hoteles','CR_Unicos',
                                'Bookings','Eficacia','ConvRate','BandaEficacia','BandaConvRate']
               if col in df_de.columns]
    add_table(ws, df_de[cols_de], start_row=current_row,
              num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                           'CR_Unicos':'#,##0','Bookings':'#,##0'},
              banda_col='BandaEficacia')
    current_row += len(df_de) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 8 · POR CHANNEL (solo Global — en canastas no aplica)
    # ─────────────────────────────────────────────────────────────────
    if is_global:
        current_row = section_sep(ws, current_row, 'Por Channel  (Producto Propio vs Third Party)')
        df_ch = (g_channel_df
                 .sort_values('Eficacia', ascending=True, na_position='last')
                 .reset_index(drop=True))
        df_ch.insert(0, 'Rk', range(1, len(df_ch)+1))
        df_ch['Grupo'] = df_ch['ExternalProviderName'].apply(
            lambda x: 'Producto Propio' if x in PP_CHANNELS else 'Third Party')
        df_ch = df_ch.rename(columns={'ExternalProviderName': 'Channel'})
        cols_ch = [col for col in ['Rk','Channel','Grupo','CR_Unicos','Successful',
                                    'Bookings','Eficacia','ConvRate','BandaEficacia','BandaConvRate']
                   if col in df_ch.columns]
        add_table(ws, df_ch[cols_ch], start_row=current_row,
                  num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                               'CR_Unicos':'#,##0','Bookings':'#,##0','Successful':'#,##0'},
                  banda_col='BandaEficacia')
        current_row += len(df_ch) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 9 · MENOR CONV RATE (BKGS>0)
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Menor Conv Rate  (BKGS>0)')
    df_mc = (df_hotel[df_hotel['Bookings'] > 0]
             .sort_values('ConvRate', ascending=True)
             .head(100)
             .reset_index(drop=True))
    df_mc.insert(0, 'Rk', range(1, len(df_mc)+1))
    cols_mc = [col for col in ['Rk','Hotel','CorpName','Channel','Destino',
                                'CR_Unicos','Bookings','Eficacia','ConvRate','BandaConvRate']
               if col in df_mc.columns]
    add_table(ws, df_mc[cols_mc], start_row=current_row,
              num_formats={'Eficacia':'0.00%','ConvRate':'0.00%',
                           'CR_Unicos':'#,##0','Bookings':'#,##0'},
              banda_col='BandaConvRate')

    return ws


# ── Construir el workbook ─────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # quitar Sheet default

# Calcular g_dest global (para hoja Global)
g_dest_global = (g_hotel.groupby('Destino')
                 .agg(CR_Unicos=('CR_Unicos','sum'), Successful=('Successful','sum'),
                      Bookings=('Bookings','sum'), Hoteles=('Hotel','nunique'))
                 .reset_index())
g_dest_global['Eficacia']      = g_dest_global['Successful'] / g_dest_global['CR_Unicos'].replace(0,1)
g_dest_global['ConvRate']      = g_dest_global['Bookings']   / g_dest_global['CR_Unicos'].replace(0,1)
g_dest_global['BandaEficacia'] = g_dest_global['Eficacia'].apply(banda_eficacia)
g_dest_global['BandaConvRate'] = g_dest_global.apply(
    lambda r: banda_convrate(r['ConvRate'], r['Bookings']), axis=1)

# ── HOJA 1: GLOBAL ──
build_sheet(wb, 'Global',
            df_hotel=p80_hotel,
            sev_ef=sev_ef_p80,
            sev_cv=sev_cv_p80,
            g_corp_df=g_corp,
            g_channel_df=g_channel,
            g_dest_df=g_dest_global)
print('  ✓ Hoja Global')

# ── HOJAS 2-4: CANASTAS ──
CANASTA_MAP = {
    'B2C':    ('B2C', 'B2C'),
    'B2B-OP': ('B2B-OP', 'B2B (OP)'),
    'CUG':    ('CUG', 'CUG (UOP)'),
}

for sheet_name, (c_key_upper, dist_cat) in CANASTA_MAP.items():
    c = CANASTA.get(c_key_upper) or CANASTA.get(c_key_upper.lower())
    if c is None:
        print(f'  ⚠ Canasta {c_key_upper} no encontrada en pickle, omitiendo')
        continue

    # Filtrar p80_hotel por DistributionCategory
    if 'DistributionCategory' in p80_hotel.columns:
        df_canasta = p80_hotel[p80_hotel['DistributionCategory'] == dist_cat].copy()
    elif 'agg_hotel' in c:
        df_canasta = c['agg_hotel'].copy()
        df_canasta['Hotel'] = df_canasta['Hotel'].apply(clean_hotel_name)
        if _hcm_clean and 'Channel' not in df_canasta.columns:
            df_canasta['Channel'] = df_canasta['Hotel'].map(_hcm_clean).fillna('—')
    else:
        print(f'  ⚠ Sin datos para canasta {c_key_upper}, omitiendo')
        continue

    # sev_ef y sev_cv de la canasta
    sev_ef_c = c.get('sev_ef', {})
    sev_cv_c = c.get('sev_cv', {})

    # g_corp de la canasta
    if 'agg_corp' in c:
        g_corp_c = c['agg_corp']
    else:
        g_corp_c = (df_canasta.groupby('CorpName')
                    .agg(CR_Unicos=('CR_Unicos','sum'), Successful=('Successful','sum'),
                         Bookings=('Bookings','sum'))
                    .reset_index())
        g_corp_c['Eficacia']      = g_corp_c['Successful'] / g_corp_c['CR_Unicos'].replace(0,1)
        g_corp_c['ConvRate']      = g_corp_c['Bookings']   / g_corp_c['CR_Unicos'].replace(0,1)
        g_corp_c['BandaEficacia'] = g_corp_c['Eficacia'].apply(banda_eficacia)
        g_corp_c['BandaConvRate'] = g_corp_c.apply(
            lambda r: banda_convrate(r['ConvRate'], r['Bookings']), axis=1)

    build_sheet(wb, sheet_name,
                df_hotel=df_canasta,
                sev_ef=sev_ef_c,
                sev_cv=sev_cv_c,
                g_corp_df=g_corp_c,
                g_channel_df=g_channel,  # channel siempre global
                g_dest_df=None)          # se calcula desde df_canasta
    print(f'  ✓ Hoja {sheet_name}')

# ── Guardar ───────────────────────────────────────────────────────────────────
out = f'{OUTPUTS}/Analisis_CheckRates_W{VOL_NUM}.xlsx'
wb.save(out)
n_sheets = len(wb.sheetnames)
print(f'\n✅ Excel CR escrito: {out}')
print(f'   {n_sheets} hojas: {" | ".join(wb.sheetnames)}')
