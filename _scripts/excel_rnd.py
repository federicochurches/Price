"""
excel_rnd.py · W21+ · Excel único consolidado RatesNoDispo
Genera: Analisis_RatesNoDispo_W{NN}.xlsx
Hojas: Global | B2C | B2B-OP | CUG
Cada hoja = filtro del mismo DataFrame base por DistributionCategory.
Reemplaza: excel_rnd.py + excel_rnd_canastas.py (4 archivos → 1)
"""
import pickle
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from engine import banda_nodispo, banda_rpm

# ── Config ────────────────────────────────────────────────────────────────────
VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '18–24 may 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_RND', f'rnd_w{VOL_NUM}_data.pkl'), 'rb') as _f:
    D = pickle.load(_f)

M        = D['M']
CANASTA  = D['CANASTA']
df18     = D['df18']
sev_nd   = D['sev_nd']
sev_rpm  = D['sev_rpm']
g_hotel  = D['g_hotel']
p80_hotel = D['p80_hotel'].copy()

# Calcular bandas si no están en df18
if 'BandaNoDispo' not in df18.columns:
    df18['BandaNoDispo'] = df18['%NoDispo'].apply(banda_nodispo)
if 'BandaRPM' not in df18.columns:
    df18['BandaRPM'] = df18.apply(lambda r: banda_rpm(r['IPM'], r['Bookings']), axis=1)

# ── Estilos ───────────────────────────────────────────────────────────────────
RND_COLOR   = 'EA0074'
HEADER_FILL = PatternFill(start_color=RND_COLOR, end_color=RND_COLOR, fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT  = Font(name='Arial', size=14, bold=True, color=RND_COLOR)
SUBTITLE_FONT = Font(name='Arial', size=11, color=RND_COLOR)
META_FONT   = Font(name='Arial', size=10, color='666666')
DATA_FONT   = Font(name='Arial', size=10)
THIN        = Side(border_style='thin', color='DDDDDD')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BAND_FILLS = {
    'Exitosa':        PatternFill(start_color='E1F5EE', end_color='E1F5EE', fill_type='solid'),
    'Aceptable':      PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
    'Revisar':        PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid'),
    'Crítica':        PatternFill(start_color='FCE4F1', end_color='FCE4F1', fill_type='solid'),
    'Súper Crítica':  PatternFill(start_color='161616', end_color='161616', fill_type='solid'),
    'Sin Conversión': PatternFill(start_color='F2EEE6', end_color='F2EEE6', fill_type='solid'),
}
BAND_FONTS = {
    'Súper Crítica':  Font(name='Arial', size=10, bold=True, color='FFFFFF'),
    'Sin Conversión': Font(name='Arial', size=10, color='8A8377'),
}

TAB_COLORS = {
    'Global': 'EA0074',
    'B2C':    'C2185B',
    'B2B-OP': 'AD1457',
    'CUG':    '880E4F',
}

RANGOS_ND  = {'Súper Crítica':'>60%','Crítica':'20–60%','Revisar':'5–20%',
               'Aceptable':'3–5%','Exitosa':'<3%'}
RANGOS_IPM = {'Sin Conversión':'BKGS=0','Crítica':'<$200','Revisar':'$200–$650',
               'Aceptable':'$650–$1.500','Exitosa':'≥$1.500'}

# ── Helpers ───────────────────────────────────────────────────────────────────
def add_header(ws, title, subtitle='', canasta_label=''):
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    if canasta_label:
        ws['A2'] = canasta_label
        ws['A2'].font = SUBTITLE_FONT
    if subtitle:
        ws['A3'] = subtitle
        ws['A3'].font = META_FONT
    ws['A4'] = f'W{VOL_NUM} · {PERIODO} · Generado {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A4'].font = META_FONT

def add_table(ws, df, start_row=6, num_formats=None, banda_col=None, banda_col2=None):
    if num_formats is None:
        num_formats = {}
    cols = list(df.columns)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for j, c in enumerate(cols, 1):
            val = row[c]
            cell = ws.cell(row=start_row + i, column=j, value=(val if not pd.isna(val) else ''))
            cell.font = DATA_FONT
            cell.border = BORDER
            if c in num_formats:
                cell.number_format = num_formats[c]
            band_val = None
            if banda_col and c == banda_col and val in BAND_FILLS:
                band_val = val
            elif banda_col2 and c == banda_col2 and val in BAND_FILLS:
                band_val = val
            if band_val:
                cell.fill = BAND_FILLS[band_val]
                if band_val in BAND_FONTS:
                    cell.font = BAND_FONTS[band_val]
    for j, c in enumerate(cols, 1):
        vals = [str(c)] + [str(row[c]) for _, row in df.iterrows() if not pd.isna(row[c])]
        max_len = max((len(v) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 52)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

def set_tab_color(ws, canasta):
    ws.sheet_properties.tabColor = TAB_COLORS.get(canasta, RND_COLOR)

def agg_dim(df_base, group_col):
    """Agrega un DataFrame RND por columna de dimensión."""
    g = (df_base.groupby(group_col)
         .agg(Trafico=('Trafico','sum'), Bookings=('Bookings','sum'),
              gb_usd=('gb_usd','sum'), TraficoNoDispo=('TraficoNoDispo','sum'),
              Hoteles=('Hotel','nunique'))
         .reset_index())
    g['%NoDispo']    = g['TraficoNoDispo'] / g['Trafico'].replace(0, 1)
    g['IPM']         = g['gb_usd'] / g['Trafico'].replace(0, 1) * 1_000_000
    g['BandaNoDispo']= g['%NoDispo'].apply(banda_nodispo)
    g['BandaIPM']    = g.apply(lambda r: banda_rpm(r['IPM'], r['Bookings']), axis=1)
    return g

# ── Función principal por hoja ─────────────────────────────────────────────────
def build_sheet(wb, canasta_label, df_base):
    """
    Construye todas las secciones de una hoja.
    df_base: df18 filtrado por DistributionCategory (o completo para Global).
    """
    is_global = (canasta_label == 'Global')
    ws = wb.create_sheet(canasta_label)
    set_tab_color(ws, canasta_label)

    add_header(ws,
               title=f'Supply Rates No Dispo · W{VOL_NUM} · {canasta_label}',
               subtitle=f'{len(df_base)} registros hotel×canasta · {PERIODO}',
               canasta_label='' if is_global else f'Canasta {canasta_label}')

    current_row = 6

    def section_sep(ws, row, title):
        cell = ws.cell(row=row, column=1, value=f'▌ {title}')
        cell.font = Font(name='Arial', size=11, bold=True, color=RND_COLOR)
        ws.row_dimensions[row].height = 20
        return row + 1

    # ── Calcular severities desde df_base ──
    sev_nd_c  = df_base.groupby('BandaNoDispo').size().to_dict() if not is_global else sev_nd
    sev_ipm_c = df_base.groupby('BandaRPM').size().to_dict()    if not is_global and 'BandaRPM' in df_base.columns else sev_rpm

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 1 · SEVERITY %NoDispo
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Severity · %NoDispo  (target < 3%)')
    total_nd = sum(sev_nd_c.values()) or 1
    data_nd = [{'Banda': n, 'Rango': RANGOS_ND[n],
                'Hoteles': int(sev_nd_c.get(n, 0)),
                '%': int(sev_nd_c.get(n, 0)) / total_nd}
               for n in ['Súper Crítica','Crítica','Revisar','Aceptable','Exitosa']]
    add_table(ws, pd.DataFrame(data_nd), start_row=current_row,
              num_formats={'%': '0.0%'}, banda_col='Banda')
    current_row += len(data_nd) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 2 · SEVERITY IPM
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Severity · IPM (Income Per Million USD)  (target ≥ $650)')
    total_ipm = sum(sev_ipm_c.values()) or 1
    data_ipm = [{'Banda': n, 'Rango': RANGOS_IPM[n],
                 'Hoteles': int(sev_ipm_c.get(n, 0)),
                 '%': int(sev_ipm_c.get(n, 0)) / total_ipm}
                for n in ['Sin Conversión','Crítica','Revisar','Aceptable','Exitosa']]
    add_table(ws, pd.DataFrame(data_ipm), start_row=current_row,
              num_formats={'%': '0.0%'}, banda_col='Banda')
    current_row += len(data_ipm) + 3

    # Agregar IPM y bandas si no están
    df_w = df_base.copy()
    if 'IPM' not in df_w.columns:
        df_w['IPM'] = df_w['gb_usd'] / df_w['Trafico'].replace(0, 1) * 1_000_000
    if 'BandaRPM' not in df_w.columns:
        df_w['BandaRPM'] = df_w.apply(lambda r: banda_rpm(r['IPM'], r['Bookings']), axis=1)
    if 'DemandaNoConvertida' not in df_w.columns:
        df_w['DemandaNoConvertida'] = (df_w['Trafico'] * df_w['%NoDispo']).round(0).astype(int)

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 3 · DEMANDA NO CONVERTIDA Top 100
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Demanda No Convertida  (%NoDispo > 0 · mayor demanda perdida)')
    df_dnc = (df_w[df_w['TraficoNoDispo'] > 0]
              .sort_values('%NoDispo', ascending=False)
              .head(100).reset_index(drop=True))
    df_dnc.insert(0, 'Rk', range(1, len(df_dnc)+1))
    cols_dnc = [col for col in ['Rk','Hotel','CorpName','PaisDestino','Destino',
                                  'Trafico','%NoDispo','Bookings','gb_usd','IPM',
                                  'BandaNoDispo','BandaRPM','DemandaNoConvertida']
                if col in df_dnc.columns]
    df_dnc_out = df_dnc[cols_dnc].rename(columns={'IPM':'IPM (USD/M)','BandaRPM':'Banda IPM'})
    add_table(ws, df_dnc_out, start_row=current_row,
              num_formats={'%NoDispo':'0.00%','gb_usd':'$#,##0','IPM (USD/M)':'$#,##0',
                           'Trafico':'#,##0','DemandaNoConvertida':'#,##0'},
              banda_col='BandaNoDispo', banda_col2='Banda IPM')
    current_row += len(df_dnc) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 4 · BAJO RENDIMIENTO Top 100
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Bajo Rendimiento  (BKGS>0 · IPM Crítica/Revisar)')
    mask_br = (df_w['Bookings'] > 0) & (df_w['BandaRPM'].isin(['Crítica','Revisar']))
    df_br = (df_w[mask_br]
             .sort_values('%NoDispo', ascending=False)
             .head(100).reset_index(drop=True))
    df_br.insert(0, 'Rk', range(1, len(df_br)+1))
    cols_br = [col for col in ['Rk','Hotel','CorpName','PaisDestino','Destino',
                                 'Trafico','%NoDispo','Bookings','gb_usd','IPM','BandaNoDispo','BandaRPM']
               if col in df_br.columns]
    df_br_out = df_br[cols_br].rename(columns={'IPM':'IPM (USD/M)','BandaRPM':'Banda IPM'})
    add_table(ws, df_br_out, start_row=current_row,
              num_formats={'%NoDispo':'0.00%','gb_usd':'$#,##0','IPM (USD/M)':'$#,##0','Trafico':'#,##0'},
              banda_col='BandaNoDispo', banda_col2='Banda IPM')
    current_row += len(df_br) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 5 · SIN CONVERSIÓN Top 100
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Top 100 · Sin Conversión  (BKGS=0 · cohorte estructural)')
    df_sc = (df_w[df_w['Bookings'] == 0]
             .sort_values('Trafico', ascending=False)
             .head(100).reset_index(drop=True))
    df_sc.insert(0, 'Rk', range(1, len(df_sc)+1))
    cols_sc = [col for col in ['Rk','Hotel','CorpName','PaisDestino','Destino',
                                 'Trafico','%NoDispo','Bookings','BandaNoDispo']
               if col in df_sc.columns]
    add_table(ws, df_sc[cols_sc], start_row=current_row,
              num_formats={'%NoDispo':'0.00%','Trafico':'#,##0'},
              banda_col='BandaNoDispo')
    current_row += len(df_sc) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 6 · POR CORPORATIVO Top 100
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Por Corporativo  (Top 100 · ordenado por tráfico ↓)')
    g_co = agg_dim(df_w, 'CorpName').sort_values('Trafico', ascending=False).head(100).reset_index(drop=True)
    g_co.insert(0, 'Rk', range(1, len(g_co)+1))
    cols_co = [col for col in ['Rk','CorpName','Hoteles','Trafico','Bookings',
                                 'gb_usd','%NoDispo','IPM','BandaNoDispo','BandaIPM']
               if col in g_co.columns]
    df_co_out = g_co[cols_co].rename(columns={'IPM':'IPM (USD/M)','BandaIPM':'Banda IPM'})
    add_table(ws, df_co_out, start_row=current_row,
              num_formats={'%NoDispo':'0.00%','gb_usd':'$#,##0','IPM (USD/M)':'$#,##0','Trafico':'#,##0'},
              banda_col='BandaNoDispo', banda_col2='Banda IPM')
    current_row += len(g_co) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 7 · POR DESTINO Top 100
    # ─────────────────────────────────────────────────────────────────
    current_row = section_sep(ws, current_row, 'Por Destino  (Top 100 · ordenado por tráfico ↓)')
    g_de = agg_dim(df_w, 'Destino').sort_values('Trafico', ascending=False).head(100).reset_index(drop=True)
    g_de.insert(0, 'Rk', range(1, len(g_de)+1))
    cols_de = [col for col in ['Rk','Destino','Hoteles','Trafico','Bookings',
                                 'gb_usd','%NoDispo','IPM','BandaNoDispo','BandaIPM']
               if col in g_de.columns]
    df_de_out = g_de[cols_de].rename(columns={'IPM':'IPM (USD/M)','BandaIPM':'Banda IPM'})
    add_table(ws, df_de_out, start_row=current_row,
              num_formats={'%NoDispo':'0.00%','gb_usd':'$#,##0','IPM (USD/M)':'$#,##0','Trafico':'#,##0'},
              banda_col='BandaNoDispo', banda_col2='Banda IPM')
    current_row += len(g_de) + 3

    # ─────────────────────────────────────────────────────────────────
    # SECCIÓN 8 · POR PAÍS Top 100 (solo Global)
    # ─────────────────────────────────────────────────────────────────
    if is_global:
        current_row = section_sep(ws, current_row, 'Por País  (todos los países · ordenado por tráfico ↓)')
        g_pa = agg_dim(df_w, 'PaisDestino').sort_values('Trafico', ascending=False).reset_index(drop=True)
        g_pa.insert(0, 'Rk', range(1, len(g_pa)+1))
        cols_pa = [col for col in ['Rk','PaisDestino','Hoteles','Trafico','Bookings',
                                     'gb_usd','%NoDispo','IPM','BandaNoDispo','BandaIPM']
                   if col in g_pa.columns]
        df_pa_out = g_pa[cols_pa].rename(columns={'IPM':'IPM (USD/M)','BandaIPM':'Banda IPM'})
        add_table(ws, df_pa_out, start_row=current_row,
                  num_formats={'%NoDispo':'0.00%','gb_usd':'$#,##0','IPM (USD/M)':'$#,##0','Trafico':'#,##0'},
                  banda_col='BandaNoDispo', banda_col2='Banda IPM')

    return ws


# ── Construir el workbook ─────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# HOJA 1: GLOBAL — usa df18 completo
build_sheet(wb, 'Global', df18)
print('  ✓ Hoja Global')

# HOJAS 2-4: CANASTAS — filtro por DistributionCategory
CANASTA_FILTER = {
    'B2C':    'B2C',
    'B2B-OP': 'B2B (OP)',
    'CUG':    'CUG (UOP)',
}

for sheet_name, dist_cat in CANASTA_FILTER.items():
    if 'DistributionCategory' in df18.columns:
        df_c = df18[df18['DistributionCategory'] == dist_cat].copy()
    else:
        # Fallback: usar agg_hotel del CANASTA dict
        c_key = sheet_name.lower().replace('-', '')  # b2c, op, cug
        c = CANASTA.get(c_key) or CANASTA.get(sheet_name)
        if c and 'agg_hotel' in c:
            df_c = c['agg_hotel'].copy()
            # Asegurar columnas mínimas
            if 'BandaNoDispo' not in df_c.columns:
                df_c['BandaNoDispo'] = df_c['%NoDispo'].apply(banda_nodispo)
            if 'BandaRPM' not in df_c.columns and 'RPM' in df_c.columns:
                df_c['BandaRPM'] = df_c.apply(lambda r: banda_rpm(r['RPM'], r['Bookings']), axis=1)
            if 'IPM' not in df_c.columns and 'RPM' in df_c.columns:
                df_c = df_c.rename(columns={'RPM': 'IPM'})
            if 'TraficoNoDispo' not in df_c.columns:
                df_c['TraficoNoDispo'] = df_c['Trafico'] * df_c['%NoDispo']
        else:
            print(f'  ⚠ Sin datos para canasta {sheet_name}, omitiendo')
            continue

    if len(df_c) == 0:
        print(f'  ⚠ Canasta {sheet_name} vacía tras filtro, omitiendo')
        continue

    build_sheet(wb, sheet_name, df_c)
    print(f'  ✓ Hoja {sheet_name}  ({len(df_c)} registros)')

# ── Guardar ───────────────────────────────────────────────────────────────────
out = f'{OUTPUTS}/Analisis_RatesNoDispo_W{VOL_NUM}.xlsx'
wb.save(out)
n_sheets = len(wb.sheetnames)
print(f'\n✅ Excel RND escrito: {out}')
print(f'   {n_sheets} hojas: {" | ".join(wb.sheetnames)}')
