"""
Excel Análisis CheckRates por Canasta · 3 Excels (B2C, OP, CUG)
Genera un Excel de 9 pestañas para cada canasta
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pickle
import os, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

with open(os.getenv('PICKLE_CR', 'cr_w20_data.pkl'),'rb') as _f:
    D = pickle.load(_f)

CANASTA = D['CANASTA']
VOL_NUM = D.get('VOL_NUM', '20')
PERIODO = D.get('PERIODO', '12–18 may 2026')
MES_AÑO = D.get('MES_AÑO', 'Mayo 2026')

# Estilos (mismo que excel_cr.py)
HEADER_FILL = PatternFill(start_color='5C469C', end_color='5C469C', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT  = Font(name='Arial', size=14, bold=True, color='5C469C')
META_FONT   = Font(name='Arial', size=10, color='666666')
DATA_FONT   = Font(name='Arial', size=10)
THIN = Side(border_style='thin', color='DDDDDD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILLS = {
    'Exitosa':       PatternFill(start_color='E8F7FD', end_color='E8F7FD', fill_type='solid'),
    'Aceptable':     PatternFill(start_color='EDE8F7', end_color='EDE8F7', fill_type='solid'),
    'Revisar':       PatternFill(start_color='FFF4E0', end_color='FFF4E0', fill_type='solid'),
    'Crítica':       PatternFill(start_color='FCE4F1', end_color='FCE4F1', fill_type='solid'),
    'Súper Crítica': PatternFill(start_color='161616', end_color='161616', fill_type='solid'),
    'Sin Conversión':PatternFill(start_color='F2EEE6', end_color='F2EEE6', fill_type='solid'),
}
BAND_FONTS = {
    'Súper Crítica': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
    'Sin Conversión': Font(name='Arial', size=10, color='8A8377'),
}

def add_title(ws, title, subtitle=''):
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    if subtitle:
        ws['A2'] = subtitle
        ws['A2'].font = META_FONT
    ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} · W{VOL_NUM} · {PERIODO}'
    ws['A3'].font = META_FONT

def add_table(ws, df, start_row=5, num_formats=None, banda_col=None):
    """Escribe DF con headers y formatos."""
    if num_formats is None: num_formats = {}
    if df.empty:
        ws[f'A{start_row}'] = "Sin datos"
        return
    
    cols = list(df.columns)
    # Header
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    # Data
    for i, (_, r) in enumerate(df.iterrows(), 1):
        for j, c in enumerate(cols, 1):
            val = r[c]
            cell = ws.cell(row=start_row+i, column=j, value=val if not pd.isna(val) else '')
            cell.font = DATA_FONT
            cell.border = BORDER
            if c in num_formats:
                cell.number_format = num_formats[c]
            # Aplicar color fill a columnas de banda
            if banda_col and c == banda_col and val in BAND_FILLS:
                cell.fill = BAND_FILLS[val]
                if val in BAND_FONTS:
                    cell.font = BAND_FONTS[val]
    # Auto width
    for j, c in enumerate(cols, 1):
        max_len = max([len(str(c))] + [len(str(r[c])) if not pd.isna(r[c]) else 0 for _, r in df.iterrows()])
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 50)
    # Freeze
    ws.freeze_panes = ws.cell(row=start_row+1, column=1)

# Mapeo de canastas para nombres de archivo
CANASTA_NAMES = {
    'B2C': 'B2C',
    'B2B-OP': 'OP',
    'CUG': 'CUG',
}

# Generar un Excel por cada canasta (B2C, OP, CUG)
for canasta_key, canasta_data in CANASTA.items():
    # Solo procesar las canastas principales
    if canasta_key in CANASTA_NAMES:
        canasta_name = CANASTA_NAMES[canasta_key]
        
        print(f"\nGenerando Excel {canasta_name}...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Extraer datos de la canasta
        agg_hotel = canasta_data.get('agg_hotel', pd.DataFrame())
        sev_ef = canasta_data.get('sev_ef', pd.Series())
        sev_cv = canasta_data.get('sev_cv', pd.Series())
        critic = canasta_data.get('critic', pd.DataFrame()).head(100)
        bajo = canasta_data.get('bajo', pd.DataFrame()).head(100)
        sin_conv = canasta_data.get('sin_conv', pd.DataFrame()).head(100)
        menor_cv = canasta_data.get('menor_cv', pd.DataFrame()).head(100)
        
        # Pestaña 1: Severity Eficacia
        ws = wb.create_sheet('Severity Eficacia')
        add_title(ws, f'Severity · Eficacia · {canasta_name}')
        data = []
        if isinstance(sev_ef, pd.Series):
            for n in ['Súper Crítica','Crítica','Revisar','Aceptable','Exitosa']:
                rng = {'Súper Crítica':'<60%','Crítica':'60-85%','Revisar':'85-93%','Aceptable':'93-97%','Exitosa':'≥97%'}[n]
                cnt = int(sev_ef.get(n, 0))
                data.append({'Banda':n,'Rango':rng,'Hoteles':cnt})
        if data:
            df_sev = pd.DataFrame(data)
            add_table(ws, df_sev, start_row=5, banda_col='Banda')
        
        # Pestaña 2: Severity ConvRate
        ws = wb.create_sheet('Severity ConvRate')
        add_title(ws, f'Severity · ConvRate · {canasta_name}')
        data = []
        if isinstance(sev_cv, pd.Series):
            for n in ['Sin Conversión','Crítica','Revisar','Aceptable','Exitosa']:
                rng = {'Sin Conversión':'0%','Crítica':'<0.8%','Revisar':'0.8-1.5%','Aceptable':'1.5-2.5%','Exitosa':'≥2.5%'}[n]
                cnt = int(sev_cv.get(n, 0))
                data.append({'Banda':n,'Rango':rng,'Hoteles':cnt})
        if data:
            df_sev = pd.DataFrame(data)
            add_table(ws, df_sev, start_row=5, banda_col='Banda')
        
        # Pestaña 3: Top 100 Críticos
        ws = wb.create_sheet('Críticos')
        add_title(ws, f'Top Críticos · {canasta_name}')
        add_table(ws, critic, start_row=5, banda_col='BandaEficacia')
        
        # Pestaña 4: Top 100 Bajo Rendimiento
        ws = wb.create_sheet('Bajo Rendimiento')
        add_title(ws, f'Top Bajo Rendimiento · {canasta_name}')
        add_table(ws, bajo, start_row=5, banda_col='BandaEficacia')
        
        # Pestaña 5: Sin Conversión
        ws = wb.create_sheet('Sin Conversión')
        add_title(ws, f'Sin Conversión · {canasta_name}')
        add_table(ws, sin_conv, start_row=5, banda_col='BandaConvRate')
        
        # Pestaña 6: Menor ConvRate
        ws = wb.create_sheet('Menor ConvRate')
        add_title(ws, f'Menor ConvRate · {canasta_name}')
        add_table(ws, menor_cv, start_row=5, banda_col='BandaConvRate')
        
        # Pestañas por dimensión (si existen)
        if 'g_corp' in canasta_data:
            ws = wb.create_sheet('Por Corporativo')
            add_title(ws, f'Top Corporativos · {canasta_name}')
            add_table(ws, canasta_data['g_corp'], start_row=5)
        
        if 'g_dest' in canasta_data:
            ws = wb.create_sheet('Por Destino')
            add_title(ws, f'Top Destinos · {canasta_name}')
            add_table(ws, canasta_data['g_dest'], start_row=5)
        
        if 'g_chan' in canasta_data:
            ws = wb.create_sheet('Por Channel')
            add_title(ws, f'Top Channels · {canasta_name}')
            add_table(ws, canasta_data['g_chan'], start_row=5)
        
        # Guardar
        filename = f'Analisis_Checkrates_{canasta_name}_7d.xlsx'
        filepath = f'/mnt/user-data/outputs/{filename}'
        wb.save(filepath)
        print(f"✅ {filename} generado ({len(wb.sheetnames)} pestañas)")

print("\n✅ Excels por canasta CR generados")
