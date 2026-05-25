"""
Excel Análisis RND por Canasta · 3 Excels (B2C, OP, CUG)
Genera un Excel de 8 pestañas para cada canasta
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pickle
import os, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

with open(os.getenv('PICKLE_RND', 'rnd_w20_data.pkl'),'rb') as _f:
    D = pickle.load(_f)

CANASTA = D['CANASTA']
VOL_NUM = D.get('VOL_NUM', '20')
PERIODO = D.get('PERIODO', '12–18 may 2026')
MES_AÑO = D.get('MES_AÑO', 'Mayo 2026')

# Estilos (mismo que excel_rnd.py)
HEADER_FILL = PatternFill(start_color='EA0074', end_color='EA0074', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT  = Font(name='Arial', size=14, bold=True, color='EA0074')
META_FONT   = Font(name='Arial', size=10, color='666666')
DATA_FONT   = Font(name='Arial', size=10)
THIN = Side(border_style='thin', color='DDDDDD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILLS = {
    'Exitosa':       PatternFill(start_color='E8F7FD', end_color='E8F7FD', fill_type='solid'),
    'Aceptable':     PatternFill(start_color='EDE8F7', end_color='EDE8F7', fill_type='solid'),
    'Revisar':       PatternFill(start_color='FFF4E0', end_color='FFF4E0', fill_type='solid'),
    'Crítica':       PatternFill(start_color='FCE4F1', end_color='FCE4F1', fill_type='solid'),
    'Súper Crítica': PatternFill(start_color='161616', end_color='161616', fill_type='solid'),
    'Sin Conversión':PatternFill(start_color='F2EEE6', end_color='F2EEE6', fill_type='solid'),
}
BAND_FONTS = {
    'Súper Crítica': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
    'Sin Conversión': Font(name='Arial', size=10, color='8A8377'),
}

def add_title(ws, title, subtitle=''):
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    if subtitle:
        ws['A2'] = subtitle
        ws['A2'].font = META_FONT
    ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} · W{VOL_NUM} · {PERIODO}'
    ws['A3'].font = META_FONT

def add_table(ws, df, start_row=5, num_formats=None, banda_col=None, banda_col2=None):
    """Escribe DF con headers y formatos."""
    if num_formats is None: num_formats = {}
    if df.empty:
        ws[f'A{start_row}'] = "Sin datos"
        return
    
    cols = list(df.columns)
    # Header
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    # Data
    for i, (_, r) in enumerate(df.iterrows(), 1):
        for j, c in enumerate(cols, 1):
            val = r[c]
            cell = ws.cell(row=start_row+i, column=j, value=val if not pd.isna(val) else '')
            cell.font = DATA_FONT
            cell.border = BORDER
            if c in num_formats:
                cell.number_format = num_formats[c]
            # Aplicar color fill a columnas de banda
            band_val = None
            if banda_col and c == banda_col and val in BAND_FILLS:
                band_val = val
            elif banda_col2 and c == banda_col2 and val in BAND_FILLS:
                band_val = val
            if band_val:
                cell.fill = BAND_FILLS[band_val]
                if band_val in BAND_FONTS:
                    cell.font = BAND_FONTS[band_val]
    # Auto width
    for j, c in enumerate(cols, 1):
        max_len = max([len(str(c))] + [len(str(r[c])) if not pd.isna(r[c]) else 0 for _, r in df.iterrows()])
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 50)
    # Freeze
    ws.freeze_panes = ws.cell(row=start_row+1, column=1)

# Mapeo de canastas para nombres de archivo
CANASTA_NAMES = {
    'B2C': 'B2C',
    'b2c': 'B2C',
    'B2B-OP': 'OP',
    'b2b-op': 'OP',
    'OP': 'OP',
    'op': 'OP',
    'CUG': 'CUG',
    'cug': 'CUG',
}

# Generar un Excel por cada canasta (B2C, OP, CUG)
for canasta_key, canasta_data in CANASTA.items():
    # Solo procesar las canastas principales (ignore duplicados en minúscula)
    if canasta_key.lower() in ['b2c', 'b2b-op', 'cug']:
        canasta_name = CANASTA_NAMES.get(canasta_key, canasta_key)
        
        print(f"\nGenerando Excel {canasta_name}...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Extraer datos de la canasta
        p80_hotel = canasta_data.get('p80_hotel', pd.DataFrame())
        sev_nd = canasta_data.get('sev_nd', {})
        sev_rpm = canasta_data.get('sev_rpm', {})
        bajo_rend = canasta_data.get('bajo_rend', pd.DataFrame()).head(100)
        bajo_rend_extra = canasta_data.get('bajo_rend_extra', pd.DataFrame())
        sin_conv = canasta_data.get('sin_conv', pd.DataFrame()).head(100)
        sin_conv_extra = canasta_data.get('sin_conv_extra', pd.DataFrame())
        
        # Pestaña 1: Severity %NoDispo
        ws = wb.create_sheet('Severity %NoDispo')
        add_title(ws, f'Severity · %NoDispo · {canasta_name}', f'P80 · {len(p80_hotel)} hoteles')
        total = int(sum(sev_nd.values())) if isinstance(sev_nd, dict) else 1
        data = []
        if isinstance(sev_nd, dict):
            for n in ['Súper Crítica','Crítica','Revisar','Aceptable','Exitosa']:
                rng = {'Súper Crítica':'>60%','Crítica':'20-60%','Revisar':'5-20%','Aceptable':'3-5%','Exitosa':'<3%'}[n]
                cnt = int(sev_nd.get(n, 0))
                data.append({'Banda':n,'Rango':rng,'Hoteles':cnt,'%':cnt/total if total else 0})
        df_sev = pd.DataFrame(data)
        add_table(ws, df_sev, start_row=5, num_formats={'%':'0.0%'}, banda_col='Banda')
        
        # Pestaña 2: Severity IPM
        ws = wb.create_sheet('Severity IPM')
        add_title(ws, f'Severity · IPM (USD/M) · {canasta_name}')
        total = int(sum(sev_rpm.values())) if isinstance(sev_rpm, dict) else 1
        data = []
        if isinstance(sev_rpm, dict):
            for n in ['Exitosa','Aceptable','Revisar','Crítica','Súper Crítica']:
                rng = {
                    'Exitosa':       '≥$650',
                    'Aceptable':     '$500–$649',
                    'Revisar':       '$200–$499',
                    'Crítica':       '<$199',
                    'Súper Crítica': '$0',
                }[n]
                cnt = int(sev_rpm.get(n, 0))
                data.append({'Banda':n,'Rango':rng,'Hoteles':cnt,'%':cnt/total if total else 0})
        df_sev = pd.DataFrame(data)
        add_table(ws, df_sev, start_row=5, num_formats={'%':'0.0%'}, banda_col='Banda')
        
        # Pestaña 3: Top 100 Bajo Rendimiento
        ws = wb.create_sheet('Bajo Rendimiento')
        add_title(ws, f'Top 100 Bajo Rendimiento · {canasta_name}')
        bajo_rend_all = pd.concat([bajo_rend, bajo_rend_extra], ignore_index=True).drop_duplicates().head(100)
        add_table(ws, bajo_rend_all, start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0','DemandaNoConvertida':'#,##0'}, banda_col='BandaNoDispo')
        
        # Pestaña 4: Top 100 Sin Conversión
        ws = wb.create_sheet('Sin Conversión')
        add_title(ws, f'Top 100 Sin Conversión · {canasta_name}')
        sin_conv_all = pd.concat([sin_conv, sin_conv_extra], ignore_index=True).drop_duplicates().head(100)
        add_table(ws, sin_conv_all, start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0','DemandaNoConvertida':'#,##0'}, banda_col='BandaNoDispo')
        
        # Pestañas por dimensión (si existen)
        if 'corps_10' in canasta_data:
            ws = wb.create_sheet('Por Corporativo')
            add_title(ws, f'Top Corporativos · {canasta_name}')
            add_table(ws, canasta_data['corps_10'], start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0'})
        
        if 'destinos_10' in canasta_data:
            ws = wb.create_sheet('Por Destino')
            add_title(ws, f'Top Destinos · {canasta_name}')
            add_table(ws, canasta_data['destinos_10'], start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0'})
        
        if 'paises_10' in canasta_data:
            ws = wb.create_sheet('Por País')
            add_title(ws, f'Top Países · {canasta_name}')
            add_table(ws, canasta_data['paises_10'], start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0'})
        
        # Pestaña: P80 completo
        ws = wb.create_sheet('P80 Completo')
        add_title(ws, f'P80 Hoteles · {canasta_name}')
        add_table(ws, p80_hotel, start_row=5, num_formats={'%NoDispo':'0.00%','IPM':'$#,##0','RPM':'$#,##0','Trafico':'#,##0','Bookings':'#,##0','gb_usd':'$#,##0','DemandaNoConvertida':'#,##0'}, banda_col='BandaNoDispo')
        
        # Guardar
        filename = f'Analisis_Rates_NoDispo_{canasta_name}_7d.xlsx'
        filepath = f'/mnt/user-data/outputs/{filename}'
        wb.save(filepath)
        print(f"✅ {filename} generado ({len(wb.sheetnames)} pestañas)")

print("\n✅ Excels por canasta RND generados")
