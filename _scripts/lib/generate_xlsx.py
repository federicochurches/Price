"""
Generación de Excels Top 50 · CR + RND
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

# Estilos compartidos
CELL_FONT = Font(name='Arial', size=10)
HEADER_FONT_WHITE = Font(name='Arial', bold=True, size=10, color='FFFFFF')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)


def write_header(ws, headers, row, color):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_row(ws, row_idx, values, alignments, fmts=None):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = alignments[col_idx - 1]
        if fmts and col_idx - 1 < len(fmts) and fmts[col_idx - 1]:
            cell.number_format = fmts[col_idx - 1]


def set_widths(ws, widths):
    for col_letter, w in zip('ABCDEFGHIJKL', widths):
        ws.column_dimensions[col_letter].width = w


def generate_cr_xlsx(data: dict, week: int, periodo: str, output_path: Path):
    """Genera Analisis_Checkrates_7d_WNN.xlsx con 11 pestañas Top 50."""
    ACCENT = '5C469C'
    HEADER_FILL = PatternFill('solid', fgColor=ACCENT)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    kpis = data['kpis']
    sev_e = data['severity_eficacia']
    sev_c = data['severity_cr']
    
    # 1 · Ficha Técnica
    ws = wb.create_sheet('Ficha Técnica')
    ws['A1'] = f'CHECKRATES · WEEK {week} · {periodo}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=ACCENT)
    ws.merge_cells('A1:C1')
    ws['A3'] = 'Ficha Técnica del Análisis'
    ws['A3'].font = Font(name='Arial', bold=True, size=12)
    ws.merge_cells('A3:C3')
    
    ficha = [
        ('SECCIÓN', 'INDICADOR', 'VALOR'),
        ('Universo', 'Total hoteles dataset', kpis['total_hot']),
        ('Universo', 'Hoteles HTS P80', kpis['p80_count']),
        ('Universo', 'CheckRates totales', f'{kpis["total_ck"]:,.0f}'),
        ('Universo', 'Bookings totales', kpis['total_bkgs']),
        ('Universo', 'Hoteles con 0 BKGS', kpis['zero_bkgs']),
        ('Universo', '% hoteles con 0 BKGS', f'{kpis["zero_bkgs_pct"]:.1f}%'),
        ('Métricas globales', 'Eficacia ponderada', f'{kpis["eficacia"]:.2f}%'),
        ('Métricas globales', 'Conv Rate ponderada', f'{kpis["cr"]:.2f}%'),
        ('Severity Eficacia', 'Exitosa (>97%)', sev_e['exitosa']),
        ('Severity Eficacia', 'Aceptable (93-97%)', sev_e['aceptable']),
        ('Severity Eficacia', 'Revisar (85-93%)', sev_e['revisar']),
        ('Severity Eficacia', 'Crítica (60-85%)', sev_e['critica']),
        ('Severity Eficacia', 'Súper Crítica (<60%)', sev_e['super']),
        ('Severity CR', 'Exitosa (>3%)', sev_c['exitosa']),
        ('Severity CR', 'Aceptable (1.74-3%)', sev_c['aceptable']),
        ('Severity CR', 'Revisar (1-1.74%)', sev_c['revisar']),
        ('Severity CR', 'Crítica (0.5-1%)', sev_c['critica']),
        ('Severity CR', 'Súper Crítica (<0.5%)', sev_c['super']),
        ('Metodología', 'Periodo', periodo),
        ('Metodología', 'Filtro P80', '80% del CK total'),
    ]
    for i, row in enumerate(ficha, start=5):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = THIN_BORDER
            if i == 5:
                c.font = HEADER_FONT_WHITE
                c.fill = HEADER_FILL
                c.alignment = CENTER
            else:
                c.font = CELL_FONT
                c.alignment = LEFT if j <= 2 else RIGHT
    set_widths(ws, [22, 35, 25])
    
    # 2 · Severidad Eficacia y CR
    ws = wb.create_sheet('Severidad Eficacia y CR')
    ws['A1'] = 'SEVERIDAD POR EFICACIA Y CONV RATE · HTS P80'
    ws['A1'].font = Font(name='Arial', bold=True, size=12, color=ACCENT)
    ws.merge_cells('A1:G1')
    ws['A3'] = 'EFICACIA'
    ws['A3'].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws['A3'].fill = HEADER_FILL
    ws['A3'].alignment = CENTER
    ws.merge_cells('A3:C3')
    ws['E3'] = 'CONV RATE'
    ws['E3'].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws['E3'].fill = HEADER_FILL
    ws['E3'].alignment = CENTER
    ws.merge_cells('E3:G3')
    write_header(ws, ['Nivel', 'Rango', 'Hoteles', '', 'Nivel', 'Rango', 'Hoteles'], 4, ACCENT)
    
    efic_rows = [
        ('Exitosa', '> 97%', sev_e['exitosa']),
        ('Aceptable', '93-97%', sev_e['aceptable']),
        ('Revisar', '85-93%', sev_e['revisar']),
        ('Crítica', '60-85%', sev_e['critica']),
        ('Súper Crítica', '< 60%', sev_e['super']),
    ]
    cr_rows = [
        ('Exitosa', '> 3%', sev_c['exitosa']),
        ('Aceptable', '1.74-3%', sev_c['aceptable']),
        ('Revisar', '1-1.74%', sev_c['revisar']),
        ('Crítica', '0.5-1%', sev_c['critica']),
        ('Súper Crítica', '< 0.5%', sev_c['super']),
    ]
    for i, (e, c) in enumerate(zip(efic_rows, cr_rows), start=5):
        write_row(ws, i, [e[0], e[1], e[2], '', c[0], c[1], c[2]],
                  [LEFT, CENTER, RIGHT, CENTER, LEFT, CENTER, RIGHT])
    ws.cell(row=10, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=10, column=3, value='=SUM(C5:C9)').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=10, column=5, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=10, column=7, value='=SUM(G5:G9)').font = Font(name='Arial', bold=True, size=10)
    for col in [1, 3, 5, 7]:
        ws.cell(row=10, column=col).border = THIN_BORDER
    set_widths(ws, [18, 14, 12, 4, 18, 14, 12])
    ws.freeze_panes = 'A5'
    
    # 3 · Hoteles Críticos
    ws = wb.create_sheet('Hoteles Críticos')
    write_header(ws, ['#', 'Hotel', 'Corporativo', 'CheckRates', 'Bookings', 'Eficacia', 'Conv Rate', 'Cluster'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_critic'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['Hotel_clean'], r['Corporate'], int(r['CK']), int(r['BKGS']), r['Eficacia']/100, r['CR']/100, r['Cluster']],
                  [CENTER, LEFT, LEFT, RIGHT, RIGHT, RIGHT, RIGHT, LEFT],
                  [None, None, None, '#,##0', '#,##0', '0.0%', '0.00%', None])
    set_widths(ws, [5, 45, 20, 12, 10, 10, 10, 22])
    ws.freeze_panes = 'A2'
    
    # 4 · Bajo Rendimiento
    ws = wb.create_sheet('Hoteles con Bajo Rend')
    write_header(ws, ['#', 'Hotel', 'Corporativo', 'CheckRates', 'Bookings', 'Eficacia', 'Conv Rate'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_bajo'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['Hotel_clean'], r['Corporate'], int(r['CK']), int(r['BKGS']), r['Eficacia']/100, r['CR']/100],
                  [CENTER, LEFT, LEFT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, None, '#,##0', '#,##0', '0.0%', '0.00%'])
    set_widths(ws, [5, 45, 20, 12, 10, 10, 10])
    ws.freeze_panes = 'A2'
    
    # 5 · Concentración por Corp
    ws = wb.create_sheet('Concentración por Corp')
    write_header(ws, ['#', 'Corporativo', 'Hoteles P80', 'CheckRates', 'Crit+ Efic', '% Portfolio Efic', '% Share Efic',
                      'Crit+ CR', '% Portfolio CR', '% Share CR'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_corp'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['Corporate'], int(r['total_hot']), int(r['CK']), int(r['crit_e']),
                   r['pct_e'], r['share_e'], int(r['crit_c']), r['pct_c'], r['share_c']],
                  [CENTER, LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, '#,##0', '#,##0', '#,##0', '0.0%', '0.0%', '#,##0', '0.0%', '0.0%'])
    set_widths(ws, [5, 25, 12, 12, 12, 16, 14, 12, 16, 14])
    ws.freeze_panes = 'C2'
    
    # 6 · Hoteles que no convierten
    ws = wb.create_sheet('Hoteles que no convierten')
    write_header(ws, ['#', 'Hotel', 'Corporativo', 'CheckRates', 'Bookings', 'Eficacia', 'Conv Rate'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_noconv'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['Hotel_clean'], r['Corporate'], int(r['CK']), int(r['BKGS']), r['Eficacia']/100, r['CR']/100],
                  [CENTER, LEFT, LEFT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, None, '#,##0', '#,##0', '0.0%', '0.00%'])
    set_widths(ws, [5, 45, 20, 12, 10, 10, 10])
    ws.freeze_panes = 'A2'
    
    # 7 · Plan de Acción (template estándar · personalizable después)
    ws = wb.create_sheet('Plan de Acción')
    write_header(ws, ['#', 'Tipo', 'Plazo', 'Acción', 'Aplica a', 'Owner sugerido'], 1, ACCENT)
    plan = [
        (1, 'Quick Win', '< 7 días', 'Auditoría Top 10 hoteles que no convierten · cluster Connectivity Issue', 'Top 10', 'Supply Tech'),
        (2, 'Quick Win', '< 7 días', 'Escalar a integración hoteles con Eficacia <60%', 'Súper Críticos', 'Integration Lead'),
        (3, 'Quick Win', '< 7 días', 'Validar pricing en Hybrids (Eficacia >95% pero CR <0.2%)', 'Hybrids', 'Pricing'),
        (4, 'Medium Priority', '2-4 semanas', 'QBR específica con Top corp con mayor %Share', 'Top Corp', 'Account Manager'),
        (5, 'Medium Priority', '2-4 semanas', 'Análisis técnico cluster geográfico crítico', 'Cluster Geo', 'Supply Tech'),
        (6, 'Medium Priority', '2-4 semanas', 'Revisar conectividad providers con Eficacia <90%', 'Por Provider', 'Integration Lead'),
        (7, 'Estratégico', 'Trimestre', 'SLA con Corporativos top de canasta OP', 'Canasta OP', 'Director Supply'),
        (8, 'Estratégico', 'Trimestre', 'Programa B2C activación', 'Canasta B2C', 'Revenue + Marketing'),
        (9, 'Estratégico', 'Trimestre', 'Dashboard semanal automatizado', 'Tooling', 'Analytics'),
    ]
    for r in plan:
        row_idx = ws.max_row + 1
        for col_idx, val in enumerate(r, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = LEFT if col_idx > 1 else CENTER
            if col_idx == 4:
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    set_widths(ws, [5, 18, 14, 70, 18, 22])
    ws.freeze_panes = 'A2'
    
    # 8-10 · Canastas
    for canasta_name in ['B2C', 'OP', 'CUG']:
        ws = wb.create_sheet(f'Canasta {canasta_name}')
        write_header(ws, ['#', 'Hotel', 'Corporativo', 'CheckRates', 'Bookings', 'Eficacia', 'Conv Rate'], 1, ACCENT)
        df_c = data['canastas'][canasta_name]
        for i, (_, r) in enumerate(df_c.iterrows(), start=2):
            write_row(ws, i,
                      [i-1, r['Hotel_clean'], r['Corporate'], int(r['CK']), int(r['BKGS']), r['Eficacia']/100, r['CR']/100],
                      [CENTER, LEFT, LEFT, RIGHT, RIGHT, RIGHT, RIGHT],
                      [None, None, None, '#,##0', '#,##0', '0.0%', '0.00%'])
        set_widths(ws, [5, 45, 20, 12, 10, 10, 10])
        ws.freeze_panes = 'A2'
    
    # 11 · Dataset HTS P80
    ws = wb.create_sheet('Dataset HTS P80')
    write_header(ws, ['Hotel', 'Corporativo', 'CheckRates', 'Bookings', 'Eficacia', 'Conv Rate'], 1, '8A8377')
    for i, (_, r) in enumerate(data['p80_df'].iterrows(), start=2):
        write_row(ws, i,
                  [r['Hotel_clean'], r['Corporate'], int(r['CK']), int(r['BKGS']), r['Eficacia']/100, r['CR']/100],
                  [LEFT, LEFT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, '#,##0', '#,##0', '0.0%', '0.00%'])
    set_widths(ws, [45, 20, 12, 10, 10, 10])
    ws.freeze_panes = 'A2'
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_rnd_xlsx(data: dict, week: int, periodo: str, output_path: Path):
    """Genera Analisis_Rates_NoDispo_7d_WNN.xlsx con 11 pestañas Top 50."""
    ACCENT = 'EA0074'
    HEADER_FILL = PatternFill('solid', fgColor=ACCENT)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    kpis = data['kpis']
    sev = data['severity']
    has_corp = kpis['has_corp']
    
    # 1 · Ficha Técnica
    ws = wb.create_sheet('Ficha Técnica')
    ws['A1'] = f'RATES NO DISPO · WEEK {week} · {periodo}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=ACCENT)
    ws.merge_cells('A1:C1')
    ws['A3'] = 'Ficha Técnica del Análisis'
    ws['A3'].font = Font(name='Arial', bold=True, size=12)
    ws.merge_cells('A3:C3')
    
    ficha = [
        ('SECCIÓN', 'INDICADOR', 'VALOR'),
        ('Universo', 'Total hoteles activos', kpis['total_hot']),
        ('Universo', 'Hoteles HTS P80', kpis['p80_count']),
        ('Universo', 'Tráfico total', f'{kpis["total_traf"]/1e6:.1f}M'),
        ('Universo', 'Bookings totales', kpis['total_bkgs']),
        ('Universo', 'GB total', f'${kpis["total_gb"]/1e6:.2f}M'),
        ('Universo', 'Hoteles con 0 BKGS', kpis['zero_bkgs']),
        ('Universo', '% hoteles con 0 BKGS', f'{kpis["zero_bkgs_pct"]:.1f}%'),
        ('Métricas globales', '%NoDispo ponderado', f'{kpis["nodispo_pond"]:.2f}%'),
        ('Severity %NoDispo', 'Exitosa (0-3%)', sev['exitosa']),
        ('Severity %NoDispo', 'Aceptable (3-5%)', sev['aceptable']),
        ('Severity %NoDispo', 'Revisar (5-20%)', sev['revisar']),
        ('Severity %NoDispo', 'Crítica (20-60%)', sev['critica']),
        ('Severity %NoDispo', 'Súper Crítica (>60%)', sev['super']),
        ('Metodología', 'Periodo', periodo),
        ('Metodología', 'Tiene CorpName?', 'SÍ' if has_corp else 'NO · usar Destino como proxy'),
    ]
    for i, row in enumerate(ficha, start=5):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = THIN_BORDER
            if i == 5:
                c.font = HEADER_FONT_WHITE
                c.fill = HEADER_FILL
                c.alignment = CENTER
            else:
                c.font = CELL_FONT
                c.alignment = LEFT if j <= 2 else RIGHT
    set_widths(ws, [22, 35, 25])
    
    # 2 · Severidad
    ws = wb.create_sheet('Severidad NoDispo')
    ws['A1'] = 'SEVERIDAD POR % NO DISPONIBILIDAD · HTS P80'
    ws['A1'].font = Font(name='Arial', bold=True, size=12, color=ACCENT)
    ws.merge_cells('A1:C1')
    write_header(ws, ['Nivel', 'Rango', 'Hoteles'], 3, ACCENT)
    sev_rows = [
        ('Exitosa', '0-3%', sev['exitosa']),
        ('Aceptable', '3-5%', sev['aceptable']),
        ('Revisar', '5-20%', sev['revisar']),
        ('Crítica', '20-60%', sev['critica']),
        ('Súper Crítica', '> 60%', sev['super']),
    ]
    for i, s in enumerate(sev_rows, start=4):
        write_row(ws, i, list(s), [LEFT, CENTER, RIGHT])
    ws.cell(row=9, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=9, column=3, value='=SUM(C4:C8)').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=9, column=3).alignment = RIGHT
    for col in range(1, 4):
        ws.cell(row=9, column=col).border = THIN_BORDER
    set_widths(ws, [18, 14, 12])
    ws.freeze_panes = 'A4'
    
    # 3 · Demanda No Convertida
    ws = wb.create_sheet('Demanda No Convertida')
    cols = ['#', 'Hotel'] + (['Corporativo'] if has_corp else []) + ['Tráfico (M)', '%NoDispo', 'Bookings', 'GB']
    write_header(ws, cols, 1, ACCENT)
    for i, (_, r) in enumerate(data['top_demanda'].iterrows(), start=2):
        vals = [i-1, r['Hotel']]
        if has_corp: vals.append(r.get('CorpName', 'N/A'))
        vals.extend([r['Trafico']/1_000_000, r['%NoDispo']/100, int(r['Bookings']), r['gb_usd']])
        aligns = [CENTER, LEFT] + ([LEFT] if has_corp else []) + [RIGHT]*4
        fmts = [None, None] + ([None] if has_corp else []) + ['0.00', '0.0%', '#,##0', '$#,##0']
        write_row(ws, i, vals, aligns, fmts)
    set_widths(ws, [5, 45, 22, 12, 10, 10, 12] if has_corp else [5, 45, 12, 10, 10, 12])
    ws.freeze_panes = 'A2'
    
    # 4 · Bajo Rendimiento
    ws = wb.create_sheet('Hoteles con Bajo Rend')
    cols = ['#', 'Hotel'] + (['Corporativo'] if has_corp else []) + ['Tráfico (M)', '%NoDispo', 'Bookings', 'GB', 'RPM']
    write_header(ws, cols, 1, ACCENT)
    for i, (_, r) in enumerate(data['top_bajo'].iterrows(), start=2):
        vals = [i-1, r['Hotel']]
        if has_corp: vals.append(r.get('CorpName', 'N/A'))
        vals.extend([r['Trafico']/1_000_000, r['%NoDispo']/100, int(r['Bookings']), r['gb_usd'], r['RPM']])
        aligns = [CENTER, LEFT] + ([LEFT] if has_corp else []) + [RIGHT]*5
        fmts = [None, None] + ([None] if has_corp else []) + ['0.00', '0.0%', '#,##0', '$#,##0', '0.00']
        write_row(ws, i, vals, aligns, fmts)
    set_widths(ws, [5, 45, 22, 12, 10, 10, 12, 10] if has_corp else [5, 45, 12, 10, 10, 12, 10])
    ws.freeze_panes = 'A2'
    
    # 5 · Concentración por Corp (si tiene CorpName)
    ws = wb.create_sheet('Concentración por Corp')
    if has_corp and len(data['top_corp']) > 0:
        write_header(ws, ['#', 'Corporativo', 'Hoteles P80', 'Tráfico (M)', 'Críticos+', '% Portfolio', '% Share'], 1, ACCENT)
        for i, (_, r) in enumerate(data['top_corp'].iterrows(), start=2):
            write_row(ws, i,
                      [i-1, r['CorpName'], int(r['total_hot']), r['Trafico']/1_000_000,
                       int(r['crit']), r['pct_port'], r['pct_share']],
                      [CENTER, LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT],
                      [None, None, '#,##0', '0.00', '#,##0', '0.0%', '0.0%'])
        set_widths(ws, [5, 30, 12, 12, 12, 14, 12])
        ws.freeze_panes = 'C2'
    else:
        ws['A1'] = '⚠ Dataset no tiene columna CorpName · usar pestaña "Por Destino" como proxy'
        ws['A1'].font = Font(name='Arial', bold=True, size=11, color='C0392B')
    
    # 6 · Por Destino
    ws = wb.create_sheet('Por Destino')
    write_header(ws, ['#', 'Destino', 'Hoteles', 'Tráfico (M)', '%NoDispo', 'Bookings', 'GB', 'RPM'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_dest'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['Destino'], int(r['Hot']), r['Trafico']/1_000_000, r['%NoDispo']/100,
                   int(r['Bookings']), r['gb_usd'], r['RPM']],
                  [CENTER, LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, '#,##0', '0.00', '0.0%', '#,##0', '$#,##0', '0.00'])
    set_widths(ws, [5, 30, 10, 12, 10, 10, 12, 10])
    ws.freeze_panes = 'A2'
    
    # 7 · Por País
    ws = wb.create_sheet('Por País')
    write_header(ws, ['#', 'País', 'Hoteles', 'Tráfico (M)', '%NoDispo', 'Bookings', 'GB', 'RPM'], 1, ACCENT)
    for i, (_, r) in enumerate(data['top_pais'].iterrows(), start=2):
        write_row(ws, i,
                  [i-1, r['PaisDestino'], int(r['Hot']), r['Trafico']/1_000_000, r['%NoDispo']/100,
                   int(r['Bookings']), r['gb_usd'], r['RPM']],
                  [CENTER, LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT],
                  [None, None, '#,##0', '0.00', '0.0%', '#,##0', '$#,##0', '0.00'])
    set_widths(ws, [5, 25, 10, 12, 10, 10, 12, 10])
    ws.freeze_panes = 'A2'
    
    # 8 · Plan de Acción
    ws = wb.create_sheet('Plan de Acción')
    write_header(ws, ['#', 'Tipo', 'Plazo', 'Acción', 'Aplica a', 'Owner sugerido'], 1, ACCENT)
    plan = [
        (1, 'Quick Win', '< 7 días', 'Auditoría Top 10 Demanda No Convertida (alto Tráfico · 0 BKGS)', 'Top 10', 'Supply Tech'),
        (2, 'Quick Win', '< 7 días', 'Revisar mapeo de inventario en hoteles con %NoDispo > 60%', 'Súper Críticos', 'Inventory Mgmt'),
        (3, 'Quick Win', '< 7 días', 'Validar contratos providers para hoteles top con BKGS = 0', 'Top Demanda', 'Account Manager'),
        (4, 'Medium Priority', '2-4 semanas', 'Auditar destino más crítico (cluster geográfico)', 'Cluster Destino', 'Pricing'),
        (5, 'Medium Priority', '2-4 semanas', 'Revisar parametrización %NoDispo con providers', 'Por Destino', 'Supply Tech'),
        (6, 'Medium Priority', '2-4 semanas', 'Activación selectiva top corp con mayor %Share crítico', 'Top Corp', 'Account Manager'),
        (7, 'Estratégico', 'Trimestre', 'SLA con corporativos top con %NoDispo > 5%', 'Por Corporativo', 'Director Supply'),
        (8, 'Estratégico', 'Trimestre', 'Programa B2C activación', 'Canasta B2C', 'Revenue + Marketing'),
        (9, 'Estratégico', 'Trimestre', 'Dashboard semanal automatizado', 'Tooling', 'Analytics'),
    ]
    for r in plan:
        row_idx = ws.max_row + 1
        for col_idx, val in enumerate(r, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = LEFT if col_idx > 1 else CENTER
            if col_idx == 4:
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    set_widths(ws, [5, 18, 14, 70, 22, 22])
    ws.freeze_panes = 'A2'
    
    # 9-11 · Canastas
    for canasta_name in ['B2C', 'OP', 'CUG']:
        ws = wb.create_sheet(f'Canasta {canasta_name}')
        df_c = data['canastas'].get(canasta_name, pd.DataFrame())
        cols = ['#', 'Hotel'] + (['Corporativo'] if has_corp else []) + ['Tráfico (M)', '%NoDispo', 'Bookings', 'GB', 'RPM']
        write_header(ws, cols, 1, ACCENT)
        if not df_c.empty:
            for i, (_, r) in enumerate(df_c.iterrows(), start=2):
                vals = [i-1, r['Hotel']]
                if has_corp: vals.append(r.get('CorpName', 'N/A'))
                vals.extend([r['Trafico']/1_000_000, r['%NoDispo']/100, int(r['Bookings']), r['gb_usd'], r['RPM']])
                aligns = [CENTER, LEFT] + ([LEFT] if has_corp else []) + [RIGHT]*5
                fmts = [None, None] + ([None] if has_corp else []) + ['0.00', '0.0%', '#,##0', '$#,##0', '0.00']
                write_row(ws, i, vals, aligns, fmts)
        set_widths(ws, [5, 45, 22, 12, 10, 10, 12, 10] if has_corp else [5, 45, 12, 10, 10, 12, 10])
        ws.freeze_panes = 'A2'
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
