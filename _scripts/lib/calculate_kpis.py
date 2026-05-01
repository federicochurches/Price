"""
Cálculo de KPIs y agregaciones desde datasets crudos.
"""
import pandas as pd
from pathlib import Path


def calculate_cr_kpis(dataset_path: Path) -> dict:
    """
    Calcula todos los KPIs y agregados de CheckRates.
    
    Returns dict con:
        - kpis: total_hot, p80_count, total_ck, total_bkgs, eficacia, cr, zero_bkgs
        - severity_eficacia: dict con counts por nivel
        - severity_cr: dict con counts por nivel
        - top_critic: DataFrame Top 50 hoteles críticos con cluster
        - top_bajo: DataFrame Top 50 bajo rendimiento
        - top_corp: DataFrame Top 50 corporativos con %Portfolio + %Share
        - top_noconv: DataFrame Top 50 que no convierten
        - canastas: dict {B2C, OP, CUG} con DataFrames Top 50 cada una
        - p80_df: DataFrame con todos los hoteles del P80
    """
    df = pd.read_excel(dataset_path, sheet_name=0)
    df['CK'] = pd.to_numeric(df['CheckRates Únicos'], errors='coerce').fillna(0)
    df['BKGS'] = pd.to_numeric(df['Bookings'], errors='coerce').fillna(0)
    df['Success'] = pd.to_numeric(df['Successful UniqueChkRts'], errors='coerce').fillna(0)
    df['Hotel_clean'] = df['Hotel'].str.replace(r'^\(\d+\)\s*-\s*', '', regex=True)
    
    # Agregación a nivel hotel
    ag = df.groupby(['Hotel_clean', 'Corporate']).agg(
        CK=('CK', 'sum'), BKGS=('BKGS', 'sum'), Success=('Success', 'sum'),
    ).reset_index()
    ag['Eficacia'] = ag['Success'] / ag['CK'] * 100
    ag['CR'] = ag['BKGS'] / ag['CK'] * 100
    ag = ag.sort_values('CK', ascending=False)
    total_ck = ag['CK'].sum()
    ag['ck_acum'] = ag['CK'].cumsum() / total_ck
    p80 = ag[ag['ck_acum'] <= 0.80].copy()
    
    # KPIs globales
    kpis = {
        'total_hot': int(len(ag)),
        'p80_count': int(len(p80)),
        'total_ck': float(total_ck),
        'total_bkgs': int(ag['BKGS'].sum()),
        'eficacia': float(ag['Success'].sum() / total_ck * 100),
        'cr': float(ag['BKGS'].sum() / total_ck * 100),
        'zero_bkgs': int((ag['BKGS'] == 0).sum()),
    }
    kpis['zero_bkgs_pct'] = kpis['zero_bkgs'] / kpis['total_hot'] * 100
    
    # Severity Eficacia
    severity_eficacia = {
        'exitosa': int((p80['Eficacia'] > 97).sum()),
        'aceptable': int(((p80['Eficacia'] >= 93) & (p80['Eficacia'] <= 97)).sum()),
        'revisar': int(((p80['Eficacia'] >= 85) & (p80['Eficacia'] < 93)).sum()),
        'critica': int(((p80['Eficacia'] >= 60) & (p80['Eficacia'] < 85)).sum()),
        'super': int((p80['Eficacia'] < 60).sum()),
    }
    
    # Severity CR
    severity_cr = {
        'exitosa': int((p80['CR'] > 3).sum()),
        'aceptable': int(((p80['CR'] >= 1.74) & (p80['CR'] <= 3)).sum()),
        'revisar': int(((p80['CR'] >= 1) & (p80['CR'] < 1.74)).sum()),
        'critica': int(((p80['CR'] >= 0.5) & (p80['CR'] < 1)).sum()),
        'super': int((p80['CR'] < 0.5).sum()),
    }
    
    # Top 50 críticos con cluster
    def cluster(ef, cr):
        if ef < 60 and cr < 0.1: return 'Connectivity Issue'
        if ef < 60: return 'Tech Issue'
        if cr < 0.5: return 'Conversion Issue'
        if cr < 1: return 'Hybrid Issue'
        return 'Quick Win'
    
    critic = p80[p80['CK'] >= 1000].copy()
    critic['rank_ef'] = critic['Eficacia'].rank()
    critic['rank_cr'] = critic['CR'].rank()
    critic['combined'] = critic['rank_ef'] + critic['rank_cr']
    critic = critic.sort_values('combined').head(50)
    critic['Cluster'] = critic.apply(lambda r: cluster(r['Eficacia'], r['CR']), axis=1)
    
    # Top 50 Bajo Rendimiento
    bajo = p80[(p80['Eficacia'] <= 95) & (p80['BKGS'] > 0) & (p80['CK'] >= 1000)].copy()
    bajo = bajo.sort_values('CK', ascending=False).head(50)
    
    # Concentración por Corp
    corp = p80.groupby('Corporate').agg(
        total_hot=('Hotel_clean', 'count'),
        CK=('CK', 'sum'),
        BKGS=('BKGS', 'sum'),
    ).reset_index()
    corp = corp[corp['total_hot'] >= 5]
    
    crit_e = p80[p80['Eficacia'] < 60].groupby('Corporate').size().reset_index(name='crit_e')
    crit_c = p80[p80['CR'] < 0.5].groupby('Corporate').size().reset_index(name='crit_c')
    corp = corp.merge(crit_e, on='Corporate', how='left').fillna(0)
    corp = corp.merge(crit_c, on='Corporate', how='left').fillna(0)
    corp['crit_e'] = corp['crit_e'].astype(int)
    corp['crit_c'] = corp['crit_c'].astype(int)
    corp['pct_e'] = corp['crit_e'] / corp['total_hot']
    corp['pct_c'] = corp['crit_c'] / corp['total_hot']
    total_e = corp['crit_e'].sum()
    total_c = corp['crit_c'].sum()
    corp['share_e'] = corp['crit_e'] / total_e if total_e > 0 else 0
    corp['share_c'] = corp['crit_c'] / total_c if total_c > 0 else 0
    corp['order'] = corp['crit_e'] + corp['crit_c']
    corp = corp.sort_values('order', ascending=False).head(50)
    
    # Top 50 No Convierten
    noconv = p80[(p80['BKGS'] >= 1) & (p80['CK'] >= 2000)].copy()
    noconv = noconv.sort_values('CR').head(50)
    
    # Canastas
    canastas = {}
    for nombre, dist_cat in [('B2C', 'B2C'), ('OP', 'B2B (OP)'), ('CUG', 'CUG (UOP)')]:
        df_sub = df[df['DistributionCategory'] == dist_cat]
        ag_c = df_sub.groupby(['Hotel_clean', 'Corporate']).agg(
            CK=('CK', 'sum'), BKGS=('BKGS', 'sum'), Success=('Success', 'sum'),
        ).reset_index()
        ag_c['Eficacia'] = ag_c['Success'] / ag_c['CK'] * 100
        ag_c['CR'] = ag_c['BKGS'] / ag_c['CK'] * 100
        ag_c = ag_c.sort_values('CK', ascending=False)
        total_c = ag_c['CK'].sum()
        ag_c['ck_acum'] = ag_c['CK'].cumsum() / total_c
        p80c = ag_c[ag_c['ck_acum'] <= 0.80].copy()
        p80c['rank_e'] = p80c['Eficacia'].rank()
        p80c['rank_c'] = p80c['CR'].rank()
        p80c['combined'] = p80c['rank_e'] + p80c['rank_c']
        canastas[nombre] = p80c.sort_values('combined').head(50)
    
    return {
        'kpis': kpis,
        'severity_eficacia': severity_eficacia,
        'severity_cr': severity_cr,
        'top_critic': critic,
        'top_bajo': bajo,
        'top_corp': corp,
        'top_noconv': noconv,
        'canastas': canastas,
        'p80_df': p80,
    }


def calculate_rnd_kpis(dataset_path: Path) -> dict:
    """
    Calcula todos los KPIs y agregados de Rates No Dispo.
    Soporta datasets con CorpName o sin (usa Destino como proxy si falta).
    """
    # Detectar pestañas · si tiene 'Canasta ALL' usar esa, si no usar la primera
    import openpyxl
    wb = openpyxl.load_workbook(dataset_path, read_only=True)
    sheet_name = 'Canasta ALL' if 'Canasta ALL' in wb.sheetnames else wb.sheetnames[0]
    
    df = pd.read_excel(dataset_path, sheet_name=sheet_name)
    # Filtrar filas vacías
    if 'Hotel' in df.columns:
        df = df[df['Hotel'] != '-'].copy()
    
    df['Trafico'] = pd.to_numeric(df['Trafico'], errors='coerce').fillna(0)
    df['Bookings'] = pd.to_numeric(df['Bookings'], errors='coerce').fillna(0)
    df['gb_usd'] = pd.to_numeric(df['gb_usd'], errors='coerce').fillna(0)
    df['%NoDispo'] = pd.to_numeric(df['%NoDispo'], errors='coerce').fillna(0)
    
    # Detectar formato de %NoDispo (decimal 0-1 o porcentaje 0-100)
    if df['%NoDispo'].max() <= 2.0:
        df['NoDispo_traf'] = df['%NoDispo'] * df['Trafico']
    else:
        df['NoDispo_traf'] = df['%NoDispo'] / 100 * df['Trafico']
    
    # Limpiar strings
    df['Hotel'] = df['Hotel'].astype(str).str.replace(r'^\\t', '', regex=True).str.replace(r'\\t$', '', regex=True).str.strip()
    df['Destino'] = df['Destino'].astype(str).str.strip()
    df['PaisDestino'] = df['PaisDestino'].astype(str).str.strip()
    
    has_corp = 'CorpName' in df.columns
    if has_corp:
        df['CorpName'] = df['CorpName'].fillna('Independiente').astype(str).str.strip()
        groupby_cols = ['Hotel', 'CorpName']
    else:
        groupby_cols = ['Hotel']
    
    # Agregación a nivel hotel
    ag = df.groupby(groupby_cols).agg(
        Trafico=('Trafico', 'sum'),
        Bookings=('Bookings', 'sum'),
        gb_usd=('gb_usd', 'sum'),
        NoDispo_traf=('NoDispo_traf', 'sum'),
    ).reset_index()
    ag = ag[ag['Trafico'] > 0].copy()
    ag['%NoDispo'] = ag['NoDispo_traf'] / ag['Trafico'] * 100
    ag['RPM'] = ag['Bookings'] * 1_000_000 / ag['Trafico']
    
    if not has_corp:
        ag['CorpName'] = 'N/A'
    
    ag = ag.sort_values('Trafico', ascending=False)
    total_traf = ag['Trafico'].sum()
    ag['traf_acum'] = ag['Trafico'].cumsum() / total_traf
    p80 = ag[ag['traf_acum'] <= 0.80].copy()
    
    # KPIs
    kpis = {
        'total_hot': int(len(ag)),
        'p80_count': int(len(p80)),
        'total_traf': float(total_traf),
        'total_bkgs': int(ag['Bookings'].sum()),
        'total_gb': float(ag['gb_usd'].sum()),
        'nodispo_pond': float(ag['NoDispo_traf'].sum() / total_traf * 100),
        'zero_bkgs': int((ag['Bookings'] == 0).sum()),
        'has_corp': has_corp,
    }
    kpis['zero_bkgs_pct'] = kpis['zero_bkgs'] / kpis['total_hot'] * 100
    
    # Severity %NoDispo
    severity = {
        'exitosa': int((p80['%NoDispo'] <= 3).sum()),
        'aceptable': int(((p80['%NoDispo'] > 3) & (p80['%NoDispo'] <= 5)).sum()),
        'revisar': int(((p80['%NoDispo'] > 5) & (p80['%NoDispo'] <= 20)).sum()),
        'critica': int(((p80['%NoDispo'] > 20) & (p80['%NoDispo'] <= 60)).sum()),
        'super': int((p80['%NoDispo'] > 60).sum()),
    }
    
    # Top 50 Demanda No Convertida
    demanda = ag[(ag['Bookings'] == 0) & (ag['Trafico'] >= 100_000)].copy()
    demanda = demanda.sort_values('Trafico', ascending=False).head(50)
    
    # Top 50 Bajo Rendimiento
    bajo = ag[(ag['Bookings'] > 0) & (ag['%NoDispo'] > 5) & (ag['Trafico'] >= 100_000)].copy()
    bajo = bajo.sort_values('Trafico', ascending=False).head(50)
    
    # Concentración por Corp (si hay) o Destino (proxy)
    if has_corp:
        corp = p80.groupby('CorpName').agg(
            total_hot=('Hotel', 'count'),
            Trafico=('Trafico', 'sum'),
        ).reset_index()
        corp = corp[corp['total_hot'] >= 3]
        crit = p80[p80['%NoDispo'] > 20].groupby('CorpName').size().reset_index(name='crit')
        corp = corp.merge(crit, on='CorpName', how='left').fillna(0)
        corp['crit'] = corp['crit'].astype(int)
        corp['pct_port'] = corp['crit'] / corp['total_hot']
        total_crit = corp['crit'].sum()
        corp['pct_share'] = corp['crit'] / total_crit if total_crit > 0 else 0
        corp = corp.sort_values('crit', ascending=False).head(50)
    else:
        corp = pd.DataFrame()  # vacío
    
    # Por Destino
    dest = df.groupby('Destino').agg(
        Trafico=('Trafico', 'sum'),
        Bookings=('Bookings', 'sum'),
        gb_usd=('gb_usd', 'sum'),
        NoDispo_traf=('NoDispo_traf', 'sum'),
        Hot=('Hotel', 'nunique'),
    ).reset_index()
    dest = dest[dest['Trafico'] > 0]
    dest['%NoDispo'] = dest['NoDispo_traf'] / dest['Trafico'] * 100
    dest['RPM'] = dest['Bookings'] * 1_000_000 / dest['Trafico']
    dest = dest.sort_values('Trafico', ascending=False).head(50)
    
    # Por País
    pais = df.groupby('PaisDestino').agg(
        Trafico=('Trafico', 'sum'),
        Bookings=('Bookings', 'sum'),
        gb_usd=('gb_usd', 'sum'),
        NoDispo_traf=('NoDispo_traf', 'sum'),
        Hot=('Hotel', 'nunique'),
    ).reset_index()
    pais = pais[pais['Trafico'] > 0]
    pais['%NoDispo'] = pais['NoDispo_traf'] / pais['Trafico'] * 100
    pais['RPM'] = pais['Bookings'] * 1_000_000 / pais['Trafico']
    pais = pais.sort_values('Trafico', ascending=False).head(50)
    
    # Canastas (si dataset tiene pestañas separadas)
    canastas = {}
    canastas_map = [('B2C', 'Canasta B2C'), ('OP', 'Canasta OP'), ('CUG', 'Canasta UOP')]
    for nombre, sheet in canastas_map:
        if sheet in wb.sheetnames:
            df_sub = pd.read_excel(dataset_path, sheet_name=sheet)
        else:
            # Fallback: filtrar por DistributionCategory
            dist_map = {'B2C': 'B2C', 'OP': 'B2B (OP)', 'CUG': 'CUG (UOP)'}
            df_sub = df[df['DistributionCategory'] == dist_map[nombre]] if 'DistributionCategory' in df.columns else df
        
        if len(df_sub) == 0:
            canastas[nombre] = pd.DataFrame()
            continue
        
        df_sub['Trafico'] = pd.to_numeric(df_sub['Trafico'], errors='coerce').fillna(0)
        df_sub['Bookings'] = pd.to_numeric(df_sub['Bookings'], errors='coerce').fillna(0)
        df_sub['gb_usd'] = pd.to_numeric(df_sub['gb_usd'], errors='coerce').fillna(0)
        df_sub['%NoDispo'] = pd.to_numeric(df_sub['%NoDispo'], errors='coerce').fillna(0)
        if df_sub['%NoDispo'].max() <= 2.0:
            df_sub['NoDispo_traf'] = df_sub['%NoDispo'] * df_sub['Trafico']
        else:
            df_sub['NoDispo_traf'] = df_sub['%NoDispo'] / 100 * df_sub['Trafico']
        df_sub['Hotel'] = df_sub['Hotel'].astype(str).str.strip()
        if 'CorpName' in df_sub.columns:
            df_sub['CorpName'] = df_sub['CorpName'].fillna('Independiente').astype(str).str.strip()
            gcols = ['Hotel', 'CorpName']
        else:
            gcols = ['Hotel']
        
        ag_c = df_sub.groupby(gcols).agg(
            Trafico=('Trafico', 'sum'),
            Bookings=('Bookings', 'sum'),
            gb_usd=('gb_usd', 'sum'),
            NoDispo_traf=('NoDispo_traf', 'sum'),
        ).reset_index()
        ag_c = ag_c[ag_c['Trafico'] > 0].copy()
        if not ag_c.empty:
            ag_c['%NoDispo'] = ag_c['NoDispo_traf'] / ag_c['Trafico'] * 100
            ag_c['RPM'] = ag_c['Bookings'] * 1_000_000 / ag_c['Trafico']
            ag_c = ag_c.sort_values('Trafico', ascending=False).head(50)
            if 'CorpName' not in ag_c.columns:
                ag_c['CorpName'] = 'N/A'
        canastas[nombre] = ag_c
    
    return {
        'kpis': kpis,
        'severity': severity,
        'top_demanda': demanda,
        'top_bajo': bajo,
        'top_corp': corp,
        'top_dest': dest,
        'top_pais': pais,
        'canastas': canastas,
        'p80_df': p80,
    }
