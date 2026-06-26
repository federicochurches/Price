"""
excel_rnd_accounts.py · W26+ · Excels RND por cuenta
Genera 2 archivos:
  Analisis_RND_GlobalAccounts_WNN.xlsx
  Analisis_RND_Estrategicas_WNN.xlsx

Filtro sobre pickle global por CorpName exacto.
5 hojas: Severity | Maestra | Críticos | Bajo Rendimiento | Sin Conversión

Diferencia vs regional: la dimensión primaria es el Corporativo, no la región.
Severity muestra ranking de corps + destinos dentro del grupo.
"""
import pickle, os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_nodispo, banda_convrate
from accounts_config import ACCOUNT_GROUPS

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_RND', f'rnd_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D = pickle.load(f)

TAB_ND = D['TAB_NoDispo']
M      = D['M']

RND = 'EA0074'

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP  = fill('EAF3DE'); WOW_UP_F  = fnt('2F6C34', bold=True)
WOW_DN  = fill('FCE8E6'); WOW_DN_F  = fnt('C0392B', bold=True)
WOW_NEU = fill('F2EEE6'); WOW_NEU_F = fnt('8A8377', bold=True)

RANGOS_ND = {
    'Exitosa':       '< 3%',      'Aceptable':     '3% – 5%',
    'Revisar':       '5% – 20%',  'Crítica':        '20% – 60%',
    'Súper Crítica': '> 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0', 'Crítica':    '< 0,8%',
    'Revisar':        '0,8% – 1,5%', 'Aceptable':  '1,5% – 2,5%',
    'Exitosa':        '≥ 2,5%',
}

_tc = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(RND, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(RND, 11, True)

def mk_hdr(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(RND)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_table(ws, hdr_row, end_row, n_cols, prefix):
    if end_row < hdr_row + 1: return
    _tc[0] += 1
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{end_row}'
    t = Table(displayName=f'T_{prefix}_{_tc[0]}', ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center'):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Split bandas ──────────────────────────────────────────────────────────────
def band_split(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['%NoDispo'].apply(lambda v: banda_nodispo(sf(v)) if sf(v) is not None else '—')
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Hoja Severity por cuenta ──────────────────────────────────────────────────
def write_severity(ws, df_grp, grp_label, corps):
    title(ws, f'{grp_label} · Severity RND W{VOL_NUM}', f'{PERIODO} · {len(corps)} corporativos')

    # KPI global de referencia
    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))
    nd_curr = m_curr.get('pct_nodispo', 0); nd_prev = m_prev.get('pct_nodispo', 0)
    nd_wow  = (nd_curr - nd_prev) * 100 if nd_prev else None

    r = 4
    section_title(ws, r, 'KPI Global (referencia)'); r += 1
    r = mk_hdr(ws, r, ['Métrica', f'W{int(VOL_NUM)-1}', f'W{VOL_NUM} Global', 'WoW'])
    ws.cell(r, 1, '%NoDispo').font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
    ws.cell(r, 2, round(nd_prev, 4) if nd_prev else None).border = BD
    ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
    ws.cell(r, 3, round(nd_curr, 4) if nd_curr else None).border = BD
    ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
    apply_wow(ws, r, 4, nd_wow, invert=True); r += 2

    # Severity del grupo
    sev_nd, sev_cv = {}, {}
    for _, row in df_grp.iterrows():
        nd  = sf(row.get('%NoDispo'))
        bk  = int(sf(row.get('Bookings')) or 0)
        trf = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        if nd is not None: b = banda_nodispo(nd); sev_nd[b] = sev_nd.get(b, 0) + 1
        b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else None)
        if b_cv: sev_cv[b_cv] = sev_cv.get(b_cv, 0) + 1
    total_nd = sum(sev_nd.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    section_title(ws, r, f'Severity %NoDispo · {grp_label}'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_nd.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_ND.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_nd, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevND')
    r += 1

    section_title(ws, r, f'Severity %Conv Rate · {grp_label}'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    # Ranking de Corporativos del grupo
    CORP_COLS = ['Corporativo', 'Banda ND', 'Banda CV', 'Hoteles', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    corp_rows = []
    for corp in corps:
        df_c = df_grp[df_grp['CorpName'] == corp]
        if len(df_c) == 0:
            corp_rows.append({'corp': corp, 'nd': None, 'conv': None, 'bk': 0, 'trf': 0,
                              'n_hoteles': 0, 'wow_nd': None})
            continue
        trf = int(df_c['Trafico'].sum()) if 'Trafico' in df_c else 0
        bk  = int(df_c['Bookings'].sum()) if 'Bookings' in df_c else 0
        nd_w = (df_c['%NoDispo'] * df_c['Trafico']).sum() / trf if trf > 0 else None
        conv = bk / trf if trf > 0 else None
        wow  = df_c['NoDispo_WoW_pp'].mean() if 'NoDispo_WoW_pp' in df_c.columns else None
        corp_rows.append({'corp': corp, 'nd': nd_w, 'conv': conv, 'bk': bk,
                          'trf': trf, 'n_hoteles': len(df_c), 'wow_nd': sf(wow)})

    df_corp_rank = pd.DataFrame(corp_rows).sort_values('nd', ascending=False, na_position='last')

    section_title(ws, r, f'Ranking Corporativos · %NoDispo DESC'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, CORP_COLS)
    for _, row in df_corp_rank.iterrows():
        nd   = sf(row.get('nd'))
        conv = sf(row.get('conv'))
        bk   = int(row.get('bk') or 0)
        trf  = int(row.get('trf') or 0)
        b_nd = banda_nodispo(nd) if nd is not None else '—'
        b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
        mk_cell(ws, r, 1, str(row.get('corp', '—')), align='left')
        mk_cell(ws, r, 2, b_nd, b_nd, is_sev=True)
        mk_cell(ws, r, 3, b_cv, b_cv, is_sev=True)
        mk_cell(ws, r, 4, int(row.get('n_hoteles') or 0))
        mk_cell(ws, r, 5, trf)
        pct_cell(ws, r, 6, nd)
        apply_wow(ws, r, 7, row.get('wow_nd'), invert=True)
        pct_cell(ws, r, 8, conv)
        mk_cell(ws, r, 9, bk)
        r += 1
    add_table(ws, hdr_r, r-1, len(CORP_COLS), 'Corps')

    # Top Destinos del grupo
    DEST_COLS = ['Destino', 'Corporativo', 'País', 'Banda ND', 'Banda CV', 'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'Bookings']
    df_dest = df_grp.copy()
    if len(df_dest) > 0:
        r += 1
        section_title(ws, r, 'Top Destinos del grupo · %NoDispo DESC'); r += 1
        dest_agg = {}
        for _, row in df_dest.iterrows():
            dest = str(row.get('Destino', '—'))
            corp = str(row.get('CorpName', '—'))
            pais = str(row.get('PaisDestino', row.get('Pais', '—')))
            nd   = sf(row.get('%NoDispo'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            wow  = sf(row.get('NoDispo_WoW_pp'))
            key  = (dest, corp)
            if key not in dest_agg:
                dest_agg[key] = {'pais': pais, 'trf': 0, 'bk': 0, 'nd_trf': 0.0, 'wow': wow}
            dest_agg[key]['trf']    += trf
            dest_agg[key]['bk']     += bk
            dest_agg[key]['nd_trf'] += (nd * trf) if nd is not None else 0

        dest_rows = []
        for (dest, corp), v in dest_agg.items():
            nd_w = v['nd_trf'] / v['trf'] if v['trf'] > 0 else None
            conv = v['bk'] / v['trf'] if v['trf'] > 0 else None
            dest_rows.append({'Destino': dest, 'Corp': corp, 'Pais': v['pais'],
                              'trf': v['trf'], 'bk': v['bk'], 'nd': nd_w,
                              'conv': conv, 'wow': v['wow']})
        df_dest_r = pd.DataFrame(dest_rows)
        if len(df_dest_r) > 0:
            df_dest_r = df_dest_r.sort_values('nd', ascending=False, na_position='last').head(50)
            hdr_r = r; r = mk_hdr(ws, r, DEST_COLS)
            for _, row in df_dest_r.iterrows():
                nd   = sf(row.get('nd'))
                conv = sf(row.get('conv'))
                bk   = int(row.get('bk') or 0)
                trf  = int(row.get('trf') or 0)
                b_nd = banda_nodispo(nd) if nd is not None else '—'
                b_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
                mk_cell(ws, r, 1, str(row.get('Destino', '—')), align='left')
                mk_cell(ws, r, 2, str(row.get('Corp', '—')), align='left')
                mk_cell(ws, r, 3, str(row.get('Pais', '—')), align='left')
                mk_cell(ws, r, 4, b_nd, b_nd, is_sev=True)
                mk_cell(ws, r, 5, b_cv, b_cv, is_sev=True)
                mk_cell(ws, r, 6, trf)
                pct_cell(ws, r, 7, nd)
                apply_wow(ws, r, 8, sf(row.get('wow')), invert=True)
                pct_cell(ws, r, 9, conv)
                mk_cell(ws, r, 10, bk)
                r += 1
            add_table(ws, hdr_r, r-1, len(DEST_COLS), 'Dest')

    autofit(ws, [28, 20, 18, 10, 10, 10, 10, 10, 10, 10])

# ── Hoja Maestra ──────────────────────────────────────────────────────────────
MAESTRA_COLS = [
    'País', 'Destino', 'Corporativo', 'Hotel',
    'Canasta', 'Tráfico', 'WoW Tráfico',
    '%NoDispo', 'WoW ND', 'Banda ND',
    '%Conv', 'WoW CV', 'Banda CV', 'Bookings'
]

def write_maestra(ws, df_grp, grp_label):
    title(ws, f'{grp_label} · Maestra RND W{VOL_NUM}',
          f'Una fila por Hotel × Canasta · {PERIODO}')
    CANASTA = D.get('CANASTA', {})
    canastas = [('Global', None), ('B2C', 'B2C'), ('Opaco', 'B2B-OP'), ('Ultra Opaco', 'CUG')]
    corps_set = set(df_grp['CorpName'].unique())

    hdr_r = 4
    r = mk_hdr(ws, hdr_r, MAESTRA_COLS)

    for can_label, can_id in canastas:
        if can_id is None:
            df_can = df_grp.copy()
        else:
            can = CANASTA.get(can_id, {})
            src = can.get('p80_hotel')
            if src is None: src = can.get('p80')
            if src is None: continue
            df_can = src[src['CorpName'].isin(corps_set)].copy() if 'CorpName' in src.columns else pd.DataFrame()
        if len(df_can) == 0: continue

        for _, row in df_can.iterrows():
            nd   = sf(row.get('%NoDispo'))
            bk   = int(sf(row.get('Bookings')) or 0)
            trf  = int(sf(row.get('Trafico')) or 0)
            conv = bk / trf if trf > 0 else None
            wow_nd = sf(row.get('NoDispo_WoW_pp'))
            wow_cv = sf(row.get('ConvRate_WoW_pp') if 'ConvRate_WoW_pp' in row.index else None)
            bnd_nd = banda_nodispo(nd) if nd is not None else '—'
            bnd_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
            pais    = str(row.get('PaisDestino', row.get('Pais', '—')))
            destino = str(row.get('Destino', '—'))
            corp    = str(row.get('CorpName', '—'))
            hotel   = str(row.get('Hotel', '—'))
            for ci, val in enumerate([pais, destino, corp, hotel], 1):
                mk_cell(ws, r, ci, val, align='left')
            mk_cell(ws, r, 5, can_label)
            mk_cell(ws, r, 6, trf)
            apply_wow(ws, r, 7, sf(row.get('Trafico_WoW_pct')), invert=False)
            pct_cell(ws, r, 8, nd)
            apply_wow(ws, r, 9, wow_nd, invert=True)
            mk_cell(ws, r, 10, bnd_nd, bnd_nd, is_sev=True)
            pct_cell(ws, r, 11, conv)
            apply_wow(ws, r, 12, wow_cv, invert=False)
            mk_cell(ws, r, 13, bnd_cv, bnd_cv, is_sev=True)
            mk_cell(ws, r, 14, bk)
            r += 1

    if r > hdr_r + 1:
        add_table(ws, hdr_r, r-1, len(MAESTRA_COLS), 'Maestra')
    autofit(ws, [16, 22, 26, 40, 14, 10, 10, 10, 10, 18, 10, 10, 18, 10])

# ── Hojas de banda ────────────────────────────────────────────────────────────
BANDA_COLS = [
    'País', 'Destino', 'Corporativo', 'Hotel',
    'Tráfico', '%NoDispo', 'WoW ND', '%Conv', 'WoW CV', 'Bookings',
    'Banda ND', 'Banda CV',
]

def write_banda(ws, df_banda, sheet_title):
    title(ws, sheet_title, f'Top 500 · banda por %NoDispo · {PERIODO}')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return
    df_s = df_banda.sort_values('%NoDispo', ascending=False).head(500)
    hdr_r = 4; r = mk_hdr(ws, hdr_r, BANDA_COLS)
    for _, row in df_s.iterrows():
        nd   = sf(row.get('%NoDispo'))
        bk   = int(sf(row.get('Bookings')) or 0)
        trf  = int(sf(row.get('Trafico')) or 0)
        conv = bk / trf if trf > 0 else None
        bnd_nd = banda_nodispo(nd) if nd is not None else '—'
        bnd_cv = banda_convrate(conv, bk) if conv is not None else ('Sin Conversión' if bk == 0 else '—')
        pais    = str(row.get('PaisDestino', row.get('Pais', '—')))
        destino = str(row.get('Destino', '—'))
        corp    = str(row.get('CorpName', '—'))
        hotel   = str(row.get('Hotel', '—'))
        for ci, val in enumerate([pais, destino, corp, hotel], 1):
            mk_cell(ws, r, ci, val, align='left')
        mk_cell(ws, r, 5, trf)
        pct_cell(ws, r, 6, nd)
        apply_wow(ws, r, 7, sf(row.get('NoDispo_WoW_pp')), invert=True)
        pct_cell(ws, r, 8, conv)
        apply_wow(ws, r, 9, sf(row.get('ConvRate_WoW_pp') if 'ConvRate_WoW_pp' in row.index else None), invert=False)
        mk_cell(ws, r, 10, bk)
        mk_cell(ws, r, 11, bnd_nd, bnd_nd, is_sev=True)
        mk_cell(ws, r, 12, bnd_cv, bnd_cv, is_sev=True)
        r += 1
    if r > hdr_r + 1:
        add_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [16, 22, 26, 40, 10, 10, 10, 10, 10, 10, 16, 16])

# ── Generar un Excel por grupo de cuentas ─────────────────────────────────────
df_global = D.get('p80_hotel', pd.DataFrame())
out_dir = Path(OUTPUTS) / 'accounts'
out_dir.mkdir(parents=True, exist_ok=True)

generated = []
for grp in ACCOUNT_GROUPS:
    corps    = grp['corps']
    label    = grp['label']
    fname_id = grp['file']

    df_grp = df_global[df_global['CorpName'].isin(corps)].copy().reset_index(drop=True)
    crit, bajo, sinc = band_split(df_grp)

    wb = Workbook(); wb.remove(wb.active)
    _tc[0] = 0

    ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = RND
    write_severity(ws, df_grp, label, corps)

    ws = wb.create_sheet('Maestra');          ws.sheet_properties.tabColor = RND
    write_maestra(ws, df_grp, label)

    ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
    write_banda(ws, crit, f'{label} · Críticos W{VOL_NUM}')

    ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
    write_banda(ws, bajo, f'{label} · Bajo Rendimiento W{VOL_NUM}')

    ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
    write_banda(ws, sinc, f'{label} · Sin Conversión W{VOL_NUM}')

    fname = f'Analisis_RND_{fname_id}_W{VOL_NUM}.xlsx'
    wb.save(out_dir / fname)
    generated.append((label, fname, len(df_grp), len(crit), len(bajo), len(sinc)))

print(f'✅ Excels RND por cuenta → {out_dir}')
print(f'   {"Grupo":<25} {"Archivo":<45} {"Total":>6} {"Crit":>5} {"BR":>5} {"SC":>5}')
for label, fname, total, c, b, s in generated:
    print(f'   {label:<25} {fname:<45} {total:>6} {c:>5} {b:>5} {s:>5}')
