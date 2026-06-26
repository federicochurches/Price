"""
excel_cr_accounts.py · W26+ · Excels CR por cuenta
Genera 2 archivos:
  Analisis_CR_GlobalAccounts_WNN.xlsx
  Analisis_CR_Estrategicas_WNN.xlsx

5 hojas: Severity | Maestra | Críticos | Bajo Rendimiento | Sin Conversión
"""
import pickle, os, re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from engine import banda_eficacia, banda_convrate
from accounts_config import ACCOUNT_GROUPS

VOL_NUM = os.getenv('VOL_NUM', '21')
PERIODO = os.getenv('PERIODO', '19-25 mayo 2026')
OUTPUTS = os.getenv('OUTPUTS_DIR', '/mnt/user-data/outputs')

with open(os.getenv('PICKLE_CR', f'cr_w{VOL_NUM}_data.pkl'), 'rb') as f:
    D = pickle.load(f)

M      = D['M']
TAB_EF = D['TAB_EF_BY_CANASTA']
TAB_CV = D['TAB_CV_BY_CANASTA']
hcm    = D.get('hotel_channel_map', {})

def clean(n):
    return re.sub(r'^\(\d+\)\s*-\s*', '', str(n)).strip() if n else str(n)

hcm_clean = {clean(k): v for k, v in hcm.items()}
p80_all = D['p80_hotel'].copy()
p80_all['Hotel']   = p80_all['Hotel'].apply(clean)
p80_all['Channel'] = p80_all['Hotel'].map(hcm_clean).fillna('—')

CR_COLOR = '5C469C'

def fill(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fnt(c='000000', sz=10, bold=False, white=False):
    return Font(name='Arial', size=sz, bold=bold, color='FFFFFF' if white else c)
T  = Side(border_style='thin', color='CCCCCC')
BD = Border(left=T, right=T, top=T, bottom=T)

BFILL = {
    'Exitosa':        fill('1A6B4A'), 'Aceptable':      fill('FBBF24'),
    'Revisar':        fill('F97316'), 'Crítica':         fill('C0392B'),
    'Súper Crítica':  fill('2D2828'), 'Sin Conversión':  fill('8A8377'),
}
BFONT = {k: fnt(white=True, bold=True) for k in BFILL}

WOW_UP  = fill('EAF3DE'); WOW_UP_F  = fnt('2F6C34', bold=True)
WOW_DN  = fill('FCE8E6'); WOW_DN_F  = fnt('C0392B', bold=True)
WOW_NEU = fill('F2EEE6'); WOW_NEU_F = fnt('8A8377', bold=True)

RANGOS_EF = {
    'Exitosa': '≥ 97%', 'Aceptable': '93% – 97%', 'Revisar': '85% – 93%',
    'Crítica': '60% – 85%', 'Súper Crítica': '< 60%',
}
RANGOS_CV = {
    'Sin Conversión': 'Bookings = 0', 'Crítica': '< 0,8%',
    'Revisar': '0,8% – 1,5%', 'Aceptable': '1,5% – 2,5%', 'Exitosa': '≥ 2,5%',
}

_tc = [0]

def sf(v):
    try: f = float(v); return None if pd.isna(f) else f
    except: return None

def fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    return round(float(v), 4)

def title(ws, t, sub=''):
    ws.cell(1, 1, t).font = fnt(CR_COLOR, 13, True)
    if sub: ws.cell(2, 1, sub).font = fnt('666666')

def section_title(ws, row, text):
    ws.cell(row, 1, text).font = fnt(CR_COLOR, 11, True)

def mk_hdr(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row, c, lbl)
        cell.font = fnt(white=True, bold=True)
        cell.fill = fill(CR_COLOR)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BD
    return row + 1

def add_table(ws, hdr_row, end_row, n_cols, prefix):
    if end_row < hdr_row + 1: return
    _tc[0] += 1
    ref = f'A{hdr_row}:{get_column_letter(n_cols)}{end_row}'
    t = Table(displayName=f'T_{prefix}_{_tc[0]}', ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1', showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(t)

def mk_cell(ws, row, col, val, banda=None, is_sev=False, align='center'):
    cell = ws.cell(row, col, val)
    cell.font = Font(name='Arial', size=10)
    cell.border = BD
    cell.alignment = Alignment(horizontal=align)
    if is_sev and banda and banda in BFILL:
        cell.fill = BFILL[banda]
        cell.font = BFONT[banda]

def apply_wow(ws, row, col, val_pp, invert=False):
    cell = ws.cell(row, col)
    if val_pp is None or (isinstance(val_pp, float) and pd.isna(val_pp)):
        cell.value = '—'; cell.font = WOW_NEU_F; cell.fill = WOW_NEU
        cell.border = BD; cell.alignment = Alignment(horizontal='center'); return
    is_up = float(val_pp) >= 0
    is_good = (is_up and not invert) or (not is_up and invert)
    s = '▲' if is_up else '▼'
    cell.value = f'{s}{abs(round(float(val_pp), 2))}'.replace('.', ',')
    cell.fill = WOW_UP if is_good else WOW_DN
    cell.font = WOW_UP_F if is_good else WOW_DN_F
    cell.border = BD; cell.alignment = Alignment(horizontal='center')

def pct_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.border = BD; c.alignment = Alignment(horizontal='center')
    if val is not None: ws.cell(row, col).number_format = '0.00%'

def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def band_split(df):
    empty = pd.DataFrame()
    if df is None or len(df) == 0: return empty, empty, empty
    d = df.copy().reset_index(drop=True)
    bk   = d['Bookings'].fillna(0) if 'Bookings' in d.columns else pd.Series(0, index=d.index)
    bcol = d['Eficacia'].apply(lambda v: banda_eficacia(round(float(v), 4)) if pd.notna(v) else '—')
    crit = d[(bk > 0) & bcol.isin(['Crítica', 'Súper Crítica'])].reset_index(drop=True)
    bajo = d[(bk > 0) & bcol.isin(['Revisar', 'Aceptable'])].reset_index(drop=True)
    sinc = d[bk == 0].reset_index(drop=True)
    return crit, bajo, sinc

# ── Severity ──────────────────────────────────────────────────────────────────
def write_severity(ws, df_grp, grp_label, corps):
    title(ws, f'{grp_label} · Severity CR W{VOL_NUM}', f'{PERIODO} · {len(corps)} corporativos')

    m_curr = M.get(f'global_w{VOL_NUM}', M.get('global_w21', {}))
    m_prev = M.get(f'global_w{int(VOL_NUM)-1}', M.get('global_w20', {}))
    ef_curr = m_curr.get('eficacia', 0); ef_prev = m_prev.get('eficacia', 0)
    cv_curr = m_curr.get('conv_rate', 0); cv_prev = m_prev.get('conv_rate', 0)
    ef_wow  = (ef_curr - ef_prev) * 100 if ef_prev else None
    cv_wow  = (cv_curr - cv_prev) * 100 if cv_prev else None

    r = 4
    section_title(ws, r, 'KPI Global (referencia)'); r += 1
    r = mk_hdr(ws, r, ['Métrica', f'W{int(VOL_NUM)-1}', f'W{VOL_NUM} Global', 'WoW'])
    for label, prev, curr, wow in [('Eficacia', ef_prev, ef_curr, ef_wow),
                                    ('Conv Rate', cv_prev, cv_curr, cv_wow)]:
        ws.cell(r, 1, label).font = Font(name='Arial', size=10, bold=True); ws.cell(r, 1).border = BD
        ws.cell(r, 2, round(prev, 4) if prev else None).border = BD
        ws.cell(r, 3, round(curr, 4) if curr else None).border = BD
        ws.cell(r, 2).number_format = '0.00%'; ws.cell(r, 2).alignment = Alignment(horizontal='center')
        ws.cell(r, 3).number_format = '0.00%'; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        apply_wow(ws, r, 4, wow, invert=False); r += 1
    r += 1

    sev_ef, sev_cv = {}, {}
    for _, row in df_grp.iterrows():
        ef = fmt_pct(row.get('Eficacia'))
        cv = fmt_pct(row.get('ConvRate'))
        bk = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        if ef is not None: b = banda_eficacia(ef); sev_ef[b] = sev_ef.get(b, 0) + 1
        if cv is not None: b = banda_convrate(cv, bk); sev_cv[b] = sev_cv.get(b, 0) + 1
    total_ef = sum(sev_ef.values()) or 1
    total_cv = sum(sev_cv.values()) or 1

    SEV_COLS = ['Severity', 'Rango', 'Hoteles', '% del Total']

    section_title(ws, r, f'Severity Eficacia · {grp_label}'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Súper Crítica']:
        n = sev_ef.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_EF.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_ef, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevEf')
    r += 1

    section_title(ws, r, f'Severity Conv Rate · {grp_label}'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, SEV_COLS)
    for b in ['Exitosa', 'Aceptable', 'Revisar', 'Crítica', 'Sin Conversión']:
        n = sev_cv.get(b, 0)
        mk_cell(ws, r, 1, b, b, is_sev=True)
        mk_cell(ws, r, 2, RANGOS_CV.get(b, ''), align='left')
        ws.cell(r, 3, n).border = BD; ws.cell(r, 3).alignment = Alignment(horizontal='center')
        ws.cell(r, 4, round(n/total_cv, 4)).number_format = '0.0%'
        ws.cell(r, 4).border = BD; ws.cell(r, 4).alignment = Alignment(horizontal='center')
        r += 1
    add_table(ws, hdr_r, r-1, len(SEV_COLS), 'SevCV')
    r += 1

    # Ranking de corps
    CORP_COLS = ['Corporativo', 'Banda Ef', 'Banda CV', 'Hoteles', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']
    corp_rows = []
    for corp in corps:
        df_c = df_grp[df_grp['CorpName'] == corp]
        if len(df_c) == 0:
            corp_rows.append({'corp': corp, 'ef': None, 'cv': None, 'bk': 0,
                              'cru': 0, 'n': 0, 'wow_ef': None, 'wow_cv': None})
            continue
        cru = int(df_c['CR_Unicos'].sum()) if 'CR_Unicos' in df_c else 0
        bk  = int(df_c['Bookings'].sum()) if 'Bookings' in df_c else 0
        ef_w = (df_c['Eficacia'] * df_c['CR_Unicos']).sum() / cru if cru > 0 else None
        cv_w = (df_c['ConvRate'] * df_c['CR_Unicos']).sum() / cru if cru > 0 else None
        wow_ef = df_c['Eficacia_WoW_pp'].mean() if 'Eficacia_WoW_pp' in df_c.columns else None
        wow_cv = df_c['ConvRate_WoW_pp'].mean() if 'ConvRate_WoW_pp' in df_c.columns else None
        corp_rows.append({'corp': corp, 'ef': fmt_pct(ef_w), 'cv': fmt_pct(cv_w),
                          'bk': bk, 'cru': cru, 'n': len(df_c),
                          'wow_ef': sf(wow_ef), 'wow_cv': sf(wow_cv)})

    df_cr = pd.DataFrame(corp_rows)
    df_cr = df_cr.sort_values('ef', ascending=True, na_position='last')

    section_title(ws, r, 'Ranking Corporativos · Eficacia ASC'); r += 1
    hdr_r = r; r = mk_hdr(ws, r, CORP_COLS)
    for _, row in df_cr.iterrows():
        ef  = row.get('ef'); cv = row.get('cv')
        bk  = int(row.get('bk') or 0); cru = int(row.get('cru') or 0)
        bef = banda_eficacia(ef) if ef is not None else '—'
        bcv = banda_convrate(cv, bk) if cv is not None else '—'
        mk_cell(ws, r, 1, str(row.get('corp', '—')), align='left')
        mk_cell(ws, r, 2, bef, bef, is_sev=True)
        mk_cell(ws, r, 3, bcv, bcv, is_sev=True)
        mk_cell(ws, r, 4, int(row.get('n') or 0))
        mk_cell(ws, r, 5, cru)
        pct_cell(ws, r, 6, ef)
        apply_wow(ws, r, 7, row.get('wow_ef'), invert=False)
        pct_cell(ws, r, 8, cv)
        apply_wow(ws, r, 9, row.get('wow_cv'), invert=False)
        mk_cell(ws, r, 10, bk)
        r += 1
    add_table(ws, hdr_r, r-1, len(CORP_COLS), 'Corps')

    # Top Destinos
    DEST_COLS = ['Destino', 'Corporativo', 'Banda Ef', 'Banda CV', 'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings']
    if len(df_grp) > 0:
        r += 1
        section_title(ws, r, 'Top Destinos del grupo · Eficacia ASC'); r += 1
        dest_agg = {}
        for _, row in df_grp.iterrows():
            dest = str(row.get('Destino', '—')); corp = str(row.get('CorpName', '—'))
            ef   = fmt_pct(row.get('Eficacia')); cv = fmt_pct(row.get('ConvRate'))
            bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            key  = (dest, corp)
            if key not in dest_agg:
                dest_agg[key] = {'cru': 0, 'bk': 0, 'ef_cru': 0.0, 'cv_cru': 0.0,
                                  'wow_ef': sf(row.get('Eficacia_WoW_pp')),
                                  'wow_cv': sf(row.get('ConvRate_WoW_pp'))}
            dest_agg[key]['cru']    += cru; dest_agg[key]['bk'] += bk
            dest_agg[key]['ef_cru'] += (ef * cru) if ef else 0
            dest_agg[key]['cv_cru'] += (cv * cru) if cv else 0

        dest_rows = []
        for (dest, corp), v in dest_agg.items():
            cru = v['cru']
            ef_w = v['ef_cru'] / cru if cru > 0 else None
            cv_w = v['cv_cru'] / cru if cru > 0 else None
            dest_rows.append({'Destino': dest, 'Corp': corp, 'cru': cru, 'bk': v['bk'],
                              'ef': fmt_pct(ef_w), 'cv': fmt_pct(cv_w),
                              'wow_ef': v['wow_ef'], 'wow_cv': v['wow_cv']})
        df_dest_r = pd.DataFrame(dest_rows)
        if len(df_dest_r) > 0:
            df_dest_r = df_dest_r.sort_values('ef', ascending=True, na_position='last').head(50)
            hdr_r = r; r = mk_hdr(ws, r, DEST_COLS)
            for _, row in df_dest_r.iterrows():
                ef  = row.get('ef'); cv = row.get('cv')
                bk  = int(row.get('bk') or 0); cru = int(row.get('cru') or 0)
                bef = banda_eficacia(ef) if ef is not None else '—'
                bcv = banda_convrate(cv, bk) if cv is not None else '—'
                mk_cell(ws, r, 1, str(row.get('Destino', '—')), align='left')
                mk_cell(ws, r, 2, str(row.get('Corp', '—')), align='left')
                mk_cell(ws, r, 3, bef, bef, is_sev=True)
                mk_cell(ws, r, 4, bcv, bcv, is_sev=True)
                mk_cell(ws, r, 5, cru)
                pct_cell(ws, r, 6, ef)
                apply_wow(ws, r, 7, sf(row.get('wow_ef')), invert=False)
                pct_cell(ws, r, 8, cv)
                apply_wow(ws, r, 9, sf(row.get('wow_cv')), invert=False)
                mk_cell(ws, r, 10, bk)
                r += 1
            add_table(ws, hdr_r, r-1, len(DEST_COLS), 'Dest')

    autofit(ws, [26, 20, 16, 16, 10, 10, 10, 10, 10, 10])

# ── Maestra ───────────────────────────────────────────────────────────────────
MAESTRA_COLS = [
    'Destino', 'Corporativo', 'Hotel', 'Channel', 'Canasta',
    'CR Únicos', 'WoW CR', 'Eficacia', 'WoW Ef', 'Banda Ef',
    'Conv Rate', 'WoW CV', 'Banda CV', 'Bookings'
]

def write_maestra(ws, df_grp, grp_label):
    title(ws, f'{grp_label} · Maestra CR W{VOL_NUM}',
          f'Una fila por Hotel × Canasta · {PERIODO}')
    CANASTA = D.get('CANASTA', {})
    corps_set = set(df_grp['CorpName'].unique())
    canastas  = [('Global', None), ('B2C', 'B2C'), ('Opaco', 'B2B-OP'), ('Ultra Opaco', 'CUG')]

    hdr_r = 4; r = mk_hdr(ws, hdr_r, MAESTRA_COLS)

    for can_label, can_id in canastas:
        if can_id is None:
            df_can = df_grp.copy()
        else:
            if 'DistributionCategory' in p80_all.columns:
                df_can = p80_all[(p80_all['DistributionCategory'] == can_id) &
                                  (p80_all['CorpName'].isin(corps_set))].copy()
            else:
                can = CANASTA.get(can_id, {})
                src = can.get('p80_hotel')
                if src is None: src = can.get('p80')
                if src is None: continue
                df_can = src[src['CorpName'].isin(corps_set)].copy() if 'CorpName' in src.columns else pd.DataFrame()
        if len(df_can) == 0: continue

        for _, row in df_can.iterrows():
            ef   = fmt_pct(row.get('Eficacia'))
            cv   = fmt_pct(row.get('ConvRate'))
            bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
            cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
            bef  = banda_eficacia(ef) if ef is not None else '—'
            bcv  = banda_convrate(cv, bk) if cv is not None else '—'
            hotel   = clean(str(row.get('Hotel', '—')))
            channel = str(row.get('Channel', '—'))
            destino = str(row.get('Destino', '—'))
            corp    = str(row.get('CorpName', '—'))
            for ci, val in enumerate([destino, corp, hotel, channel, can_label], 1):
                mk_cell(ws, r, ci, val, align='left')
            mk_cell(ws, r, 6, cru)
            apply_wow(ws, r, 7, sf(row.get('CR_WoW_pct')), invert=False)
            pct_cell(ws, r, 8, ef)
            apply_wow(ws, r, 9, sf(row.get('Eficacia_WoW_pp')), invert=False)
            mk_cell(ws, r, 10, bef, bef, is_sev=True)
            pct_cell(ws, r, 11, cv)
            apply_wow(ws, r, 12, sf(row.get('ConvRate_WoW_pp')), invert=False)
            mk_cell(ws, r, 13, bcv, bcv, is_sev=True)
            mk_cell(ws, r, 14, bk)
            r += 1

    if r > hdr_r + 1:
        add_table(ws, hdr_r, r-1, len(MAESTRA_COLS), 'Maestra')
    autofit(ws, [22, 26, 40, 18, 14, 10, 10, 10, 10, 18, 10, 10, 18, 10])

# ── Banda ─────────────────────────────────────────────────────────────────────
BANDA_COLS = [
    'Destino', 'Corporativo', 'Hotel', 'Channel',
    'CR Únicos', 'Eficacia', 'WoW Ef', 'Conv Rate', 'WoW CV', 'Bookings',
    'Banda Ef', 'Banda CV',
]

def write_banda(ws, df_banda, sheet_title):
    title(ws, sheet_title, f'Top 500 · banda por Eficacia ASC · {PERIODO}')
    if df_banda is None or len(df_banda) == 0:
        ws.cell(4, 1, 'Sin datos'); return
    df_s = df_banda.sort_values('Eficacia', ascending=True).head(500)
    tab_cv_g = TAB_CV.get('global', {})
    if 'ConvRate_WoW_pp' not in df_s.columns and tab_cv_g.get('hotel') is not None:
        df_s = df_s.merge(tab_cv_g['hotel'][['Hotel', 'ConvRate_WoW_pp']], on='Hotel', how='left')
    hdr_r = 4; r = mk_hdr(ws, hdr_r, BANDA_COLS)
    for _, row in df_s.iterrows():
        ef   = fmt_pct(row.get('Eficacia')); cv = fmt_pct(row.get('ConvRate'))
        bk   = int(row.get('Bookings', 0)) if pd.notna(row.get('Bookings', 0)) else 0
        cru  = int(row.get('CR_Unicos', 0)) if pd.notna(row.get('CR_Unicos', 0)) else 0
        bef  = banda_eficacia(ef) if ef is not None else '—'
        bcv  = banda_convrate(cv, bk) if cv is not None else '—'
        hotel   = clean(str(row.get('Hotel', '—')))
        channel = str(row.get('Channel', '—'))
        destino = str(row.get('Destino', '—'))
        corp    = str(row.get('CorpName', '—'))
        for ci, val in enumerate([destino, corp, hotel, channel], 1):
            mk_cell(ws, r, ci, val, align='left')
        mk_cell(ws, r, 5, cru)
        pct_cell(ws, r, 6, ef)
        apply_wow(ws, r, 7, sf(row.get('Eficacia_WoW_pp')), invert=False)
        pct_cell(ws, r, 8, cv)
        apply_wow(ws, r, 9, sf(row.get('ConvRate_WoW_pp')), invert=False)
        mk_cell(ws, r, 10, bk)
        mk_cell(ws, r, 11, bef, bef, is_sev=True)
        mk_cell(ws, r, 12, bcv, bcv, is_sev=True)
        r += 1
    if r > hdr_r + 1:
        add_table(ws, hdr_r, r-1, len(BANDA_COLS), 'Banda')
    autofit(ws, [22, 26, 40, 18, 10, 10, 10, 10, 10, 10, 16, 16])

# ── Build ─────────────────────────────────────────────────────────────────────
out_dir = Path(OUTPUTS) / 'accounts'
out_dir.mkdir(parents=True, exist_ok=True)

generated = []
for grp in ACCOUNT_GROUPS:
    corps    = grp['corps']
    label    = grp['label']
    fname_id = grp['file']

    df_grp = p80_all[p80_all['CorpName'].isin(corps)].copy().reset_index(drop=True)
    crit, bajo, sinc = band_split(df_grp)

    wb = Workbook(); wb.remove(wb.active)
    _tc[0] = 0

    ws = wb.create_sheet('Severity');         ws.sheet_properties.tabColor = CR_COLOR
    write_severity(ws, df_grp, label, corps)

    ws = wb.create_sheet('Maestra');          ws.sheet_properties.tabColor = CR_COLOR
    write_maestra(ws, df_grp, label)

    ws = wb.create_sheet('Críticos');         ws.sheet_properties.tabColor = 'C0392B'
    write_banda(ws, crit, f'{label} · Críticos CR W{VOL_NUM}')

    ws = wb.create_sheet('Bajo Rendimiento'); ws.sheet_properties.tabColor = 'F97316'
    write_banda(ws, bajo, f'{label} · Bajo Rendimiento CR W{VOL_NUM}')

    ws = wb.create_sheet('Sin Conversión');   ws.sheet_properties.tabColor = '8A8377'
    write_banda(ws, sinc, f'{label} · Sin Conversión CR W{VOL_NUM}')

    fname = f'Analisis_CR_{fname_id}_W{VOL_NUM}.xlsx'
    wb.save(out_dir / fname)
    generated.append((label, fname, len(df_grp), len(crit), len(bajo), len(sinc)))

print(f'✅ Excels CR por cuenta → {out_dir}')
print(f'   {"Grupo":<25} {"Archivo":<45} {"Total":>6} {"Crit":>5} {"BR":>5} {"SC":>5}')
for label, fname, total, c, b, s in generated:
    print(f'   {label:<25} {fname:<45} {total:>6} {c:>5} {b:>5} {s:>5}')
